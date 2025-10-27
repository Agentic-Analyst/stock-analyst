#!/usr/bin/env python3
"""report_agent.py

Professional Financial Report Generator

This module generates comprehensive, banker-grade financial analyst reports by:
1. Collecting data from all pipeline stages (financial data, Excel models, news screening)
2. Extracting key metrics from original and adjusted financial models
3. Using LLM to synthesize all data into a professional markdown report

Data Sources:
- Financial JSON: Company info, historical financials, market data
- Original Excel Model: Base case valuation, assumptions, projections
- Screening JSON: News analysis (catalysts, risks, mitigations)
- Adjusted Excel Model: News-adjusted valuation and forecasts

Output: Professional analyst report in markdown format with:
- Executive Summary with key metrics
- Investment Thesis with catalysts and risks
- Financial Analysis with model comparison
- Valuation Summary with original vs adjusted
- Risk Assessment and recommendations
"""
from __future__ import annotations
import time
from datetime import datetime
from pathlib import Path
import pathlib
from typing import Dict, Any, List, Optional, Tuple
import json
import openpyxl
from dataclasses import dataclass
from llms.config import get_llm
from logger import StockAnalystLogger
from open_excel import excel_to_json_values_portable


@dataclass
class CompanyData:
    """Comprehensive company information from financial JSON."""
    # Basic Info
    ticker: str
    name: str
    sector: str
    industry: str
    description: str
    website: str
    employees: int
    country: str
    exchange: str
    
    # Market Data
    current_price: float
    market_cap: float
    enterprise_value: float
    shares_outstanding: float
    shares_outstanding_diluted: float
    week_52_high: float
    week_52_low: float
    
    # Valuation Multiples
    pe_trailing: float
    pe_forward: float
    peg_ratio: float
    price_to_book: float
    price_to_sales: float
    ev_to_revenue: float
    ev_to_ebitda: float
    
    # Financial Health
    total_debt: float
    total_cash: float
    net_debt: float
    debt_to_equity: float
    current_ratio: float
    quick_ratio: float
    beta: float
    
    # Profitability
    gross_margin: float
    operating_margin: float
    ebitda_margin: float
    net_margin: float
    roe: float
    roa: float
    
    # Growth
    revenue_growth: float
    earnings_growth: float
    
    # Dividends
    dividend_yield: float
    dividend_rate: float
    payout_ratio: float
    
    # Analyst Data
    target_mean_price: float
    target_high_price: float
    target_low_price: float
    recommendation: str
    num_analysts: int
    forward_eps: float


@dataclass
class ValuationMetrics:
    """Valuation metrics from Excel model."""
    # DCF Perpetual Growth
    perpetual_equity_value: float
    perpetual_enterprise_value: float
    perpetual_price_per_share: float
    
    # DCF Exit Multiple  
    exit_equity_value: float
    exit_enterprise_value: float
    exit_price_per_share: float
    
    # Blended/Average
    average_price_per_share: float
    upside_vs_market: float
    
    # Key assumptions
    wacc: float
    terminal_growth: float
    exit_multiple: float
    
    # Balance sheet items
    cash: float
    debt: float
    shares_outstanding_diluted: float


@dataclass
class FinancialProjections:
    """5-year financial projections from Excel model."""
    revenue_fy1_to_fy5: List[float]
    ebitda_fy1_to_fy5: List[float]
    fcf_fy1_to_fy5: List[float]
    
    revenue_growth_rates: List[float]
    ebitda_margins: List[float]
    fcf_margins: List[float]


def extract_company_data(financial_json_path: Path) -> CompanyData:
    """Extract comprehensive company information from financial JSON file.
    
    Args:
        financial_json_path: Path to financials_annual_modeling_latest.json
        
    Returns:
        CompanyData object with comprehensive company information
    """
    with open(financial_json_path, 'r') as f:
        data = json.load(f)
    
    company_data = data.get('company_data', {})
    basic_info = company_data.get('basic_info', {})
    market_data = company_data.get('market_data', {})
    valuation_metrics = company_data.get('valuation_metrics', {})
    capital_structure = company_data.get('capital_structure', {})
    growth_profitability = company_data.get('growth_profitability', {})
    forward_guidance = company_data.get('forward_guidance', {})
    
    return CompanyData(
        # Basic Info
        ticker=data.get('ticker', 'N/A'),
        name=basic_info.get('long_name', 'Unknown Company'),
        sector=basic_info.get('sector', 'N/A'),
        industry=basic_info.get('industry', 'N/A'),
        description=basic_info.get('business_summary', '') or basic_info.get('long_business_summary', ''),
        website=basic_info.get('website', ''),
        employees=basic_info.get('employees', 0) or basic_info.get('fullTimeEmployees', 0),
        country=basic_info.get('country', 'N/A'),
        exchange=basic_info.get('exchange', 'N/A'),
        
        # Market Data
        current_price=market_data.get('current_price', 0),
        market_cap=market_data.get('market_cap', 0),
        enterprise_value=market_data.get('enterprise_value', 0),
        shares_outstanding=market_data.get('shares_outstanding_basic', 0),
        shares_outstanding_diluted=market_data.get('shares_outstanding_diluted', 0),
        week_52_high=market_data.get('52_week_high', 0),
        week_52_low=market_data.get('52_week_low', 0),
        
        # Valuation Multiples
        pe_trailing=valuation_metrics.get('pe_ratio_trailing', 0),
        pe_forward=valuation_metrics.get('pe_ratio_forward', 0),
        peg_ratio=valuation_metrics.get('peg_ratio', 0),
        price_to_book=valuation_metrics.get('price_to_book', 0),
        price_to_sales=valuation_metrics.get('price_to_sales', 0),
        ev_to_revenue=valuation_metrics.get('enterprise_to_revenue', 0),
        ev_to_ebitda=valuation_metrics.get('enterprise_to_ebitda', 0),
        
        # Financial Health
        total_debt=capital_structure.get('total_debt', 0),
        total_cash=capital_structure.get('total_cash', 0),
        net_debt=capital_structure.get('net_debt', 0),
        debt_to_equity=capital_structure.get('debt_to_equity', 0),
        current_ratio=capital_structure.get('current_ratio', 0),
        quick_ratio=capital_structure.get('quick_ratio', 0),
        beta=capital_structure.get('beta', 0),
        
        # Profitability
        gross_margin=growth_profitability.get('gross_margins', 0),
        operating_margin=growth_profitability.get('operating_margins', 0),
        ebitda_margin=growth_profitability.get('ebitda_margins', 0),
        net_margin=growth_profitability.get('profit_margins', 0),
        roe=growth_profitability.get('return_on_equity', 0),
        roa=growth_profitability.get('return_on_assets', 0),
        
        # Growth
        revenue_growth=growth_profitability.get('revenue_growth', 0),
        earnings_growth=growth_profitability.get('earnings_growth', 0),
        
        # Dividends
        dividend_yield=market_data.get('dividend_yield', 0),
        dividend_rate=market_data.get('dividend_rate', 0),
        payout_ratio=market_data.get('payout_ratio', 0),
        
        # Analyst Data
        target_mean_price=forward_guidance.get('target_mean_price', 0),
        target_high_price=forward_guidance.get('target_high_price', 0),
        target_low_price=forward_guidance.get('target_low_price', 0),
        recommendation=forward_guidance.get('recommendation_key', 'N/A'),
        num_analysts=forward_guidance.get('number_of_analyst_opinions', 0),
        forward_eps=forward_guidance.get('forward_eps', 0)
    )


def extract_valuation_metrics(excel_path: Path) -> ValuationMetrics:
    """Extract valuation metrics from Excel file or companion JSON.
    
    First tries to load from {ticker}_valuation_metrics.json (exported by model builder).
    Falls back to using xlwings to evaluate Excel formulas directly.
    
    Args:
        excel_path: Path to financial model Excel file
        
    Returns:
        ValuationMetrics object with valuation data
    """
    # Try to load from companion JSON file first (fast path)
    json_path = excel_path.parent / f"{excel_path.stem.replace('_financial_model', '').replace('_adjusted_model', '')}_valuation_metrics.json"
    
    if json_path.exists():
        try:
            with open(json_path, 'r') as f:
                data = json.load(f)
            
            perp = data.get('perpetual_growth_dcf', {})
            exit_mult = data.get('exit_multiple_dcf', {})
            blended = data.get('blended', {})
            assumptions = data.get('assumptions', {})
            balance_sheet = data.get('balance_sheet', {})
            
            return ValuationMetrics(
                perpetual_equity_value=perp.get('equity_value'),
                perpetual_enterprise_value=perp.get('enterprise_value'),
                perpetual_price_per_share=perp.get('price_per_share'),
                exit_equity_value=exit_mult.get('equity_value'),
                exit_enterprise_value=exit_mult.get('enterprise_value'),
                exit_price_per_share=exit_mult.get('price_per_share'),
                average_price_per_share=blended.get('average_price_per_share'),
                upside_vs_market=blended.get('upside_vs_market'),
                wacc=assumptions.get('wacc'),
                terminal_growth=assumptions.get('terminal_growth'),
                exit_multiple=assumptions.get('exit_multiple'),
                cash=balance_sheet.get('cash'),
                debt=balance_sheet.get('debt'),
                shares_outstanding_diluted=balance_sheet.get('shares_outstanding_diluted', 0.0)
            )
        except Exception as e:
            # Fall through to xlwings extraction if JSON fails
            pass
    
    # Fallback: Use xlwings to evaluate Excel formulas
    try:
        # Use open_excel.py to evaluate all formulas
        data = excel_to_json_values_portable(str(excel_path))
        
        # Extract from Summary tab (all key metrics in one place)
        summary = data.get("Summary", {})
        llm_inferred = data.get("LLM_Inferred", {})
        valuation_exit = data.get("Valuation (Exit Multiple)", {})
        
        return ValuationMetrics(
            # From Summary tab - format is "(row,col)" = value
            perpetual_equity_value=summary.get("(13,2)", 0.0),  # B13
            perpetual_enterprise_value=summary.get("(17,2)", 0.0),  # B17
            perpetual_price_per_share=summary.get("(18,2)", 0.0),  # B18
            exit_equity_value=summary.get("(21,2)", 0.0),  # B21
            exit_enterprise_value=summary.get("(20,2)", 0.0),  # B20
            exit_price_per_share=summary.get("(22,2)", 0.0),  # B22
            average_price_per_share=summary.get("(26,2)", 0.0),  # B26 (average)
            upside_vs_market=summary.get("(27,2)", 0.0),  # B27
            # Assumptions from LLM_Inferred (simple values, not formulas)
            wacc=llm_inferred.get("(2,2)", 0.085),  # B2
            terminal_growth=llm_inferred.get("(3,2)", 0.025),  # B3
            exit_multiple=valuation_exit.get("(3,2)", 12.0),  # Valuation (Exit Multiple)!B3
            # Balance sheet items from Summary
            cash=summary.get("(14,2)", 0.0),  # B14
            debt=summary.get("(15,2)", 0.0),  # B15
            shares_outstanding_diluted=summary.get("(8,2)", 0.0)  # B8
        )
        
    except Exception as e:
        # Last resort: Try openpyxl for assumptions only
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            # Get assumptions from LLM_Inferred tab
            llm_inferred = wb['LLM_Inferred']
            wacc = float(llm_inferred['B2'].value or 0.085)
            terminal_growth = float(llm_inferred['B3'].value or 0.025)
            
            # Get exit multiple from Valuation (Exit Multiple) tab
            exit_multiple = 12.0
            if 'Valuation (Exit Multiple)' in wb.sheetnames:
                valuation_exit = wb['Valuation (Exit Multiple)']
                exit_mult_val = valuation_exit['B3'].value
                if exit_mult_val and isinstance(exit_mult_val, (int, float)):
                    exit_multiple = float(exit_mult_val)
            
            # Get balance sheet items from Summary tab (simpler than Historical)
            cash = 0.0
            debt = 0.0
            shares = 0.0
            
            if 'Summary' in wb.sheetnames:
                summary = wb['Summary']
                cash_val = summary['B14'].value
                debt_val = summary['B15'].value
                shares_val = summary['B8'].value
                
                if cash_val and isinstance(cash_val, (int, float)):
                    cash = float(cash_val)
                if debt_val and isinstance(debt_val, (int, float)):
                    debt = float(debt_val)
                if shares_val and isinstance(shares_val, (int, float)):
                    shares = float(shares_val)
            
            # Note: Valuation prices require formula evaluation
            return ValuationMetrics(
                perpetual_equity_value=0.0,  # Requires Excel formula evaluation
                perpetual_enterprise_value=0.0,
                perpetual_price_per_share=0.0,
                exit_equity_value=0.0,
                exit_enterprise_value=0.0,
                exit_price_per_share=0.0,
                average_price_per_share=0.0,
                upside_vs_market=0.0,
                wacc=wacc,
                terminal_growth=terminal_growth,
                exit_multiple=exit_multiple,
                cash=cash,
                debt=debt,
                shares_outstanding_diluted=shares
            )
            
        except Exception as e2:
            # Return default values on error
            return ValuationMetrics(
                perpetual_equity_value=0.0,
                perpetual_enterprise_value=0.0,
                perpetual_price_per_share=0.0,
                exit_equity_value=0.0,
                exit_enterprise_value=0.0,
                exit_price_per_share=0.0,
                average_price_per_share=0.0,
                upside_vs_market=0.0,
                wacc=0.085,
                terminal_growth=0.025,
                exit_multiple=12.0,
                cash=0.0,
                debt=0.0,
                shares_outstanding_diluted=0.0
            )


def extract_financial_projections(excel_path: Path) -> FinancialProjections:
    """Extract financial projections from Excel file.
    
    First tries to use xlwings to read evaluated formulas from Projections tab.
    Falls back to manual calculation from Raw tab and LLM_Inferred assumptions.
    
    Args:
        excel_path: Path to financial model Excel file
        
    Returns:
        FinancialProjections object with 5-year projections
    """
    try:
        # Try xlwings first - reads ACTUAL calculated values from Projections tab
        data = excel_to_json_values_portable(str(excel_path))
        
        proj = data.get("Projections", {})
        llm = data.get("LLM_Inferred", {})
        
        # Extract actual projection values
        revenues = [
            proj.get("(3,2)", 0.0),  # B3 (Revenue FY1)
            proj.get("(3,3)", 0.0),  # C3 (Revenue FY2)
            proj.get("(3,4)", 0.0),  # D3 (Revenue FY3)
            proj.get("(3,5)", 0.0),  # E3 (Revenue FY4)
            proj.get("(3,6)", 0.0),  # F3 (Revenue FY5)
        ]
        
        ebitdas = [
            proj.get("(21,2)", 0.0),  # B21 (EBITDA FY1)
            proj.get("(21,3)", 0.0),
            proj.get("(21,4)", 0.0),
            proj.get("(21,5)", 0.0),
            proj.get("(21,6)", 0.0),
        ]
        
        fcfs = [
            proj.get("(19,2)", 0.0),  # B19 (FCF FY1)
            proj.get("(19,3)", 0.0),
            proj.get("(19,4)", 0.0),
            proj.get("(19,5)", 0.0),
            proj.get("(19,6)", 0.0),
        ]
        
        # Growth rates and margins from LLM_Inferred
        growth_rates = [
            llm.get("(4,2)", 0.0),  # B4 (Revenue Growth FY1)
            llm.get("(4,3)", 0.0),
            llm.get("(4,4)", 0.0),
            llm.get("(4,5)", 0.0),
            llm.get("(4,6)", 0.0),
        ]
        
        ebitda_margins = [
            llm.get("(6,2)", 0.0),  # B6 (EBITDA Margin FY1)
            llm.get("(6,3)", 0.0),
            llm.get("(6,4)", 0.0),
            llm.get("(6,5)", 0.0),
            llm.get("(6,6)", 0.0),
        ]
        
        # Calculate FCF margins from actual values
        fcf_margins = [
            (fcfs[i] / revenues[i]) if revenues[i] > 0 else 0.0
            for i in range(5)
        ]
        
        return FinancialProjections(
            revenue_fy1_to_fy5=revenues,
            ebitda_fy1_to_fy5=ebitdas,
            fcf_fy1_to_fy5=fcfs,
            revenue_growth_rates=growth_rates,
            ebitda_margins=ebitda_margins,
            fcf_margins=fcf_margins
        )
        
    except Exception as e:
        # Fallback: Manual calculation from Raw tab and assumptions
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            # Step 1: Get base revenue from Raw tab
            raw_tab = wb['Raw']
            base_revenue = 0.0
            latest_year = None
            
            # Find latest revenue
            for row_idx in range(1, min(700, raw_tab.max_row + 1)):
                field_name = raw_tab.cell(row_idx, 2).value
                year = raw_tab.cell(row_idx, 3).value
                value = raw_tab.cell(row_idx, 4).value
                
                if field_name and value and isinstance(value, (int, float)):
                    field_str = str(field_name).lower()
                    if ('operating revenue' in field_str or 'total revenue' in field_str) and 'cost' not in field_str:
                        if latest_year is None or (year and str(year) > str(latest_year)):
                            latest_year = year
                            base_revenue = float(value)
            
            if base_revenue == 0:
                raise ValueError("Could not find revenue in Raw tab")
            
            # Step 2: Get assumptions from LLM_Inferred
            llm_tab = wb['LLM_Inferred']
            
            growth_rates = []
            for col_idx in range(2, 7):
                val = llm_tab.cell(4, col_idx).value
                growth_rates.append(float(val) if val else 0.0)
            
            ebitda_margins = []
            for col_idx in range(2, 7):
                val = llm_tab.cell(6, col_idx).value
                ebitda_margins.append(float(val) if val else 0.0)
            
            # FCF margin estimate
            fcf_margins = [margin * 0.8 for margin in ebitda_margins]
            
            # Step 3: Calculate projections
            revenues = []
            ebitdas = []
            fcfs = []
            
            current_revenue = base_revenue
            for i in range(5):
                current_revenue *= (1 + growth_rates[i])
                revenues.append(current_revenue)
                ebitdas.append(current_revenue * ebitda_margins[i])
                fcfs.append(current_revenue * fcf_margins[i])
            
            return FinancialProjections(
                revenue_fy1_to_fy5=revenues,
                ebitda_fy1_to_fy5=ebitdas,
                fcf_fy1_to_fy5=fcfs,
                revenue_growth_rates=growth_rates,
                ebitda_margins=ebitda_margins,
                fcf_margins=fcf_margins
            )
            
        except Exception as e2:
            # Return empty projections on error
            return FinancialProjections(
                revenue_fy1_to_fy5=[0.0] * 5,
                ebitda_fy1_to_fy5=[0.0] * 5,
                fcf_fy1_to_fy5=[0.0] * 5,
                revenue_growth_rates=[0.0] * 5,
                ebitda_margins=[0.0] * 5,
                fcf_margins=[0.0] * 5
            )


def load_screening_data(screening_json_path: Path) -> Dict[str, Any]:
    """Load news screening data from JSON file.
    
    Args:
        screening_json_path: Path to screening_data.json
        
    Returns:
        Dictionary with catalysts, risks, mitigations, and analysis summary
    """
    with open(screening_json_path, 'r') as f:
        return json.load(f)


def collect_report_data(analysis_path: Path, ticker: str, logger: Optional[StockAnalystLogger] = None) -> Dict[str, Any]:
    """Collect all data needed for professional financial report.
    
    This function gathers data from all pipeline stages:
    1. Financial JSON: Company info, historical data
    2. Original Excel Model: Base case valuation
    3. Screening JSON: News analysis (catalysts/risks)
    4. Adjusted Excel Model: News-adjusted valuation
    
    Args:
        analysis_path: Base path to analysis folder
        ticker: Stock ticker symbol
        logger: Optional logger instance
        
    Returns:
        Dictionary containing all collected data for report generation
    """
    if logger:
        logger.stage_start("DATA COLLECTION FOR REPORT", f"Collecting data from all pipeline outputs - {ticker}")
    
    # Paths to data sources
    financial_json = analysis_path / "financials" / "financials_annual_modeling_latest.json"
    original_model = analysis_path / "models" / f"{ticker}_financial_model.xlsx"
    screening_json = analysis_path / "screened" / "screening_data.json"
    adjusted_model = analysis_path / "models" / f"{ticker}_adjusted_model.xlsx"
    
    # Collect data
    if logger:
        logger.info("[1/4] Extracting company data from financial JSON...")
    company_data = extract_company_data(financial_json)
    if logger:
        logger.info(f"      ✅ {company_data.name} ({company_data.sector})")
    
    if logger:
        logger.info("[2/4] Extracting valuation metrics from original model...")
    original_valuation = extract_valuation_metrics(original_model)
    original_projections = extract_financial_projections(original_model)
    if logger:
        logger.info(f"      ✅ Base case: ${original_valuation.average_price_per_share:.2f}/share")
    
    if logger:
        logger.info("[3/4] Loading news screening data...")
    screening_data = load_screening_data(screening_json)
    num_catalysts = len(screening_data.get('catalysts', []))
    num_risks = len(screening_data.get('risks', []))
    if logger:
        logger.info(f"      ✅ {num_catalysts} catalysts, {num_risks} risks identified")
    
    if logger:
        logger.info("[4/4] Extracting metrics from adjusted model...")
    adjusted_valuation = extract_valuation_metrics(adjusted_model)
    adjusted_projections = extract_financial_projections(adjusted_model)
    price_change = adjusted_valuation.average_price_per_share - original_valuation.average_price_per_share
    if logger:
        logger.info(f"      ✅ Adjusted: ${adjusted_valuation.average_price_per_share:.2f}/share ({price_change:+.2f})")
        logger.stage_end("DATA COLLECTION", success=True)
    
    return {
        'company': company_data,
        'original_valuation': original_valuation,
        'original_projections': original_projections,
        'adjusted_valuation': adjusted_valuation,
        'adjusted_projections': adjusted_projections,
        'screening': screening_data,
        'analysis_path': analysis_path,
        'ticker': ticker
    }


def format_catalyst_with_evidence(catalyst: Dict, index: int) -> str:
    """Format catalyst with rich evidence, quotes, and sources.
    
    Args:
        catalyst: Catalyst dictionary from screening data
        index: Catalyst number
        
    Returns:
        Formatted markdown string with comprehensive catalyst details
    """
    output = f"\n### {index}. {catalyst.get('type', 'general').title()}: {catalyst.get('description', 'Unknown Event')}\n\n"
    
    # Key metrics banner
    conf = catalyst.get('confidence', 0.5)
    timeline = catalyst.get('timeline', 'N/A')
    impact = catalyst.get('potential_impact', 'N/A')
    
    output += f"**Timeline:** {timeline.title()} | "
    output += f"**Confidence:** {conf*100:.0f}% | "
    output += f"**Impact:** {impact}\n\n"
    
    # Supporting evidence (numbered list)
    evidence = catalyst.get('supporting_evidence', [])
    if evidence:
        output += "**Supporting Evidence:**\n"
        for i, item in enumerate(evidence, 1):
            output += f"{i}. {item}\n"
        output += "\n"
    
    # Direct quotes (block quotes with attribution)
    quotes = catalyst.get('direct_quotes', [])
    if quotes:
        output += "**Direct Evidence from News:**\n\n"
        for quote_obj in quotes[:3]:  # Show top 3 quotes
            quote_text = quote_obj.get('quote', '')
            source_title = quote_obj.get('source_article', '')
            source_url = quote_obj.get('source_url', '')
            context = quote_obj.get('context', '')
            
            if quote_text:
                output += f"> \"{quote_text}\"\n>\n"
                if context:
                    output += f"> *Context:* {context}\n>\n"
                if source_url:
                    output += f"> Source: [{source_title}]({source_url})\n\n"
                elif source_title:
                    output += f"> Source: {source_title}\n\n"
    
    # LLM reasoning (if available)
    llm_reasoning = catalyst.get('llm_reasoning') or catalyst.get('reasoning')
    if llm_reasoning and len(llm_reasoning) > 10:
        output += f"**AI Analysis:**\n{llm_reasoning}\n\n"
    
    # Source articles (bulleted list with links)
    sources = catalyst.get('source_articles', [])
    if sources:
        output += "**Related News Articles:**\n"
        for source in sources[:5]:  # Show top 5 sources
            if isinstance(source, dict):
                title = source.get('title', 'Article')
                url = source.get('url', '')
                if url:
                    output += f"- [{title}]({url})\n"
                else:
                    output += f"- {title}\n"
            else:
                output += f"- {source}\n"
        output += "\n"
    
    return output


def format_risk_with_evidence(risk: Dict, index: int) -> str:
    """Format risk with rich evidence, quotes, and sources.
    
    Args:
        risk: Risk dictionary from screening data
        index: Risk number
        
    Returns:
        Formatted markdown string with comprehensive risk details
    """
    output = f"\n### {index}. {risk.get('type', 'general').title()}: {risk.get('description', 'Unknown Risk')}\n\n"
    
    # Key metrics banner
    conf = risk.get('confidence', 0.5)
    severity = risk.get('severity', 'medium').upper()
    likelihood = risk.get('likelihood', 'medium').title()
    impact = risk.get('potential_impact', 'N/A')
    
    output += f"**Severity:** {severity} | "
    output += f"**Likelihood:** {likelihood} | "
    output += f"**Confidence:** {conf*100:.0f}%\n\n"
    output += f"**Potential Impact:** {impact}\n\n"
    
    # Supporting evidence
    evidence = risk.get('supporting_evidence', [])
    if evidence:
        output += "**Supporting Evidence:**\n"
        for i, item in enumerate(evidence, 1):
            output += f"{i}. {item}\n"
        output += "\n"
    
    # Direct quotes
    quotes = risk.get('direct_quotes', [])
    if quotes:
        output += "**Direct Evidence from News:**\n\n"
        for quote_obj in quotes[:3]:
            quote_text = quote_obj.get('quote', '')
            source_title = quote_obj.get('source_article', '')
            source_url = quote_obj.get('source_url', '')
            context = quote_obj.get('context', '')
            
            if quote_text:
                output += f"> \"{quote_text}\"\n>\n"
                if context:
                    output += f"> *Context:* {context}\n>\n"
                if source_url:
                    output += f"> Source: [{source_title}]({source_url})\n\n"
                elif source_title:
                    output += f"> Source: {source_title}\n\n"
    
    # LLM reasoning
    llm_reasoning = risk.get('llm_reasoning') or risk.get('reasoning')
    if llm_reasoning and len(llm_reasoning) > 10:
        output += f"**AI Analysis:**\n{llm_reasoning}\n\n"
    
    # Source articles
    sources = risk.get('source_articles', [])
    if sources:
        output += "**Related News Articles:**\n"
        for source in sources[:5]:
            if isinstance(source, dict):
                title = source.get('title', 'Article')
                url = source.get('url', '')
                if url:
                    output += f"- [{title}]({url})\n"
                else:
                    output += f"- {title}\n"
            else:
                output += f"- {source}\n"
        output += "\n"
    
    return output


def format_mitigation_with_evidence(mitigation: Dict, index: int) -> str:
    """Format mitigation strategy with evidence.
    
    Args:
        mitigation: Mitigation dictionary from screening data
        index: Mitigation number
        
    Returns:
        Formatted markdown string with mitigation details
    """
    output = f"\n### {index}. {mitigation.get('strategy', 'Unknown Strategy')}\n\n"
    
    # Key metrics
    conf = mitigation.get('confidence', 0.5)
    effectiveness = mitigation.get('effectiveness', 'medium').title()
    risk_addressed = mitigation.get('risk_addressed', 'N/A')
    
    output += f"**Risk Addressed:** {risk_addressed}\n"
    output += f"**Effectiveness:** {effectiveness} | "
    output += f"**Confidence:** {conf*100:.0f}%\n\n"
    
    # Company action
    company_action = mitigation.get('company_action')
    if company_action:
        output += f"**Company Action:** {company_action}\n\n"
    
    # Timeline
    timeline = mitigation.get('implementation_timeline')
    if timeline:
        output += f"**Implementation Timeline:** {timeline}\n\n"
    
    # Supporting evidence
    evidence = mitigation.get('supporting_evidence', [])
    if evidence:
        output += "**Supporting Evidence:**\n"
        for i, item in enumerate(evidence, 1):
            output += f"{i}. {item}\n"
        output += "\n"
    
    # Direct quotes
    quotes = mitigation.get('direct_quotes', [])
    if quotes:
        output += "**Evidence from News:**\n\n"
        for quote_obj in quotes[:2]:
            quote_text = quote_obj.get('quote', '')
            source_title = quote_obj.get('source_article', '')
            source_url = quote_obj.get('source_url', '')
            
            if quote_text:
                output += f"> \"{quote_text}\"\n>\n"
                if source_url:
                    output += f"> Source: [{source_title}]({source_url})\n\n"
                elif source_title:
                    output += f"> Source: {source_title}\n\n"
    
    # Source articles
    sources = mitigation.get('source_articles', [])
    if sources:
        output += "**Related Articles:**\n"
        for source in sources[:3]:
            if isinstance(source, dict):
                title = source.get('title', 'Article')
                url = source.get('url', '')
                if url:
                    output += f"- [{title}]({url})\n"
                else:
                    output += f"- {title}\n"
            else:
                output += f"- {source}\n"
        output += "\n"
    
    return output


def format_currency(value: float, decimals: int = 2) -> str:
    """Format value as currency (billions or millions)."""
    if abs(value) >= 1e9:
        return f"${value/1e9:.{decimals}f}B"
    elif abs(value) >= 1e6:
        return f"${value/1e6:.{decimals}f}M"
    else:
        return f"${value:,.{decimals}f}"


def format_percentage(value: float, decimals: int = 1) -> str:
    """Format value as percentage."""
    return f"{value*100:.{decimals}f}%"


def calculate_recommendation(upside_pct: float) -> str:
    """Calculate BUY/HOLD/SELL recommendation based on upside/downside.
    
    Thresholds:
    - BUY: upside >= +15%
    - HOLD: -10% <= upside < +15%
    - SELL: upside < -10%
    
    Args:
        upside_pct: Upside/downside as decimal (e.g., 0.15 for +15%, -0.70 for -70%)
        
    Returns:
        Recommendation string: "BUY", "HOLD", or "SELL"
    """
    if upside_pct >= 0.15:
        return "BUY"
    elif upside_pct >= -0.10:
        return "HOLD"
    else:
        return "SELL"


def generate_news_analysis_summary(screening: Dict[str, Any]) -> str:
    """Generate CONCISE news analysis summary for LLM prompt (detailed version in appendix).
    
    Args:
        screening: Screening data dictionary with catalysts, risks, mitigations
        
    Returns:
        Concise formatted markdown string with news analysis summary
    """
    catalysts = screening.get('catalysts', [])
    risks = screening.get('risks', [])
    mitigations = screening.get('mitigations', [])
    analysis_summary = screening.get('analysis_summary', {})
    
    output = ""
    
    # Overall assessment
    sentiment = analysis_summary.get('overall_sentiment', 'neutral').upper()
    confidence = analysis_summary.get('confidence_score', 0.5)
    
    output += f"**Overall Sentiment**: {sentiment} (Confidence: {confidence*100:.0f}%)\n\n"
    
    # High-level summary only
    if catalysts:
        output += f"**Positive Developments**: {len(catalysts)} catalysts identified across {len(set(c.get('type', 'general') for c in catalysts))} categories\n"
    else:
        output += "**Positive Developments**: No significant catalysts identified\n"
    
    if risks:
        output += f"**Risk Factors**: {len(risks)} risks identified - "
        risk_severities = [r.get('severity', 'medium') for r in risks]
        severity_counts = {s: risk_severities.count(s) for s in set(risk_severities)}
        output += ", ".join([f"{count} {sev.upper()}" for sev, count in sorted(severity_counts.items())])
        output += "\n"
    else:
        output += "**Risk Factors**: No significant risks identified\n"
    
    if mitigations:
        output += f"**Mitigations**: {len(mitigations)} management strategies identified\n"
    
    output += "\n*Detailed evidence with quotes and sources provided in appendix*\n"
    
    return output


def generate_source_article_index(screening: Dict[str, Any]) -> str:
    """Generate CONCISE index of source articles (full index in appendix).
    
    Args:
        screening: Screening data dictionary with catalysts, risks, mitigations
        
    Returns:
        Concise formatted markdown string with article count
    """
    # Collect all unique articles
    articles_set = set()
    
    # From catalysts
    for catalyst in screening.get('catalysts', []):
        for article in catalyst.get('source_articles', []):
            url = article.get('url', '')
            if url:
                articles_set.add(url)
    
    # From risks
    for risk in screening.get('risks', []):
        for article in risk.get('source_articles', []):
            url = article.get('url', '')
            if url:
                articles_set.add(url)
    
    # From mitigations
    for mitigation in screening.get('mitigations', []):
        for article in mitigation.get('source_articles', []):
            url = article.get('url', '')
            if url:
                articles_set.add(url)
    
    if not articles_set:
        return "*No source articles available*"
    
    return f"**Total Unique Sources Referenced**: {len(articles_set)} articles\n\n*Full source index with links provided in appendix*"


def generate_full_source_article_index(screening: Dict[str, Any]) -> str:
    """Generate complete index of all source articles for appendix.
    
    Args:
        screening: Screening data dictionary with catalysts, risks, mitigations
        
    Returns:
        Formatted markdown string with full article index
    """
    # Collect all unique articles
    articles_dict = {}  # url -> {title, date, referenced_in}
    
    # From catalysts
    for i, catalyst in enumerate(screening.get('catalysts', []), 1):
        for article in catalyst.get('source_articles', []):
            url = article.get('url', '')
            if url and url not in articles_dict:
                articles_dict[url] = {
                    'title': article.get('title', 'Untitled'),
                    'referenced_in': []
                }
            if url:
                articles_dict[url]['referenced_in'].append(f"Catalyst {i}")
    
    # From risks
    for i, risk in enumerate(screening.get('risks', []), 1):
        for article in risk.get('source_articles', []):
            url = article.get('url', '')
            if url and url not in articles_dict:
                articles_dict[url] = {
                    'title': article.get('title', 'Untitled'),
                    'referenced_in': []
                }
            if url:
                articles_dict[url]['referenced_in'].append(f"Risk {i}")
    
    # From mitigations
    for i, mitigation in enumerate(screening.get('mitigations', []), 1):
        for article in mitigation.get('source_articles', []):
            url = article.get('url', '')
            if url and url not in articles_dict:
                articles_dict[url] = {
                    'title': article.get('title', 'Untitled'),
                    'referenced_in': []
                }
            if url:
                articles_dict[url]['referenced_in'].append(f"Mitigation {i}")
    
    if not articles_dict:
        return "*No source articles available*"
    
    output = "## Complete Source Article Index\n\n"
    for idx, (url, info) in enumerate(sorted(articles_dict.items(), key=lambda x: x[1]['title']), 1):
        output += f"{idx}. [{info['title']}]({url})\n"
        output += f"   - Referenced in: {', '.join(info['referenced_in'])}\n\n"
    
    return output


def generate_professional_report(data: Dict[str, Any], logger: Optional[StockAnalystLogger] = None) -> str:
    """Generate professional financial analyst report using LLM.
    
    Takes all collected data and synthesizes it into a comprehensive,
    banker-grade investment report with sections:
    - Executive Summary
    - Investment Thesis  
    - Company Overview
    - Financial Analysis (Original + Adjusted)
    - Valuation Summary
    - Risk Assessment
    - Investment Recommendation
    
    Args:
        data: Dictionary containing all collected report data
        logger: Optional logger instance
        
    Returns:
        Markdown-formatted professional analyst report
    """
    if logger:
        logger.stage_start("REPORT GENERATION", f"Generating professional analyst report - {data['ticker']}")
    
    # Extract key data for prompt
    company = data['company']
    orig_val = data['original_valuation']
    adj_val = data['adjusted_valuation']
    orig_proj = data['original_projections']
    adj_proj = data['adjusted_projections']
    screening = data['screening']
    
    # Load prompt template from file
    prompt_template_path = Path(__file__).parent.parent / "prompts" / "professional_report_generation.md"
    with open(prompt_template_path, 'r') as f:
        prompt_template = f.read()
    
    # OPTIMIZATION: Create CONCISE summary of catalysts/risks for LLM (not full detail)
    # The full detailed evidence sections will be appended AFTER LLM generation
    catalysts_summary = ""
    for i, catalyst in enumerate(screening.get('catalysts', []), 1):
        cat_type = catalyst.get('type', 'general').title()
        desc = catalyst.get('description', 'Unknown')
        conf = catalyst.get('confidence', 0.5)
        timeline = catalyst.get('timeline', 'N/A').title()
        catalysts_summary += f"{i}. **{cat_type}** (Confidence: {conf*100:.0f}%, Timeline: {timeline}): {desc}\n"
    
    risks_summary = ""
    for i, risk in enumerate(screening.get('risks', []), 1):
        risk_type = risk.get('type', 'general').title()
        desc = risk.get('description', 'Unknown')
        severity = risk.get('severity', 'medium').upper()
        conf = risk.get('confidence', 0.5)
        risks_summary += f"{i}. **{risk_type}** [{severity} Severity] (Confidence: {conf*100:.0f}%): {desc}\n"
    
    mitigations_summary = ""
    for i, mitigation in enumerate(screening.get('mitigations', []), 1):
        strategy = mitigation.get('strategy', 'Unknown')
        effectiveness = mitigation.get('effectiveness', 'medium').title()
        mitigations_summary += f"{i}. ({effectiveness} Effectiveness): {strategy}\n"
    
    # Calculate accurate CAGRs
    base_revenue_cagr = (orig_proj.revenue_fy1_to_fy5[-1] / orig_proj.revenue_fy1_to_fy5[0]) ** (1/4) - 1
    adj_revenue_cagr = (adj_proj.revenue_fy1_to_fy5[-1] / adj_proj.revenue_fy1_to_fy5[0]) ** (1/4) - 1
    
    # Calculate deltas (adjusted - original) in basis points
    revenue_growth_deltas = [
        (adj - orig) * 100 for adj, orig in zip(adj_proj.revenue_growth_rates, orig_proj.revenue_growth_rates)
    ]
    ebitda_margin_deltas = [
        (adj - orig) * 100 for adj, orig in zip(adj_proj.ebitda_margins, orig_proj.ebitda_margins)
    ]
    fcf_margin_deltas = [
        (adj - orig) * 100 for adj, orig in zip(adj_proj.fcf_margins, orig_proj.fcf_margins)
    ]
    
    # Format price adjustor deltas
    adjustment_details = f"""**LLM-Inferred Adjustments Based on News Analysis:**

- **Revenue Growth Rates**: {', '.join([f'FY{i+1}: {d:+.0f}bps' for i, d in enumerate(revenue_growth_deltas)])}
- **EBITDA Margins**: {', '.join([f'FY{i+1}: {d:+.0f}bps' for i, d in enumerate(ebitda_margin_deltas)])}
- **FCF Margins**: {', '.join([f'FY{i+1}: {d:+.0f}bps' for i, d in enumerate(fcf_margin_deltas)])}

*These adjustments reflect the AI model's assessment of how identified catalysts and risks will impact financial performance.*
"""
    
    # Format market cap professionally
    market_cap_formatted = f"${company.market_cap / 1e12:.1f}T" if company.market_cap > 1e12 else format_currency(company.market_cap)
    
    # Calculate recommendation based on upside/downside
    recommendation = calculate_recommendation(orig_val.upside_vs_market)
    
    # Handle adjusted valuation section
    price_delta = adj_val.average_price_per_share - orig_val.average_price_per_share
    
    if orig_val.average_price_per_share > 0:
        price_delta_pct = price_delta / orig_val.average_price_per_share
        adjusted_valuation_section = f"""
### DCF Perpetual Growth Method
- Enterprise Value: {format_currency(adj_val.perpetual_enterprise_value)}
- Equity Value: {format_currency(adj_val.perpetual_equity_value)}
- **Price per Share**: ${adj_val.perpetual_price_per_share:.2f}

### DCF Exit Multiple Method
- Enterprise Value: {format_currency(adj_val.exit_enterprise_value)}
- Equity Value: {format_currency(adj_val.exit_equity_value)}
- **Price per Share**: ${adj_val.exit_price_per_share:.2f}

### Adjusted Blended Valuation
- **Adjusted Price Target**: ${adj_val.average_price_per_share:.2f}
- **Change from Base Case**: {price_delta:+.2f} ({format_percentage(price_delta_pct, 1)})
- **Implied Upside vs Current**: {format_percentage(adj_val.upside_vs_market)}
"""
    else:
        adjusted_valuation_section = "**NOTE**: Valuation prices require Excel formula evaluation. Focus on financial projection changes instead."
    
    # Fill prompt template with CONCISE data
    prompt_context = prompt_template.format(
        ticker=company.ticker,
        company_name=company.name,
        sector=company.sector,
        industry=company.industry,
        market_cap=market_cap_formatted,
        current_price=f"{company.current_price:.2f}",
        shares_outstanding=f"{company.shares_outstanding/1e6:.1f}M",
        employees=f"{company.employees:,}",
        # Original valuation
        orig_perpetual_ev=format_currency(orig_val.perpetual_enterprise_value),
        orig_perpetual_equity=format_currency(orig_val.perpetual_equity_value),
        orig_perpetual_price=f"{orig_val.perpetual_price_per_share:.2f}",
        orig_exit_ev=format_currency(orig_val.exit_enterprise_value),
        orig_exit_equity=format_currency(orig_val.exit_equity_value),
        orig_exit_price=f"{orig_val.exit_price_per_share:.2f}",
        orig_average_price=f"{orig_val.average_price_per_share:.2f}",
        orig_upside=format_percentage(orig_val.upside_vs_market),
        recommendation=recommendation,
        wacc=format_percentage(orig_val.wacc),
        terminal_growth=format_percentage(orig_val.terminal_growth),
        exit_multiple=f"{orig_val.exit_multiple:.1f}x",
        cash=format_currency(orig_val.cash),
        debt=format_currency(orig_val.debt),
        # Original projections
        orig_revenue_projections=' → '.join([format_currency(r) for r in orig_proj.revenue_fy1_to_fy5]),
        orig_revenue_cagr=format_percentage(base_revenue_cagr, 1),
        orig_revenue_growth_rates=' | '.join([format_percentage(g) for g in orig_proj.revenue_growth_rates[1:]]),
        orig_ebitda_projections=' → '.join([format_currency(e) for e in orig_proj.ebitda_fy1_to_fy5]),
        orig_ebitda_margins=' | '.join([format_percentage(m, 1) for m in orig_proj.ebitda_margins]),
        orig_fcf_projections=' → '.join([format_currency(f) for f in orig_proj.fcf_fy1_to_fy5]),
        orig_fcf_margins=' | '.join([format_percentage(m, 1) for m in orig_proj.fcf_margins]),
        # News analysis - CONCISE SUMMARY (detailed evidence appended after)
        articles_analyzed=screening.get('analysis_summary', {}).get('articles_analyzed', 0),
        analysis_period="Recent news coverage (past 30 days)",
        overall_sentiment=screening.get('analysis_summary', {}).get('overall_sentiment', 'neutral').upper(),
        key_themes=', '.join(screening.get('analysis_summary', {}).get('key_themes', ['No major themes identified'])),
        num_catalysts=len(screening.get('catalysts', [])),
        catalysts_detail=catalysts_summary if catalysts_summary else "*No significant catalysts identified*\n",
        num_risks=len(screening.get('risks', [])),
        risks_detail=risks_summary if risks_summary else "*No significant risks identified*\n",
        num_mitigations=len(screening.get('mitigations', [])),
        mitigations_detail=mitigations_summary if mitigations_summary else "*No mitigations identified*\n",
        news_analysis_summary=generate_news_analysis_summary(screening),
        source_article_index=generate_source_article_index(screening),
        # Adjusted valuation
        adjusted_valuation_section=adjusted_valuation_section,
        adjustment_details=adjustment_details,
        # Adjusted projections
        adj_revenue_projections=' → '.join([format_currency(r) for r in adj_proj.revenue_fy1_to_fy5]),
        adj_revenue_cagr=format_percentage(adj_revenue_cagr, 1),
        adj_revenue_growth_rates=' | '.join([format_percentage(g) for g in adj_proj.revenue_growth_rates[1:]]),
        adj_ebitda_projections=' → '.join([format_currency(e) for e in adj_proj.ebitda_fy1_to_fy5]),
        adj_ebitda_margins=' | '.join([format_percentage(m, 1) for m in adj_proj.ebitda_margins]),
        adj_fcf_projections=' → '.join([format_currency(f) for f in adj_proj.fcf_fy1_to_fy5]),
        adj_fcf_margins=' | '.join([format_percentage(m, 1) for m in adj_proj.fcf_margins])
    )
    
    if logger:
        logger.info("🤖 Calling LLM to generate professional report...")
    llm = get_llm()
    
    try:
        # LLM expects a list of message dicts
        messages = [
            {"role": "user", "content": prompt_context}
        ]
        response, cost = llm(messages, temperature=0.7)
        report = response.strip()
        
        if logger:
            logger.llm_call("Report Generation", cost)
            logger.info(f"✅ Report generated ({len(report):,} characters)")
        
        # POST-PROCESSING: Append detailed evidence sections AFTER LLM generation
        # This keeps the LLM prompt short while preserving all evidence in final report
        
        # Find where to insert detailed evidence (after executive summary, before metadata)
        # Look for "## News Analysis" or similar section header
        detailed_evidence = "\n\n---\n\n"
        detailed_evidence += "# APPENDIX: Detailed News Evidence & Sources\n\n"
        detailed_evidence += "*The following sections provide comprehensive evidence, direct quotes, and source links for all catalysts, risks, and mitigations identified in the news analysis.*\n\n"
        
        # Append detailed catalysts with full evidence
        if screening.get('catalysts'):
            detailed_evidence += "## Detailed Catalyst Evidence\n\n"
            for i, catalyst in enumerate(screening.get('catalysts', []), 1):
                detailed_evidence += format_catalyst_with_evidence(catalyst, i)
        
        # Append detailed risks with full evidence
        if screening.get('risks'):
            detailed_evidence += "## Detailed Risk Evidence\n\n"
            for i, risk in enumerate(screening.get('risks', []), 1):
                detailed_evidence += format_risk_with_evidence(risk, i)
        
        # Append detailed mitigations with full evidence
        if screening.get('mitigations'):
            detailed_evidence += "## Detailed Mitigation Evidence\n\n"
            for i, mitigation in enumerate(screening.get('mitigations', []), 1):
                detailed_evidence += format_mitigation_with_evidence(mitigation, i)
        
        # Append full source article index
        detailed_evidence += generate_full_source_article_index(screening)
        
        # Insert detailed evidence before metadata footer
        report += detailed_evidence
        
        # Add metadata footer
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        report += f"""

---
*Report generated on {timestamp} by AI-powered Financial Analysis System*
*Data sources: Financial statements, news analysis, DCF valuation models*
*Generation cost: ${cost:.4f}*
"""
        
        if logger:
            logger.stage_end("REPORT GENERATION", success=True, stats={
                "Report Length": f"{len(report):,} characters",
                "LLM Cost": f"${cost:.4f}"
            })
        
        return report
        
    except Exception as e:
        if logger:
            logger.error(f"❌ Error generating report: {e}")
            logger.stage_end("REPORT GENERATION", success=False)
        raise


def save_professional_report(report: str, analysis_path: Path, ticker: str, logger: Optional[StockAnalystLogger] = None) -> Path:
    """Save professional report to markdown file.
    
    Args:
        report: Generated report content
        analysis_path: Base path to analysis folder
        ticker: Stock ticker symbol
        logger: Optional logger instance
        
    Returns:
        Path to saved report file
    """
    # Create reports directory if needed
    reports_dir = analysis_path / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    # Also save as "latest" version
    latest_path = reports_dir / f"{ticker}_professional_report_latest.md"
    with open(latest_path, 'w') as f:
        f.write(report)
    
    if logger:
        logger.file_operation("Latest report saved", latest_path)
    
    return latest_path


def generate_and_save_professional_report(
    analysis_path: Path,
    ticker: str,
    logger: Optional[StockAnalystLogger] = None
) -> Tuple[str, Path]:
    """Main entry point: Collect data, generate report, save to file.
    
    This is the primary function to call from main.py. It orchestrates
    the entire report generation process:
    1. Collect data from all pipeline outputs
    2. Generate professional report using LLM
    3. Save report to markdown file
    
    Args:
        analysis_path: Path to analysis folder (e.g., data/user@email.com/AAPL/timestamp/)
        ticker: Stock ticker symbol
        logger: Optional logger instance
        
    Returns:
        Tuple of (report_content, report_file_path)
        
    Example:
        >>> report, path = generate_and_save_professional_report(
        ...     Path("data/test@example.com/AAPL/2026-1"),
        ...     "AAPL",
        ...     logger
        ... )
    """
    # Step 1: Collect all data
    data = collect_report_data(analysis_path, ticker, logger)
    
    # Step 2: Generate professional report
    report = generate_professional_report(data, logger)
    
    # Step 3: Save to file
    report_path = save_professional_report(report, analysis_path, ticker, logger)
    
    return report, report_path

