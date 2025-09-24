
#!/usr/bin/env python3
"""
financial_model_generator.py - Enhanced LLM-optional financial model generator.

This module creates comprehensive financial models (Excel/CSV) using structured
financial JSON (e.g., financials_annual_modeling_latest.json). It includes:
- Deterministic, auditable DCF to implied price (no LLM required)
- Comparable analysis (company-level with EV/EBITDA)
- Optional LLM narrative sections (safe fallback)
- Configurable WACC, terminal growth, projection years
- Robust handling of missing fields and clearer logging
- Optional explicit --data-file path (instead of directory probing)
- Clean OOP design preserved; logger supported if provided
- Auto-saves both timestamped and "latest" versions of Excel/CSV files

File Outputs:
- Timestamped: financial_model_comprehensive_TICKER_20250823_153045.xlsx
- Latest: financial_model_comprehensive_latest.xlsx (overwrites previous)
- CSV components: dcf_model_latest.csv, comparable_analysis_latest.csv, etc.

▶ Usage examples:
    python financial_model_generator.py --ticker NVDA --model dcf --data-file /path/to/financials_annual_modeling_latest.json
    python financial_model_generator.py --ticker AAPL --model comparable --years 5 --save-csv
    python financial_model_generator.py --ticker TSLA --model comprehensive --wacc 0.095 --term-growth 0.025
"""

from __future__ import annotations
import os, json, argparse, pathlib, math, re, shutil, sys, csv
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
import pandas as pd
import pandas as pd

# Import configuration
from model_config import DEFAULTS, BOUNDS, EXCEL_CONFIG, PROMPTS
from path_utils import get_latest_analysis_path
from financial_scraper import FinancialScraper

# Excel manipulation libraries
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Optional LLM integration (guarded import)
def _try_load_llm():
    try:
        import sys
        sys.path.insert(0, str(pathlib.Path(__file__).parent))
        from llms import gpt_4o_mini
        return gpt_4o_mini
    except Exception:
        return None


class FinancialModelGenerator:
    """Enhanced, auditable financial model generator for valuation analysis."""

    def __init__(self, ticker: str, base_path: pathlib.Path, email: str, data_file: Optional[str] = None):
        """Constructor sets up path references, caches, and optional LLM client.

        Parameters
        ----------
        ticker : str
            Company ticker.
        data_file : Optional[str]
            Explicit path to modeling JSON (bypasses probing) if provided.
        """
        # --- Core paths ---
        self.ticker = ticker.upper()
        self.company_dir = base_path
        self.email = email
        self.financials_dir = self.company_dir / "financials"
        self.models_dir = self.company_dir / "models"

        # --- Logging ---
        self.logger = None

        # --- Optional LLM client (narrative + parameter assist) ---
        self.llm_function =_try_load_llm()

        # --- Stats ---
        self.models_generated = 0
        self.analysis_sections = 0
        self.data_points_processed = 0

        # --- Cached data ---
        self._financial_data: Optional[Dict[str, Any]] = None
        self._company_info: Optional[Dict[str, Any]] = None

        # --- Explicit data file (if provided) ---
        self._explicit_data_file = pathlib.Path(data_file) if data_file else None

        # --- User overrides dictionary (will be merged with LLM proposals) ---
        self.overrides: Dict[str, Any] = {}

        # --- Diagnostics (warnings about missing or defaulted fields) ---
        self.diagnostics: List[str] = []

        # --- Forecast result cache ---
        self._forecast_cache: Dict[Tuple[Any, ...], Dict[str, Any]] = {}

        # --- Audit log container for LLM parameter assist ---
        self.llm_param_audit: Dict[str, Any] = {}

    # ---------- Logging helpers ----------
    def set_logger(self, logger):
        self.logger = logger

    def _log(self, level: str, message: str):
        if self.logger:
            getattr(self.logger, level, self.logger.info)(message)
        else:
            print(f"[{level.upper()}] {message}")

    # ---------- Data loading ----------
    def _load_financial_data(self) -> Dict[str, Any]:
        if self._financial_data is not None:
            return self._financial_data

        if self._explicit_data_file and self._explicit_data_file.exists():
            path = self._explicit_data_file
        else:
            # default probing
            modeling_file = self.financials_dir / "financials_annual_modeling_latest.json"
            if not modeling_file.exists():
                candidates = sorted(self.financials_dir.glob("*modeling*.json"), key=lambda f: f.stat().st_mtime, reverse=True)
                if not candidates:
                    raise FileNotFoundError(
                        f"No financial modeling data found for {self.ticker}. "
                        f"Provide --data-file or place the JSON under {self.financials_dir}."
                    )
                modeling_file = candidates[0]
            path = modeling_file

        with open(path, "r", encoding="utf-8") as f:
            self._financial_data = json.load(f)
        self._log("info", f"Loaded financial data for {self.ticker} from {path}")
        return self._financial_data

    # ---------- Safe numeric helpers ----------
    @staticmethod
    def _num(v: Any, default: Optional[float] = None) -> Optional[float]:
        try:
            if v is None: return default
            return float(v)
        except Exception:
            return default

    def _latest_from_map(self, m: Dict[str, Dict[str, Any]], keys: List[str]) -> Optional[float]:
        if not m: return None
        latest = max(m.keys())
        row = m.get(latest, {})
        for k in keys:
            if k in row and row[k] is not None:
                return self._num(row[k])
        return None

    # ---------- Metric extraction ----------
    def _extract_key_financial_metrics(self, data: Dict[str, Any]) -> Dict[str, Any]:
        metrics = {
            "company_info": {},
            "historical_financials": {},
            "valuation_inputs": {},
            "market_data": {},
            "company_data": data.get("company_data", {}),  # keep full block for WACC
        }

        cd = data.get("company_data", {})
        md = cd.get("market_data", {})
        basic = cd.get("basic_info", {})

        # Company info (no LLM classification – narrative only elsewhere)
        metrics["company_info"] = {
            "name": basic.get("long_name", self.ticker),
            "sector": basic.get("sector"),
            "industry": basic.get("industry"),
            "market_cap": md.get("market_cap"),
            "enterprise_value": md.get("enterprise_value"),
            "current_price": md.get("current_price"),
            "shares_outstanding": md.get("shares_outstanding_basic"),
        }

        fs = data.get("financial_statements", {})
        metrics["historical_financials"] = {
            "income_statement": fs.get("income_statement", {}),
            "balance_sheet": fs.get("balance_sheet", {}),
            "cash_flow": fs.get("cash_flow", {}),
        }

        vm = cd.get("valuation_metrics", {})
        cs = cd.get("capital_structure", {})
        metrics["valuation_inputs"] = {
            "beta": cs.get("beta"),
            "debt_to_equity": cs.get("debt_to_equity"),
            "current_ratio": cs.get("current_ratio"),
            "pe_ratio": vm.get("pe_ratio_trailing"),
            "price_to_book": vm.get("price_to_book"),
            "price_to_sales": vm.get("price_to_sales"),
        }

        # Market data (close-only volatility omitted here to keep core clean)
        hd = data.get("market_data", {}).get("historical_prices", {})
        metrics["market_data"] = {
            "period": hd.get("period"),
            "data_points": hd.get("data_points"),
        }

        # Count stats
        # Collect basic diagnostics for key nulls
        for field in ["market_cap","enterprise_value","current_price","shares_outstanding"]:
            if metrics["company_info"].get(field) is None:
                self.diagnostics.append(f"missing_company_info:{field}")
        self.data_points_processed = len(json.dumps(metrics))
        return metrics

    # ---------- WACC & working capital ----------
    def _infer_tax_rate(self, base_is: Dict[str, Any]) -> float:
        tr = self._num(base_is.get("Tax Rate For Calcs"))
        if tr and DEFAULTS.TAX_RATE_BOUNDS[0] < tr < DEFAULTS.TAX_RATE_BOUNDS[1]:
            return tr
        prov = self._num(base_is.get("Tax Provision"), DEFAULTS.DEFAULT_WORKING_CAPITAL_VALUE)
        pretax = self._num(base_is.get("Pretax Income"))
        if pretax and pretax != 0:
            cand = prov / pretax
            if DEFAULTS.TAX_RATE_BOUNDS[0] < cand < DEFAULTS.TAX_RATE_BOUNDS[1]:
                return float(cand)
        return DEFAULTS.DEFAULT_TAX_RATE

    def _compute_nwc(self, bs_row: Dict[str, Any]) -> float:
        ar = self._num(bs_row.get("Accounts Receivable"), DEFAULTS.DEFAULT_WORKING_CAPITAL_VALUE)
        inv = DEFAULTS.DEFAULT_WORKING_CAPITAL_VALUE
        for k in ["Inventory", "Finished Goods", "Work In Process", "Raw Materials"]:
            inv += self._num(bs_row.get(k), DEFAULTS.DEFAULT_WORKING_CAPITAL_VALUE) or DEFAULTS.DEFAULT_WORKING_CAPITAL_VALUE
        ap = self._num(bs_row.get("Accounts Payable"), DEFAULTS.DEFAULT_WORKING_CAPITAL_VALUE)
        return float(ar + inv - ap)

    def _get_wacc(self, company_data: Dict[str, Any], override_wacc: Optional[float]) -> float:
        if isinstance(override_wacc, (int, float)) and DEFAULTS.WACC_OVERRIDE_BOUNDS[0] < override_wacc < DEFAULTS.WACC_OVERRIDE_BOUNDS[1]:
            return float(override_wacc)

        cs = (company_data or {}).get("capital_structure", {}) or {}
        md = (company_data or {}).get("market_data", {}) or {}

        beta = self._num(cs.get("beta"), DEFAULTS.DEFAULT_BETA) or DEFAULTS.DEFAULT_BETA
        rf = self._num(cs.get("risk_free_rate"), DEFAULTS.DEFAULT_RISK_FREE_RATE) or DEFAULTS.DEFAULT_RISK_FREE_RATE
        erp = self._num(cs.get("equity_risk_premium"), DEFAULTS.DEFAULT_EQUITY_RISK_PREMIUM) or DEFAULTS.DEFAULT_EQUITY_RISK_PREMIUM
        ke = rf + beta * erp

        debt = self._num(cs.get("total_debt"), DEFAULTS.DEFAULT_WORKING_CAPITAL_VALUE) or DEFAULTS.DEFAULT_WORKING_CAPITAL_VALUE
        cash = self._num(cs.get("total_cash"), DEFAULTS.DEFAULT_WORKING_CAPITAL_VALUE) or DEFAULTS.DEFAULT_WORKING_CAPITAL_VALUE
        net_debt = debt - cash

        E = self._num(md.get("market_cap"), DEFAULTS.DEFAULT_WORKING_CAPITAL_VALUE) or DEFAULTS.DEFAULT_WORKING_CAPITAL_VALUE
        D = max(net_debt, 0.0) if net_debt is not None else 0.0  # if net cash, treat D~0 for weights
        V = max(E + D, 1.0)
        kd = max(rf, DEFAULTS.COST_OF_DEBT_FLOOR)  # conservative floor
        tax = DEFAULTS.DEFAULT_CORPORATE_TAX_RATE
        wacc = (E / V) * ke + (D / V) * kd * (1 - tax)
        self._log("info", f"Computed WACC: {wacc:.4f} (Ke={ke:.4f}, Kd={kd:.4f}, E={E:.2f}, D={D:.2f})")
        return float(max(DEFAULTS.WACC_MIN_FLOOR, min(wacc, DEFAULTS.WACC_MAX_CEILING)))

    # ---------- Strategy selection helpers ----------
    def _select_strategy(self, requested: Optional[str], metrics: Dict[str, Any]):
        from forecast_strategies import STRATEGIES, get_strategy_by_name
        def _normalize(name: Optional[str]) -> Optional[str]:
            if not name:
                return None
            n = str(name).strip().lower()
            mapping = {
                # Generic
                'generic': 'generic_dcf', 'generic_dcf': 'generic_dcf',
                # SaaS / Software
                'saas': 'saas_dcf', 'saas_dcf': 'saas_dcf', 'software': 'saas_dcf',
                # REIT
                'reit': 'reit_dcf', 'real estate': 'reit_dcf', 'reit_dcf': 'reit_dcf',
                # Bank / Financials
                'bank': 'bank_excess_returns', 'residual': 'bank_excess_returns', 'excess returns': 'bank_excess_returns', 'bank_excess_returns': 'bank_excess_returns',
                # Utility
                'utility': 'utility_dcf', 'utilities': 'utility_dcf', 'utility_dcf': 'utility_dcf',
                # Energy
                'energy': 'energy_nav_dcf', 'oil': 'energy_nav_dcf', 'gas': 'energy_nav_dcf', 'o&g': 'energy_nav_dcf', 'energy_nav_dcf': 'energy_nav_dcf',
                # Hardware / Semis
                'hardware': 'hardware_dcf', 'semiconductor': 'hardware_dcf', 'semiconductors': 'hardware_dcf', 'chip': 'hardware_dcf', 'hardware_dcf': 'hardware_dcf',
                # Marketplace / Platform / E-comm
                'marketplace': 'marketplace_dcf', 'ecommerce': 'marketplace_dcf', 'e-commerce': 'marketplace_dcf', 'platform': 'marketplace_dcf', 'marketplace_dcf': 'marketplace_dcf',
                # Telecom
                'telecom': 'telecom_dcf', 'telecommunications': 'telecom_dcf', 'wireless': 'telecom_dcf', 'cable': 'telecom_dcf', 'telecom_dcf': 'telecom_dcf',
            }
            if n in mapping:
                return mapping[n]
            # fuzzy: contains registry name
            names = [s.name for s in STRATEGIES]
            for nm in names:
                if n in nm:
                    return nm
            return None

        if requested:
            norm = _normalize(requested)
            strat = get_strategy_by_name(requested) or (get_strategy_by_name(norm) if norm else None)
            if strat:
                self._log("info", f"Using requested strategy: {strat.name}")
                return strat
            self._log("warning", f"Requested strategy '{requested}' not found. Falling back to auto selection.")
        sector = metrics.get("company_info", {}).get("sector")
        industry = metrics.get("company_info", {}).get("industry")
        for strat in STRATEGIES:
            if strat.name != "generic_dcf" and strat.applies_to(sector, industry):
                self._log("info", f"Auto-selected strategy '{strat.name}' for sector='{sector}' industry='{industry}'")
                return strat
        # fallback (last one is generic in registry)
        generic = [s for s in STRATEGIES if s.name == "generic_dcf"]
        chosen = generic[0] if generic else STRATEGIES[-1]
        self._log("info", f"Fallback strategy '{chosen.name}' selected.")
        return chosen

    # ---------- Comparable analysis ----------
    def _create_comparable_analysis_dataframe(self, metrics: Dict[str, Any]) -> pd.DataFrame:
        cd = metrics.get("company_data", {}) or {}
        basic = (cd.get("basic_info") or {})
        md = (cd.get("market_data") or {})
        vm = (cd.get("valuation_metrics") or {})
        cs = (cd.get("capital_structure") or {})

        cap = self._num(md.get("market_cap"))
        enterprise_val = self._num(md.get("enterprise_value"))
        if enterprise_val is None:
            debt = self._num(cs.get("total_debt"), 0.0) or 0.0
            cash = self._num(cs.get("total_cash"), 0.0) or 0.0
            enterprise_val = (cap or 0.0) + debt - cash

        is_map = metrics["historical_financials"].get("income_statement", {})
        latest = max(is_map.keys()) if is_map else None
        ebitda = self._num(is_map.get(latest, {}).get("EBITDA") if latest else None) or \
                 self._num(is_map.get(latest, {}).get("Normalized EBITDA") if latest else None) or 0.0
        ev_ebitda = (enterprise_val / ebitda) if (enterprise_val and ebitda) else None

        df = pd.DataFrame([{
            "Company": basic.get("long_name") or metrics.get("company_info", {}).get("name", self.ticker),
            "Ticker": self.ticker,
            "Market Cap ($M)": (cap / 1e6) if cap else None,
            "Enterprise Value ($M)": (enterprise_val / 1e6) if enterprise_val else None,
            "P/E (trailing)": self._num(vm.get("pe_ratio_trailing")),
            "EV/EBITDA": ev_ebitda,
            "P/S": self._num(vm.get("price_to_sales")),
            "P/B": self._num(vm.get("price_to_book")),
            "Debt/Equity": self._num(cs.get("debt_to_equity")),
        }])
        return df

    # Peer comparables (multi-row) if peer tickers supplied
    def _create_peer_comps(self, base_metrics: Dict[str, Any], peers: List[str]) -> Optional[pd.DataFrame]:
        rows = []
        # include base first
        rows.append(self._create_comparable_analysis_dataframe(base_metrics).iloc[0].to_dict())
        
        successful_peers = 0
        for pt in peers:
            try:
                self._log("info", f"Scraping fresh financial data for peer: {pt}")
                
                # Create a temporary directory for peer data scraping
                peer_base_path = self.company_dir / f"temp_peer_{pt.upper()}"
                
                # Initialize financial scraper for the peer
                peer_scraper = FinancialScraper(pt.upper(), base_path=peer_base_path)
                if self.logger:
                    peer_scraper.set_logger(self.logger)
                
                # Scrape fresh comprehensive financial modeling data
                peer_data = peer_scraper.scrape_financial_modeling_data(annual=True)
                
                if not peer_data or not peer_data.get("company_data"):
                    self._log("warning", f"Failed to scrape data for peer {pt}; skipping")
                    continue

                # Extract metrics from the fresh data
                pmetrics = self._extract_key_financial_metrics(peer_data)
                
                # Validate that we have essential data for comparison
                if not pmetrics.get("company_info", {}).get("market_cap"):
                    self._log("warning", f"Peer {pt} missing essential market data; skipping")
                    continue
                
                # Adjust ticker context temporarily
                orig = self.ticker
                self.ticker = pt.upper()
                prow = self._create_comparable_analysis_dataframe(pmetrics).iloc[0].to_dict()
                self.ticker = orig
                rows.append(prow)
                successful_peers += 1
                
                self._log("info", f"Successfully added peer {pt} to comparison")
                file_path = peer_scraper.save_financial_data(peer_data, annual=True, statements_scraped=["modeling"])
                self._log("info", f"💾 Financial data saved to: {file_path}")

            except Exception as e:
                self._log("warning", f"Peer {pt} failed during fresh data scraping: {e}")
        
        if successful_peers == 0:
            self._log("warning", "No peer data could be successfully scraped")
            return None
        
        self._log("info", f"Peer comparison table created with {successful_peers} peer(s)")
        return pd.DataFrame(rows)

    # Sensitivities
    def _generate_sensitivities(self, strategy, metrics: Dict[str, Any], projection_years: int,
                                base_wacc: float, base_term_growth: float, term_growth: float,
                                override_wacc: Optional[float]) -> Dict[str, pd.DataFrame]:
        sens: Dict[str, pd.DataFrame] = {}
        try:
            # WACC vs Terminal Growth grid
            wacc_points = sorted(set([max(0.04, round(base_wacc - 0.02,4)), max(0.04, round(base_wacc - 0.01,4)), round(base_wacc,4), min(0.15, round(base_wacc + 0.01,4)), min(0.15, round(base_wacc + 0.02,4))]))
            tg_points = sorted(set([max(0.0, round(base_term_growth - 0.01,4)), round(base_term_growth,4), min(0.05, round(base_term_growth + 0.01,4))]))
            grid = []
            for w in wacc_points:
                row = {"WACC": w}
                for g in tg_points:
                    out = strategy.forecast(self, metrics, projection_years, g, w)
                    price = out.get("valuation_summary", {}).get("Implied Price") or out.get("valuation_summary", {}).get("Implied Price (blended)")
                    row[f"g={g:.2%}"] = price
                grid.append(row)
            sens["sensitivity_wacc_term"] = pd.DataFrame(grid)

            # Growth vs Margin (vary first-year growth delta & margin uplift)
            first_year_growth = 0.0
            # attempt to infer from strategy default by running once (base already ran)
            first_run = strategy.forecast(self, metrics, projection_years, term_growth, override_wacc)
            # not storing growth sequence; we'll assume ~ first revenue growth from DCF model
            df = first_run.get("dcf_model")
            if isinstance(df, pd.DataFrame) and "Revenue" in df.columns and len(df) >= 2:
                try:
                    r0 = float(df.iloc[0]["Revenue"])
                    r1 = float(df.iloc[1]["Revenue"])
                    if r0:
                        first_year_growth = (r1 / r0) - 1
                except Exception:
                    pass
            growth_deltas = [-0.02, 0.0, 0.02]
            margin_uplifts = [-0.05, 0.0, 0.05]
            mgrid = []
            for gd in growth_deltas:
                row = {"ΔGrowth": gd}
                for mu in margin_uplifts:
                    # Build a transient override set (no mutation of self.overrides)
                    temp_overrides = dict(self.overrides)
                    temp_overrides["first_year_growth"] = max(0.0, first_year_growth + gd)
                    temp_overrides["margin_uplift"] = mu
                    # Monkey-pass via attribute (strategy reads generator.overrides); swap briefly
                    original = self.overrides
                    self.overrides = temp_overrides
                    out = strategy.forecast(self, metrics, projection_years, term_growth, override_wacc)
                    price = out.get("valuation_summary", {}).get("Implied Price") or out.get("valuation_summary", {}).get("Implied Price (blended)")
                    row[f"MarginΔ={mu:.0%}"] = price
                    self.overrides = original
                mgrid.append(row)
            sens["sensitivity_growth_margin"] = pd.DataFrame(mgrid)
        except Exception as e:
            self._log("warning", f"Sensitivity generation failed: {e}")
        return sens

    def _validation_summary(self, strategy_name: str, valuation_summary: Dict[str, Any]) -> Dict[str, Any]:
        missing = [k for k,v in valuation_summary.items() if v is None]
        return {
            "strategy": strategy_name,
            "overrides": self.overrides,
            "missing_valuation_fields": missing,
            "diagnostics": self.diagnostics,
            "models_generated": self.models_generated + 1,
        }

    # ---------- LLM narrative (optional) ----------
    def _generate_llm_financial_analysis(self, metrics: Dict[str, Any], model_type: str) -> Dict[str, str]:
        if not self.llm_function:
            self._log("info", "LLM disabled or unavailable; skipping narrative analysis.")
            self.analysis_sections = 0  # Set explicitly when LLM disabled
            return {}
        try:
            # Load financial narrative prompt from file
            prompt_path = pathlib.Path(PROMPTS.FINANCIAL_NARRATIVE)
            with open(prompt_path, 'r', encoding='utf-8') as f:
                prompt_template = f.read()
            
            # Format the prompt with ticker
            prompt = prompt_template.format(ticker=self.ticker)
            
            msgs = [
                {"role": "system", "content": "You are a senior financial analyst producing concise research-ready notes."},
                {"role": "user", "content": prompt},
            ]
            resp = self.llm_function(msgs, temperature=DEFAULTS.TEMP_NARRATIVE_ANALYSIS)
            # Support either string or (string, cost)
            if isinstance(resp, tuple):
                content, _ = resp
            else:
                content = resp
            # Simple section split
            sections = {}
            current = "executive_summary"
            buf = []
            for line in str(content).splitlines():
                up = line.upper()
                if "EXECUTIVE" in up: 
                    if buf: sections[current] = "\n".join(buf).strip(); buf=[]
                    current="executive_summary"
                elif "FINANCIAL PROJECTIONS" in up:
                    if buf: sections[current] = "\n".join(buf).strip(); buf=[]
                    current="financial_projections"
                elif "VALUATION" in up:
                    if buf: sections[current] = "\n".join(buf).strip(); buf=[]
                    current="valuation_analysis"
                elif "SENSITIVIT" in up:
                    if buf: sections[current] = "\n".join(buf).strip(); buf=[]
                    current="sensitivity_analysis"
                elif "INVESTMENT" in up:
                    if buf: sections[current] = "\n".join(buf).strip(); buf=[]
                    current="investment_recommendation"
                else:
                    buf.append(line)
            if buf:
                sections[current] = "\n".join(buf).strip()
            self.analysis_sections = len(sections)
            return sections
        except Exception as e:
            self._log("warning", f"LLM analysis failed: {e}")
            return {}

    # ---------- Public API ----------
    def generate_financial_model(self, model_type: str = "comprehensive", projection_years: int = 5,
                                 term_growth: Optional[float] = None, override_wacc: Optional[float] = None,
                                 strategy: Optional[str] = None, peers: Optional[List[str]] = None,
                                 generate_sensitivities: bool = False, lean: bool = False) -> Dict[str, Any]:
        data = self._load_financial_data()
        metrics = self._extract_key_financial_metrics(data)
        
        # --- AGENTIC PARAMETER CALCULATION ---
        # Instead of using defaults, let LLM calculate optimal parameters based on company analysis
        if self.llm_function and not self.overrides:  # Only if no manual overrides provided
            self._log("info", "🤖 Generating agentic parameter recommendations using LLM analysis...")
            
            # Create a baseline model for context (using minimal defaults)
            temp_term_growth = term_growth or 0.1  # Temporary for baseline
            baseline_model = {"valuation_summary": {"WACC": 0.1}}  # Minimal baseline
            
            try:
                agentic_params = self.propose_llm_parameter_overrides(
                    baseline_model, projection_years, temp_term_growth, strategy
                )
                
                if agentic_params:
                    # Apply agentic parameters as overrides
                    for param, value in agentic_params.items():
                        if param not in ['confidence_level', 'analysis_summary', 'reasoning']:
                            self.overrides[param] = value
                            
                    self._log("info", f"Applied agentic parameters: {list(agentic_params.keys())}")
                    
                    # Use agentic terminal growth if provided and not manually set
                    if term_growth is None and 'terminal_growth' in agentic_params:
                        term_growth = agentic_params['terminal_growth']
                        self._log("info", f"Using agentic terminal growth: {term_growth:.3f}")
                        
                    # Use agentic WACC if provided and not manually set
                    if override_wacc is None and 'wacc' in agentic_params:
                        override_wacc = agentic_params['wacc']
                        self._log("info", f"Using agentic WACC: {override_wacc:.3f}")
                        
                else:
                    self._log("warning", "Agentic parameter generation returned no parameters")
                    
            except Exception as e:
                self._log("error", f"Agentic parameter generation failed: {e}")
        
        # --- Terminal growth auto-inference (fallback if not set by agentic system) ---
        if term_growth is None:
            inferred_tg = None; tg_conf = None; tg_reason = None
            if self.llm_function:
                try:
                    prompt_path = pathlib.Path(PROMPTS.PARAMETER_INFERENCE)
                    with open(prompt_path,'r',encoding='utf-8') as f:
                        param_prompt = f.read()
                    fin = metrics.get('historical_financials', {})
                    def tail_map(m: Dict[str, Any]):
                        if not isinstance(m, dict): return m
                        keys = sorted(m.keys())[-3:]
                        return {k:m[k] for k in keys}
                    ctx_payload = {
                        'ticker': self.ticker,
                        'projection_years': projection_years,
                        'available_missing_params': ['terminal_growth'],
                        'income_statement_tail': tail_map(fin.get('income_statement', {})),
                        'balance_sheet_tail': tail_map(fin.get('balance_sheet', {})),
                        'cash_flow_tail': tail_map(fin.get('cash_flow', {})),
                        'company_info': metrics.get('company_info'),
                    }
                    llm_resp = self.llm_function([
                        {"role": "system", "content": param_prompt},
                        {"role": "user", "content": json.dumps(ctx_payload)}
                    ], temperature=DEFAULTS.TEMP_TERMINAL_GROWTH)
                    if isinstance(llm_resp, tuple): llm_resp = llm_resp[0]
                    txt = str(llm_resp).strip()
                    mm = re.search(r"\{[\s\S]+\}", txt)
                    if mm:
                        inferred_json = json.loads(mm.group(0))
                        inferred_block = inferred_json.get('inferred', {}) or {}
                        conf_block = inferred_json.get('confidence', {}) or {}
                        rationale_block = inferred_json.get('rationale', {}) or {}
                        if 'terminal_growth' in inferred_block and isinstance(inferred_block['terminal_growth'], (int,float)):
                            inferred_tg = float(inferred_block['terminal_growth'])
                            tg_conf = conf_block.get('terminal_growth', 1.0)
                            tg_reason = rationale_block.get('terminal_growth','')
                except Exception as e:
                    self._log('warning', f"LLM terminal growth inference failed: {e}")
            if inferred_tg is not None and (tg_conf is None or tg_conf >= DEFAULTS.LLM_CONFIDENCE_THRESHOLD):
                term_growth = max(0.0, min(DEFAULTS.MAX_TERMINAL_GROWTH_LLM, inferred_tg))
                self.llm_param_audit.setdefault('parameter_inference', {}).setdefault('applied', {})['terminal_growth'] = {
                    'value': term_growth,
                    'reason': tg_reason or 'llm_inferred',
                    'confidence': tg_conf
                }
                self.llm_param_audit.setdefault('parameter_inference_meta', {})['confidence_threshold'] = DEFAULTS.LLM_CONFIDENCE_THRESHOLD
            else:
                # Deterministic fallback using recent revenue CAGR * 0.3, clipped 0–0.03
                try:
                    is_map = metrics.get('historical_financials', {}).get('income_statement', {}) or {}
                    years_sorted = sorted(is_map.keys())
                    rev_vals = []
                    for y in years_sorted[-3:]:
                        row = is_map.get(y, {}) or {}
                        rv = row.get('Total Revenue') or row.get('Operating Revenue')
                        if rv is not None:
                            rev_vals.append(float(rv))
                    fallback = 0.02
                    if len(rev_vals) >= 2 and rev_vals[0] > 0 and rev_vals[-1] > 0:
                        cagr = (rev_vals[-1]/rev_vals[0]) ** (1/max(len(rev_vals)-1,1)) - 1
                        fallback = max(0.0, min(DEFAULTS.MAX_TERMINAL_GROWTH_DETERMINISTIC, cagr * DEFAULTS.TERMINAL_GROWTH_CAGR_SCALE))
                    term_growth = fallback
                except Exception:
                    term_growth = DEFAULTS.DEFAULT_TERMINAL_GROWTH_FALLBACK
                self.llm_param_audit.setdefault('parameter_inference', {}).setdefault('applied', {})['terminal_growth'] = {
                    'value': term_growth,
                    'reason': 'deterministic_fallback',
                    'confidence': 0.0
                }
        wacc_str = 'auto' if override_wacc is None else f"{override_wacc:.3f}"
        tg_str = 'n/a' if term_growth is None else f"{term_growth:.3f}"
        self._log("info", f"Generating {model_type} model for {self.ticker} ("\
                          f"{projection_years}y, g={tg_str}, wacc={wacc_str})")

        components: Dict[str, Any] = {}
        valuation_summary: Dict[str, Any] = {}

        chosen_strategy_name = None
        strat = None
        if model_type in ("dcf", "comprehensive"):
            # LLM assisted strategy selection if no explicit strategy passed and LLM available
            chosen_strategy_text = None
            if strategy is None and self.llm_function:
                try:
                    prompt_path = pathlib.Path(PROMPTS.STRATEGY_SELECTION)
                    context = {
                        "company_info": metrics.get('company_info'),
                        "quick_hist": {
                            "rev_years": len(metrics.get('historical_financials', {}).get('income_statement', {}) or {}),
                        },
                        "available_strategies": []
                    }
                    # import strategies list names
                    from forecast_strategies import STRATEGIES
                    context['available_strategies'] = [s.name for s in STRATEGIES]
                    with open(prompt_path, 'r', encoding='utf-8') as f:
                        base_prompt = f.read()
                    llm_resp = self.llm_function([
                        {"role": "system", "content": base_prompt},
                        {"role": "user", "content": json.dumps(context)}
                    ], temperature=DEFAULTS.TEMP_STRATEGY_SELECTION)
                    if isinstance(llm_resp, tuple): llm_resp = llm_resp[0]
                    txt = str(llm_resp).strip()
                    m = re.search(r"\{[\s\S]+\}", txt)
                    if m:
                        parsed = json.loads(m.group(0))
                        chosen_strategy_text = parsed.get('strategy')
                        # Allow LLM to suggest a projection horizon
                        try:
                            pj = parsed.get('projection_years')
                            if isinstance(pj, int) and 3 <= pj <= 10:
                                self._log('info', f"LLM suggested projection_years={pj}; applying")
                                projection_years = pj
                        except Exception:
                            pass
                        self.llm_param_audit.setdefault('strategy_selection', parsed)
                except Exception as e:
                    self._log('warning', f"LLM strategy selection failed: {e}; falling back to rule-based")
            strat = self._select_strategy(chosen_strategy_text or strategy, metrics)
            chosen_strategy_name = strat.name
            # Caching key
            hashable_overrides = {}
            for k, v in self.overrides.items():
                hashable_overrides[k] = str(v) if isinstance(v, list) else v
            ov_hash = hash(tuple(sorted(hashable_overrides.items())))
            cache_key = (strat.name, projection_years, round(term_growth,6), override_wacc if override_wacc is None else round(override_wacc,6), ov_hash)
            if cache_key in self._forecast_cache:
                strat_outputs = self._forecast_cache[cache_key]
                self._log("info", f"Cache hit for strategy {strat.name}")
            else:
                # Before first forecast, perform LLM parameter inference for any unset override fields
                if self.llm_function:
                    try:
                        # Determine which parameters are unset (not provided by user or prior LLM override)
                        candidate_params = [
                            'first_year_growth','margin_target','margin_ramp','capex_rate','da_rate',
                            'nwc_method','nwc_ratio','margin_curve','growth_sequence'
                        ]  # terminal_growth handled earlier to ensure caching key stability
                        missing = [p for p in candidate_params if p not in self.overrides]
                        if missing:
                            prompt_path = pathlib.Path(PROMPTS.PARAMETER_INFERENCE)
                            with open(prompt_path,'r',encoding='utf-8') as f:
                                param_prompt = f.read()
                            # Build compact context for inference (recent IS/BS rows)
                            fin = metrics.get('historical_financials', {})
                            # Limit to last 3 periods per statement to keep prompt size manageable
                            def tail_map(m: Dict[str, Any]):
                                if not isinstance(m, dict): return m
                                keys = sorted(m.keys())[-3:]
                                return {k:m[k] for k in keys}
                            ctx_payload = {
                                'ticker': self.ticker,
                                'projection_years': projection_years,
                                'term_growth': term_growth,
                                'available_missing_params': missing,
                                'income_statement_tail': tail_map(fin.get('income_statement', {})),
                                'balance_sheet_tail': tail_map(fin.get('balance_sheet', {})),
                                'cash_flow_tail': tail_map(fin.get('cash_flow', {})),
                                'company_info': metrics.get('company_info'),
                            }
                            llm_resp = self.llm_function([
                                {"role": "system", "content": param_prompt},
                                {"role": "user", "content": json.dumps(ctx_payload)}
                            ], temperature=DEFAULTS.TEMP_PARAMETER_INFERENCE)
                            if isinstance(llm_resp, tuple): llm_resp = llm_resp[0]
                            txt = str(llm_resp).strip()
                            mm = re.search(r"\{[\s\S]+\}", txt)
                            if mm:
                                inferred_json = json.loads(mm.group(0))
                                inferred = inferred_json.get('inferred', {}) or {}
                                rationale = inferred_json.get('rationale', {}) or {}
                                conf = inferred_json.get('confidence', {}) or {}
                                applied = {}
                                skipped_low_conf = []
                                confidence_threshold = DEFAULTS.LLM_CONFIDENCE_THRESHOLD
                                self.llm_param_audit.setdefault('parameter_inference_meta', {})['confidence_threshold'] = confidence_threshold
                                bounds = BOUNDS.get_bounds_dict()
                                for k,v in inferred.items():
                                    cval = conf.get(k, 1.0)
                                    if cval < confidence_threshold:
                                        skipped_low_conf.append({'param': k, 'reason': 'low_confidence', 'confidence': cval})
                                        continue
                                    if k == 'growth_sequence' and isinstance(v, list) and k not in self.overrides:
                                        seq = [float(x) for x in v if isinstance(x,(int,float))]
                                        if seq:
                                            # Enforce length; if shorter, extend with last value; if longer, truncate
                                            if len(seq) < projection_years:
                                                seq += [seq[-1]] * (projection_years - len(seq))
                                            if len(seq) > projection_years:
                                                seq = seq[:projection_years]
                                            # Clip bounds per element
                                            seq = [max(0.0, min(0.60, x)) for x in seq]
                                            self.overrides[k] = seq
                                            applied[k] = {'value': seq, 'reason': rationale.get(k,''), 'confidence': cval}
                                    elif k == 'margin_curve' and isinstance(v, list) and k not in self.overrides:
                                        curve = [float(x) for x in v if isinstance(x,(int,float))]
                                        if curve:
                                            if len(curve) < projection_years:
                                                curve += [curve[-1]] * (projection_years - len(curve))
                                            if len(curve) > projection_years:
                                                curve = curve[:projection_years]
                                            curve = [max(0.0, min(0.90, x)) for x in curve]
                                            self.overrides[k] = curve
                                            applied[k] = {'value': curve, 'reason': rationale.get(k,''), 'confidence': cval}
                                    elif k in candidate_params and k not in self.overrides and isinstance(v,(int,float)):
                                        lo,hi = bounds.get(k,(None,None))
                                        fv = float(v)
                                        if lo is not None:
                                            fv = max(lo, min(hi, fv))
                                        self.overrides[k] = fv
                                        applied[k] = {'value': fv, 'reason': rationale.get(k,''), 'confidence': cval}
                                    elif k=='nwc_method' and k not in self.overrides and v in ('ratio','delta2pct'):
                                        if cval >= confidence_threshold:
                                            self.overrides[k] = v
                                            applied[k] = {'value': v, 'reason': rationale.get(k,''), 'confidence': cval}
                                        else:
                                            skipped_low_conf.append({'param': k, 'reason': 'low_confidence', 'confidence': cval})
                                if applied:
                                    self.llm_param_audit.setdefault('parameter_inference', {})['applied'] = applied
                                    self.llm_param_audit.setdefault('parameter_inference', {})['raw'] = txt[:1500]
                                if skipped_low_conf:
                                    self.llm_param_audit.setdefault('parameter_inference', {})['skipped_low_confidence'] = skipped_low_conf
                    except Exception as e:
                        self._log('warning', f"LLM parameter inference failed: {e}")
                strat_outputs = strat.forecast(self, metrics, projection_years, term_growth, override_wacc)
                self._forecast_cache[cache_key] = strat_outputs
            dcf_df = strat_outputs.get("dcf_model")
            valuation_summary = strat_outputs.get("valuation_summary", {})
            valuation_summary["Strategy"] = strat_outputs.get("strategy_name")
            components["dcf_model"] = dcf_df
            components["valuation_summary"] = valuation_summary
            # Extra components (e.g., FFO/AFFO)
            for k,v in (strat_outputs.get("extra_components") or {}).items():
                components[k] = v
            if (not lean) and generate_sensitivities and strat and valuation_summary.get("WACC"):
                sens = self._generate_sensitivities(strat, metrics, projection_years,
                                                    valuation_summary.get("WACC"), term_growth, term_growth, override_wacc)
                for k,v in sens.items():
                    components[k] = v

        if (not lean) and model_type in ("comparable", "comprehensive"):
            comps_df = self._create_comparable_analysis_dataframe(metrics)
            components["comparable_analysis"] = comps_df
            if peers:
                peer_list = [p.strip().upper() for p in peers if p.strip().upper() not in (self.ticker, "")]
                if peer_list:
                    peer_df = self._create_peer_comps(metrics, peer_list)
                    if peer_df is not None:
                        components["peer_comparables"] = peer_df

        llm_analysis = {} if lean else self._generate_llm_financial_analysis(metrics, model_type)

        model = {
            "ticker": self.ticker,
            "company_name": metrics["company_info"].get("name", self.ticker),
            "model_type": model_type,
            "generated_at": datetime.utcnow().isoformat(),
            "projection_years": projection_years,
            "llm_analysis": llm_analysis,
            "financial_metrics": metrics,
            "model_components": components,
            "valuation_summary": valuation_summary,
            "generation_stats": {
                "models_generated": 1,
                "analysis_sections": self.analysis_sections,
                "data_points_processed": self.data_points_processed,
                "forecast_cache_entries": len(self._forecast_cache),
            },
            "parameters": {
                "term_growth": term_growth,
                "wacc": valuation_summary.get("WACC"),
                "strategy": valuation_summary.get("Strategy"),
                "peers": peers,
                "lean": lean,
            },
            "validation": self._validation_summary(chosen_strategy_name or "n/a", valuation_summary),
        }
        if self.llm_param_audit:
            model["llm_parameter_overrides"] = self.llm_param_audit
            # Persist audit to JSON for reproducibility
            try:
                self._ensure_dirs()
                audit_path = self.models_dir / f"llm_audit_{self.ticker}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
                with open(audit_path,'w',encoding='utf-8') as f:
                    json.dump(self.llm_param_audit, f, indent=2)
                model['llm_audit_file'] = str(audit_path)
                self._log('info', f"LLM audit saved: {audit_path}")
            except Exception as e:
                self._log('warning', f"Failed to persist LLM audit: {e}")
        self.models_generated += 1
        return model

    # ---------- LLM Assisted Parameter Overrides (bounded & auditable) ----------
    def propose_llm_parameter_overrides(self, baseline_model: Dict[str, Any], projection_years: int,
                                        term_growth: float, strategy: Optional[str],
                                        caps: Optional[Dict[str, Any]] = None, temperature: float = 0.1) -> Dict[str, Any]:
        """Ask LLM for refined parameter overrides using comprehensive agentic analysis.

        Returns dict of validated overrides without mutating state.
        Safe if LLM unavailable (returns empty dict).
        """
        if not self.llm_function:
            self._log("warning", "LLM unavailable - using default parameter calculation")
            return {}
            
        # Load financial data for comprehensive analysis
        try:
            financial_data = self._load_financial_data()
            return self._generate_agentic_parameters(financial_data, baseline_model, projection_years, 
                                                   term_growth, strategy, temperature)
        except Exception as e:
            self._log("error", f"Agentic parameter calculation failed: {e}")
            return {}

    def _generate_agentic_parameters(self, financial_data: Dict[str, Any], baseline_model: Dict[str, Any],
                                   projection_years: int, term_growth: float, 
                                   strategy: Optional[str], temperature: float) -> Dict[str, Any]:
        """Generate comprehensive parameter recommendations using agentic LLM analysis."""
        
        # Extract comprehensive company context
        context = self._build_company_context(financial_data)
        
        # Get peer analysis if available
        peer_analysis = self._get_peer_context(financial_data)
        
        # Build strategy-specific guidance
        strategy_guidance = self._get_strategy_guidance(strategy, context)
        
        # Load agentic parameter calculator prompt
        prompt_path = pathlib.Path(__file__).parent.parent / "prompts" / "agentic_parameter_calculator.md"
        
        try:
            with open(prompt_path, 'r', encoding='utf-8') as f:
                prompt_template = f.read()
        except Exception:
            self._log("warning", "Could not load agentic parameter prompt - using fallback")
            return self._fallback_parameter_calculation(context)
        
        # Build a safe prompt: prepend structured JSON context, then include the template as-is (no .format on template)
        context_payload = {
            'ticker': self.ticker,
            'company_name': context.get('company_name', self.ticker),
            'sector': context.get('sector', 'Unknown'),
            'industry': context.get('industry', 'Unknown'),
            'market_cap': context.get('market_cap', 0),
            'enterprise_value': context.get('enterprise_value', 0),
            'current_price': context.get('current_price', 0),
            'latest_revenue': context.get('latest_revenue', 0),
            'latest_ebitda': context.get('latest_ebitda', 0),
            'ebitda_margin': context.get('ebitda_margin', 0),
            'historical_growth': context.get('historical_growth', 0),
            'historical_summary': context.get('historical_summary', 'No historical data available'),
            'peer_analysis': peer_analysis,
            'risk_free_rate': context.get('risk_free_rate', 0.04),
            'market_risk_premium': context.get('market_risk_premium', 0.06),
            'beta': context.get('beta', 1.0),
            'economic_context': context.get('economic_context', 'Current market conditions'),
            'strategy_specific_guidance': strategy_guidance,
            'projection_years': projection_years,
            'term_growth': term_growth,
        }
        try:
            context_json = json.dumps(context_payload, indent=2)
        except Exception:
            # As a fallback, stringify keys one-by-one
            context_json = str(context_payload)
        prompt = (
            "You are provided with company context below in JSON. Use it along with the instructions to propose modeling parameters.\n" 
            + context_json + "\n\n" + prompt_template
        )
        
        # Make LLM call with comprehensive context
        try:
            messages = [
                {"role": "system", "content": "You are a senior financial analyst specializing in DCF modeling and valuation parameters. Provide realistic, defensible parameter recommendations."},
                {"role": "user", "content": prompt}
            ]
            
            response, cost = self.llm_function(messages, temperature=temperature)
            
            # Track LLM usage
            if hasattr(self, 'total_llm_cost'):
                self.total_llm_cost += cost
            if hasattr(self, 'llm_call_count'):
                self.llm_call_count += 1
                
            # Parse and validate LLM response
            parameters = self._parse_agentic_response(response, context)
            
            # Store audit trail
            self.llm_param_audit = {
                "method": "agentic_comprehensive",
                "context_used": list(context.keys()),
                "llm_cost": cost,
                "parameters_generated": list(parameters.keys()),
                "confidence": parameters.get('confidence_level', 'Unknown')
            }
            
            return parameters
            
        except Exception as e:
            self._log("error", f"Agentic LLM parameter generation failed: {e}")
            return self._fallback_parameter_calculation(context)

    def _build_company_context(self, financial_data: Dict[str, Any]) -> Dict[str, Any]:
        """Build comprehensive company context for agentic analysis."""
        
        cd = financial_data.get("company_data", {})
        basic = cd.get("basic_info", {})
        md = cd.get("market_data", {})
        vm = cd.get("valuation_metrics", {})
        cs = cd.get("capital_structure", {})
        
        # Financial statements
        fs = financial_data.get("financial_statements", {})
        is_map = fs.get("income_statement", {})
        
        # Latest financials
        latest_year = max(is_map.keys()) if is_map else None
        latest_is = is_map.get(latest_year, {}) if latest_year else {}
        
        latest_revenue = self._num(latest_is.get("Total Revenue")) or 0
        latest_ebitda = self._num(latest_is.get("EBITDA")) or 0
        ebitda_margin = (latest_ebitda / latest_revenue) if latest_revenue else 0
        
        # Historical growth calculation
        historical_growth = self._calculate_historical_growth(is_map)
        historical_summary = self._generate_historical_summary(is_map)
        
        return {
            'company_name': basic.get('long_name', self.ticker),
            'sector': basic.get('sector', 'Unknown'),
            'industry': basic.get('industry', 'Unknown'),
            'market_cap': self._num(md.get('market_cap'), 0),
            'enterprise_value': self._num(md.get('enterprise_value'), 0),
            'current_price': self._num(md.get('current_price'), 0),
            'latest_revenue': latest_revenue,
            'latest_ebitda': latest_ebitda,
            'ebitda_margin': ebitda_margin,
            'historical_growth': historical_growth,
            'historical_summary': historical_summary,
            'risk_free_rate': self._num(cs.get('risk_free_rate'), 0.04),
            'market_risk_premium': self._num(cs.get('equity_risk_premium'), 0.06),
            'beta': self._num(cs.get('beta'), 1.0),
            'economic_context': self._assess_economic_context(cd)
        }

    def _calculate_historical_growth(self, is_map: Dict[str, Any]) -> float:
        """Calculate compound annual revenue growth from historical data."""
        if len(is_map) < 2:
            return 0.0
            
        years = sorted(is_map.keys())
        revenues = []
        
        for year in years[-3:]:  # Last 3 years
            revenue = self._num(is_map[year].get("Total Revenue"))
            if revenue:
                revenues.append(revenue)
        
        if len(revenues) < 2:
            return 0.0
            
        # Calculate CAGR
        n_years = len(revenues) - 1
        cagr = (revenues[-1] / revenues[0]) ** (1/n_years) - 1
        return cagr

    def _generate_historical_summary(self, is_map: Dict[str, Any]) -> str:
        """Generate summary of historical financial performance."""
        if not is_map:
            return "No historical financial data available"
            
        years = sorted(is_map.keys())[-3:]  # Last 3 years
        summary_lines = []
        
        for year in years:
            data = is_map[year]
            revenue = self._num(data.get("Total Revenue"))
            ebitda = self._num(data.get("EBITDA"))
            margin = (ebitda / revenue * 100) if (revenue and ebitda) else 0
            
            summary_lines.append(f"{year}: Revenue ${revenue/1e6:.0f}M, EBITDA ${ebitda/1e6:.0f}M ({margin:.1f}%)")
        
        return "\n".join(summary_lines)

    def _get_peer_context(self, financial_data: Dict[str, Any]) -> str:
        """Get peer company analysis context."""
        # This would integrate with peer comparison data if available
        return "Peer analysis: Industry benchmarks suggest similar companies trade at 15-25x EBITDA with 10-20% revenue growth"

    def _get_strategy_guidance(self, strategy: Optional[str], context: Dict[str, Any]) -> str:
        """Get strategy-specific parameter guidance."""
        sector = context.get('sector', '').lower()
        industry = context.get('industry', '').lower()
        
        if strategy == 'saas_dcf' or 'software' in industry:
            return """
SaaS/Software Strategy Considerations:
- Growth: Typically 15-40% revenue growth, declining over time
- Margins: Should improve due to operating leverage (target 25-40% EBITDA)
- CapEx: Low (2-5% of revenue) - mainly technology infrastructure
- Working Capital: Often negative due to deferred revenue
- WACC: Typically 8-12% depending on growth stage and profitability
"""
        elif strategy == 'reit_dcf' or 'reit' in industry:
            return """
REIT Strategy Considerations:
- Growth: Modest 2-6% revenue growth from rent increases and acquisitions
- Margins: Focus on FFO/AFFO rather than EBITDA
- CapEx: Split between maintenance (50-80% of D&A) and growth capex
- Working Capital: Minimal impact
- WACC: Lower due to real estate backing (6-9%)
"""
        else:
            return """
Generic Strategy Considerations:
- Growth: Analyze competitive position and market dynamics
- Margins: Consider scale economies and competitive pressures
- CapEx: Asset-intensive vs asset-light business models
- Working Capital: Industry-specific patterns
- WACC: Risk-adjusted cost of capital based on business model
"""

    def _assess_economic_context(self, company_data: Dict[str, Any]) -> str:
        """Assess current economic environment for modeling."""
        return "Current environment: Moderate growth, elevated interest rates, focus on profitability over growth"

    def _parse_agentic_response(self, response: str, context: Dict[str, Any]) -> Dict[str, Any]:
        """Parse and validate LLM agentic parameter response."""
        
        try:
            # Extract JSON from response
            import re
            json_match = re.search(r'\{[\s\S]*\}', response)
            if not json_match:
                raise ValueError("No JSON found in response")
            
            import json
            parsed = json.loads(json_match.group())
            
            # Extract parameters with validation
            parameters = parsed.get('parameters', {})
            validated_params = {}
            
            # Validate and bound parameters
            bounds = {
                'first_year_growth': (0.0, 0.6),
                'terminal_growth': (0.0, 0.04),
                'margin_target': (0.05, 0.6),
                'margin_ramp': (0.0, 0.05),
                'capex_rate': (0.0, 0.2),
                'da_rate': (0.0, 0.15),
                'nwc_ratio': (-0.1, 0.4),
                'wacc': (0.04, 0.2)
            }
            
            for param, value in parameters.items():
                if param in bounds:
                    min_val, max_val = bounds[param]
                    bounded_value = max(min_val, min(max_val, float(value)))
                    validated_params[param] = bounded_value
            
            # Add growth sequence if provided
            if 'growth_sequence' in parameters:
                growth_seq = parameters['growth_sequence']
                if isinstance(growth_seq, list) and len(growth_seq) >= 5:
                    validated_params['growth_sequence'] = [max(0.0, min(0.6, float(g))) for g in growth_seq[:5]]
            
            # Store additional analysis
            validated_params['confidence_level'] = parsed.get('confidence_level', 'Medium')
            validated_params['analysis_summary'] = parsed.get('analysis_summary', '')
            validated_params['reasoning'] = parsed.get('reasoning', {})
            
            self._log("info", f"Agentic parameters generated: {list(validated_params.keys())}")
            return validated_params
            
        except Exception as e:
            self._log("error", f"Failed to parse agentic response: {e}")
            return self._fallback_parameter_calculation(context)

    def _fallback_parameter_calculation(self, context: Dict[str, Any]) -> Dict[str, Any]:
        """Fallback parameter calculation when LLM fails."""
        
        # Use intelligent defaults based on company context
        market_cap = context.get('market_cap', 0)
        sector = context.get('sector', '').lower()
        historical_growth = context.get('historical_growth', 0)
        ebitda_margin = context.get('ebitda_margin', 0)
        
        # Size-based adjustments
        if market_cap > 100e9:  # Large cap
            growth_mult = 0.7
            margin_target = min(0.35, max(0.2, ebitda_margin * 1.1))
        elif market_cap > 10e9:  # Mid cap
            growth_mult = 1.0
            margin_target = min(0.4, max(0.15, ebitda_margin * 1.15))
        else:  # Small cap
            growth_mult = 1.3
            margin_target = min(0.45, max(0.1, ebitda_margin * 1.2))
        
        # Sector-based adjustments
        if 'technology' in sector or 'software' in sector:
            base_growth = max(0.15, historical_growth * growth_mult)
            capex_rate = 0.03
        elif 'utility' in sector or 'infrastructure' in sector:
            base_growth = max(0.03, historical_growth * 0.5)
            capex_rate = 0.08
        else:
            base_growth = max(0.08, historical_growth * growth_mult)
            capex_rate = 0.05
        
        return {
            'first_year_growth': min(0.4, base_growth),
            'margin_target': margin_target,
            'margin_ramp': 0.01,
            'capex_rate': capex_rate,
            'da_rate': 0.04,
            'nwc_ratio': 0.05,
            'confidence_level': 'Low - Fallback calculation'
        }

    # ---------- Excel / CSV writers ----------
    def _ensure_dirs(self):
        self.company_dir.mkdir(parents=True, exist_ok=True)
        self.models_dir.mkdir(parents=True, exist_ok=True)

    def save_model_to_excel(self, model: Dict[str, Any]) -> pathlib.Path:
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl required. Install with: pip install openpyxl")

        self._ensure_dirs()
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"financial_model_{model['model_type']}_{model['ticker']}_{ts}.xlsx"
        out = self.models_dir / filename
        tmp = out.with_suffix(".xlsx.tmp")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Summary
        ws = wb.create_sheet("Executive Summary")
        ws["A1"] = f"{model['company_name']} ({model['ticker']}) — Financial Model"
        ws["A1"].font = Font(size=16, bold=True)

        ws["A3"] = "Model Type:"; ws["B3"] = model["model_type"].title()
        ws["A4"] = "Generated:" ; ws["B4"] = model["generated_at"][:19].replace("T"," ")
        ws["A5"] = "Projection Years:"; ws["B5"] = model["projection_years"]

        # Company info
        ws["A7"] = "Company Information:"; ws["A7"].font = Font(bold=True)
        r = 8
        for k, v in model.get("financial_metrics", {}).get("company_info", {}).items():
            if v is not None:
                ws[f"A{r}"] = k.replace("_", " ").title() + ":"
                ws[f"B{r}"] = v
                r += 1

        # Valuation block if available
        val = model.get("valuation_summary") or {}
        if val:
            r += 1
            ws[f"A{r}"] = "Valuation"; ws[f"A{r}"].font = Font(bold=True); r += 1
            for k in ["WACC","Terminal Growth","PV of FCFF","PV of TV","Enterprise Value","Net Debt","Equity Value","Shares (basic)","Implied Price"]:
                ws[f"A{r}"] = k + ":"
                ws[f"B{r}"] = val.get(k)
                r += 1

        # Dynamic DataFrame tabs
        for key, df in model["model_components"].items():
            if isinstance(df, pd.DataFrame):
                sheet_name = EXCEL_CONFIG.SHEET_NAMES.get(key, key[:EXCEL_CONFIG.MAX_SHEET_NAME_LENGTH])
                wsdf = wb.create_sheet(sheet_name)
                for ridx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for cidx, value in enumerate(row, 1):
                        cell = wsdf.cell(row=ridx, column=cidx, value=value)
                        if ridx == 1:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color=EXCEL_CONFIG.HEADER_FILL_COLOR, end_color=EXCEL_CONFIG.HEADER_FILL_COLOR, fill_type="solid")

        # Narrative
        if model.get("llm_analysis"):
            na_ws = wb.create_sheet("LLM Analysis")
            rr = 1
            for sec, txt in model["llm_analysis"].items():
                na_ws[f"A{rr}"] = sec.replace("_"," ").title()
                na_ws[f"A{rr}"].font = Font(bold=True)
                rr += 1
                for line in str(txt).splitlines():
                    na_ws[f"A{rr}"] = line
                    rr += 1
                rr += 1

        # Validation sheet (structured formatting)
        val_ws = wb.create_sheet("Validation")
        vdata = model.get("validation", {}) or {}
        # Strategy
        val_ws["A1"] = "Strategy"; val_ws["A1"].font = Font(bold=True)
        val_ws["B1"] = vdata.get("strategy")
        # Overrides
        val_ws["A3"] = "Overrides"; val_ws["A3"].font = Font(bold=True)
        overrides = vdata.get("overrides") or {}
        row = 4
        for k in sorted(overrides.keys()):
            val_ws[f"A{row}"] = k
            # Handle arrays and other non-Excel-compatible types
            value = overrides[k]
            if isinstance(value, (list, tuple)):
                # Convert arrays to comma-separated strings
                val_ws[f"B{row}"] = ", ".join(str(x) for x in value)
            elif isinstance(value, dict):
                # Convert dicts to string representation
                val_ws[f"B{row}"] = str(value)
            else:
                val_ws[f"B{row}"] = value
            row += 1
        # Missing valuation fields
        row += 1
        val_ws[f"A{row}"] = "Missing Valuation Fields"; val_ws[f"A{row}"].font = Font(bold=True)
        mvf = vdata.get("missing_valuation_fields") or []
        row += 1
        if mvf:
            fill_warn = PatternFill(start_color=EXCEL_CONFIG.WARNING_FILL_COLOR, end_color=EXCEL_CONFIG.WARNING_FILL_COLOR, fill_type="solid")
            for fld in mvf:
                val_ws[f"A{row}"] = fld
                val_ws[f"A{row}"].fill = fill_warn
                row += 1
        else:
            val_ws[f"A{row}"] = "(none)"; row += 1
        # Diagnostics
        row += 1
        val_ws[f"A{row}"] = "Diagnostics"; val_ws[f"A{row}"].font = Font(bold=True)
        row += 1
        diags = vdata.get("diagnostics") or []
        if diags:
            fill_info = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
            for d in diags:
                val_ws[f"A{row}"] = d
                val_ws[f"A{row}"].fill = fill_info
                row += 1
        else:
            val_ws[f"A{row}"] = "(none)"; row += 1
        # Meta stats
        row += 1
        val_ws[f"A{row}"] = "Models Generated"; val_ws[f"B{row}"] = vdata.get("models_generated")
        row += 1
        val_ws[f"A{row}"] = "Cache Entries"; val_ws[f"B{row}"] = model.get("generation_stats", {}).get("forecast_cache_entries")

        wb.save(tmp)            # write to temp file
        wb.close()              # close workbook resources
        os.replace(tmp, out)    # atomic on POSIX/NTFS

        self._log("info", f"Excel saved: {out}")
        
        # Also save a "latest" version for easy access (similar to financials_annual_modeling_latest.json)
        latest_filename = f"financial_model_{model['model_type']}_latest.xlsx"
        latest_path = self.models_dir / latest_filename
        shutil.copy2(out, latest_path)
        self._log("info", f"Latest Excel saved: {latest_path}")
        
        return out

    def save_model_to_csv(self, model: Dict[str, Any]) -> List[pathlib.Path]:
        self._ensure_dirs()
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        saved: List[pathlib.Path] = []
        for name, data in model.get("model_components", {}).items():
            if isinstance(data, pd.DataFrame):
                # Save timestamped version
                p = self.models_dir / f"{name}_{model['ticker']}_{ts}.csv"
                data.to_csv(p, index=False)
                saved.append(p)
                self._log("info", f"CSV saved: {p}")
                
                # Also save latest version
                latest_p = self.models_dir / f"{name}_latest.csv"
                data.to_csv(latest_p, index=False)
                self._log("info", f"Latest CSV saved: {latest_p}")
        return saved


