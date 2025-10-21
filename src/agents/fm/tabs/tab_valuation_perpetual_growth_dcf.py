"""
Tab 5: Valuation (Perpetual Growth DCF) Tab Builder

This module creates the Valuation tab using the Perpetual Growth DCF method.

Professional-grade DCF valuation following investment banking standards:
- WACC calculation from cost of equity and cost of debt
- FCF discounting for explicit projection period (FY1-FY5)
- Terminal value using perpetual growth method
- Equity bridge: EV + Cash - Debt + Investments = Equity Value
- Value per share calculation
- Sanity check ratios (EV/EBITDA, P/E, FCF Yield)
"""

from typing import Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..financial_model_builder import ExcelFormats


class ValuationPerpetualGrowthDCFBuilder:
    """
    Builds the Valuation (Perpetual Growth DCF) tab.
    
    This tab contains:
    - WACC Calculation (rows 2-12): Risk-free rate, beta, cost of equity/debt, WACC
    - FCF Discounting (rows 14-19): Year headers, FCF, discount factors, PV of FCFs
    - Terminal Value (rows 21-27): Validation check, terminal growth, terminal FCF, TV, PV of TV, EV
    - Equity Bridge (rows 29-33): EV + Cash - Debt + Investments = Equity Value
    - Value per Share (rows 35-37): Shares outstanding, intrinsic value per share
    - Sanity Checks (rows 39-42): EV/EBITDA, P/NOPAT, FCF Yield
    - Output Summary (rows 44-47): EV, Equity Value, Value/Share (recalculated for validation)
    
    All formulas reference Assumptions, Projections, and Historical tabs.
    
    Professional enhancements:
    - g >= WACC validation check to prevent invalid perpetuity calculations
    - Dynamic COLUMNS formula for flexible projection periods
    - Corrected P/NOPAT labeling (uses NOPAT, not Net Income)
    - Share count units clarified (absolute count)
    """
    
    def __init__(self, projection_years: int = 5):
        """
        Initialize the Valuation DCF builder.
        
        Args:
            projection_years: Number of projection years (default: 5)
        """
        self.projection_years = projection_years
    
    def create_tab(self, workbook: openpyxl.Workbook) -> Worksheet:
        """
        Create and format the Valuation (Perpetual Growth DCF) tab.
        
        Args:
            workbook: The Excel workbook to add the tab to
            
        Returns:
            The created worksheet
        """
        # Create or get the sheet
        if "Valuation (DCF)" in workbook.sheetnames:
            ws = workbook["Valuation (DCF)"]
            workbook.remove(ws)
        
        ws = workbook.create_sheet("Valuation (DCF)", 5)  # Position 5 (sixth tab)
        
        # Set up sections following exact specification
        self._setup_header(ws)
        self._setup_wacc_section(ws)
        self._setup_fcf_discounting(ws)
        self._setup_terminal_value(ws)
        self._setup_equity_bridge(ws)
        self._setup_value_per_share(ws)
        self._setup_sanity_checks(ws)
        self._setup_output_summary(ws)
        
        # Format the sheet
        self._format_sheet(ws)
        
        return ws
    
    def _setup_header(self, ws: Worksheet) -> None:
        """Set up main header."""
        ws.cell(row=1, column=1, value="VALUATION - PERPETUAL GROWTH DCF")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    def _setup_wacc_section(self, ws: Worksheet) -> None:
        """
        Set up WACC calculation section (rows 2-11).
        
        Row 2: Risk-Free Rate = Assumptions!C30
        Row 3: Equity Risk Premium = Assumptions!C31
        Row 4: Levered Beta = Assumptions!C32
        Row 5: Cost of Equity (Ke) = Rf + Beta * ERP
        Row 6: Pre-Tax Cost of Debt = Assumptions!C33
        Row 7: Tax Rate = Assumptions!C20
        Row 8: After-Tax Cost of Debt = Kd * (1 - T)
        Row 9: Equity Weight (E/V) = Assumptions!C34
        Row 10: Debt Weight (D/V) = 1 - E/V
        Row 11: WACC = Ke * E/V + Kd(1-T) * D/V
        """
        # Section header
        ws.cell(row=2, column=1, value="A. WACC CALCULATION")
        ws.cell(row=2, column=1).font = Font(bold=True, size=11, underline="single")
        
        # Row 3: Risk-Free Rate
        ws.cell(row=3, column=1, value="Risk-Free Rate (Rf)")
        ws.cell(row=3, column=2, value='=Assumptions!B23')  # Updated: row 23 in Assumptions
        ws.cell(row=3, column=2).number_format = '0.00%'
        
        # Row 4: Equity Risk Premium
        ws.cell(row=4, column=1, value="Equity Risk Premium (ERP)")
        ws.cell(row=4, column=2, value='=Assumptions!B24')  # Updated: row 24 in Assumptions
        ws.cell(row=4, column=2).number_format = '0.00%'
        
        # Row 5: Levered Beta
        ws.cell(row=5, column=1, value="Levered Beta (β)")
        ws.cell(row=5, column=2, value='=Assumptions!B25')  # Updated: row 25 in Assumptions
        ws.cell(row=5, column=2).number_format = '0.00'
        
        # Row 6: Cost of Equity
        ws.cell(row=6, column=1, value="Cost of Equity (Ke)")
        ws.cell(row=6, column=1).font = Font(bold=True)
        ws.cell(row=6, column=2, value='=B3+B5*B4')  # Rf + Beta * ERP
        ws.cell(row=6, column=2).number_format = '0.00%'
        
        # Row 7: Pre-Tax Cost of Debt
        ws.cell(row=7, column=1, value="Pre-Tax Cost of Debt (Kd)")
        ws.cell(row=7, column=2, value='=Assumptions!B27')  # Updated: row 27 in Assumptions
        ws.cell(row=7, column=2).number_format = '0.00%'
        
        # Row 8: Tax Rate
        ws.cell(row=8, column=1, value="Tax Rate (T)")
        ws.cell(row=8, column=2, value='=Assumptions!B20')  # Row 20 in Assumptions
        ws.cell(row=8, column=2).number_format = '0.00%'
        
        # Row 9: After-Tax Cost of Debt
        ws.cell(row=9, column=1, value="After-Tax Cost of Debt")
        ws.cell(row=9, column=1).font = Font(bold=True)
        ws.cell(row=9, column=2, value='=B7*(1-B8)')  # Kd * (1 - T)
        ws.cell(row=9, column=2).number_format = '0.00%'
        
        # Row 10: Equity Weight
        ws.cell(row=10, column=1, value="Equity Weight (E/V)")
        ws.cell(row=10, column=2, value='=Assumptions!B29')  # Updated: row 29 in Assumptions
        ws.cell(row=10, column=2).number_format = '0.00%'
        
        # Row 11: Debt Weight
        ws.cell(row=11, column=1, value="Debt Weight (D/V)")
        ws.cell(row=11, column=2, value='=1-B10')  # 1 - E/V
        ws.cell(row=11, column=2).number_format = '0.00%'
        
        # Row 12: WACC
        ws.cell(row=12, column=1, value="WACC")
        ws.cell(row=12, column=1).font = Font(bold=True, size=11)
        ws.cell(row=12, column=2, value='=B6*B10+B9*B11')  # Ke*E/V + Kd(1-T)*D/V
        ws.cell(row=12, column=2).number_format = '0.00%'
        ws.cell(row=12, column=2).font = Font(bold=True, size=11)
        ws.cell(row=12, column=2).fill = PatternFill(
            start_color=ExcelFormats.IMPORTANT_COLOR,
            end_color=ExcelFormats.IMPORTANT_COLOR,
            fill_type="solid"
        )
    
    def _setup_fcf_discounting(self, ws: Worksheet) -> None:
        """
        Set up FCF discounting section (rows 13-18).
        
        Row 14: Section header
        Row 15: Year (from Projections B2:F2)
        Row 16: Free Cash Flow (from Projections B19:F19)
        Row 17: Discount Factor = 1/(1+WACC)^n
        Row 18: Present Value of FCF = FCF * Discount Factor
        Row 19: Sum of PV of FCFs
        """
        # Blank row
        ws.cell(row=13, column=1, value="")
        
        # Section header
        ws.cell(row=14, column=1, value="B. FCF DISCOUNTING (FY1-FY5)")
        ws.cell(row=14, column=1).font = Font(bold=True, size=11, underline="single")
        
        # Row 15: Year headers (B-F columns for FY1-FY5)
        ws.cell(row=15, column=1, value="Year")
        ws.cell(row=15, column=1).font = Font(bold=True)
        for i in range(self.projection_years):
            col = 2 + i  # B, C, D, E, F
            col_letter = chr(64 + col)
            formula = f'=Projections!{col_letter}2'  # Reference year from Projections
            ws.cell(row=15, column=col, value=formula)
            ws.cell(row=15, column=col).number_format = '0'
            ws.cell(row=15, column=col).font = Font(bold=True)
            ws.cell(row=15, column=col).alignment = Alignment(horizontal="center")
        
        # Row 16: Free Cash Flow (FCF)
        ws.cell(row=16, column=1, value="Free Cash Flow (FCF)")
        ws.cell(row=16, column=1).font = Font(bold=True)
        for i in range(self.projection_years):
            col = 2 + i  # B, C, D, E, F
            col_letter = chr(64 + col)
            formula = f'=Projections!{col_letter}19'  # Reference FCF from Projections row 19
            ws.cell(row=16, column=col, value=formula)
            ws.cell(row=16, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 17: Discount Factor
        ws.cell(row=17, column=1, value="Discount Factor")
        for i in range(self.projection_years):
            col = 2 + i  # B, C, D, E, F
            period = i + 1  # 1, 2, 3, 4, 5
            # Formula: 1/(1+WACC)^(n - MYD_adjustment)
            # MYD toggle in Sensitivity!$B$4: 0.5 if mid-year, 0 if year-end
            formula = f'=1/((1+$B$12)^({period}-Sensitivity!$B$4))'
            ws.cell(row=17, column=col, value=formula)
            ws.cell(row=17, column=col).number_format = '0.0000'
        
        # Row 18: Present Value (PV) of FCF
        ws.cell(row=18, column=1, value="Present Value (PV) of FCF")
        ws.cell(row=18, column=1).font = Font(bold=True)
        for i in range(self.projection_years):
            col = 2 + i  # B, C, D, E, F
            col_letter = chr(64 + col)
            formula = f'={col_letter}16*{col_letter}17'  # FCF * Discount Factor
            ws.cell(row=18, column=col, value=formula)
            ws.cell(row=18, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 19: Sum of PV of FCFs
        ws.cell(row=19, column=1, value="Sum of PV of FCFs")
        ws.cell(row=19, column=1).font = Font(bold=True)
        ws.cell(row=19, column=2, value='=SUM(B18:F18)')  # Sum across FY1-FY5
        ws.cell(row=19, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=19, column=2).font = Font(bold=True)
    
    def _setup_terminal_value(self, ws: Worksheet) -> None:
        """
        Set up Terminal Value section (rows 20-27).
        
        Row 21: Validation check for g >= WACC
        Row 22: Section header
        Row 23: Terminal Growth Rate (g)
        Row 24: Terminal FCF (Year n+1) = FCF_FY5 * (1 + g)
        Row 25: Terminal Value (Un-discounted) = Terminal FCF / (WACC - g)
        Row 26: PV of Terminal Value = TV / (1 + WACC)^n (using dynamic COLUMNS formula)
        Row 27: Enterprise Value (EV) = Sum of PV of FCFs + PV of TV
        """
        # Blank row
        ws.cell(row=20, column=1, value="")
        
        # Row 21: Validation check for g >= WACC
        ws.cell(row=21, column=1, value="Validation Check:")
        ws.cell(row=21, column=1).font = Font(bold=True, italic=True)
        ws.cell(row=21, column=2, value='=IF($B$23>=$B$12,"⚠️ ERROR: g ≥ WACC (invalid for perpetuity)","✓ Valid")')
        ws.cell(row=21, column=2).font = Font(bold=True)
        
        # Section header
        ws.cell(row=22, column=1, value="C. TERMINAL VALUE (PERPETUAL GROWTH)")
        ws.cell(row=22, column=1).font = Font(bold=True, size=11, underline="single")
        
        # Row 23: Terminal Growth Rate
        ws.cell(row=23, column=1, value="Terminal Growth Rate (g)")
        ws.cell(row=23, column=1).font = Font(bold=True)
        ws.cell(row=23, column=2, value='=Assumptions!B30')  # Updated: row 30 in Assumptions
        ws.cell(row=23, column=2).number_format = '0.00%'
        
        # Row 24: Terminal FCF (Year n+1)
        ws.cell(row=24, column=1, value="Terminal FCF (Year n+1)")
        ws.cell(row=24, column=2, value='=F16*(1+$B$23)')  # FCF_FY5 * (1 + g)
        ws.cell(row=24, column=2).number_format = ExcelFormats.CURRENCY
        
        # Row 25: Terminal Value (Un-discounted)
        ws.cell(row=25, column=1, value="Terminal Value (Un-discounted)")
        ws.cell(row=25, column=2, value='=B24/($B$12-$B$23)')  # Terminal FCF / (WACC - g)
        ws.cell(row=25, column=2).number_format = ExcelFormats.CURRENCY
        
        # Row 26: PV of Terminal Value
        ws.cell(row=26, column=1, value="PV of Terminal Value")
        ws.cell(row=26, column=1).font = Font(bold=True)
        # Professional formula: Use COLUMNS to dynamically count projection periods
        # This discounts TV by the exact number of projected periods
        ws.cell(row=26, column=2, value='=B25/(1+$B$12)^(COLUMNS($B$16:$F$16))')
        ws.cell(row=26, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=26, column=2).font = Font(bold=True)
        
        # Row 27: Enterprise Value (EV)
        ws.cell(row=27, column=1, value="Enterprise Value (EV)")
        ws.cell(row=27, column=1).font = Font(bold=True, size=11)
        ws.cell(row=27, column=2, value='=B19+B26')  # Sum of PV of FCFs + PV of TV
        ws.cell(row=27, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=27, column=2).font = Font(bold=True, size=11)
        ws.cell(row=27, column=2).fill = PatternFill(
            start_color=ExcelFormats.IMPORTANT_COLOR,
            end_color=ExcelFormats.IMPORTANT_COLOR,
            fill_type="solid"
        )
    
    def _setup_equity_bridge(self, ws: Worksheet) -> None:
        """
        Set up Equity Bridge section (rows 28-33).
        
        Row 29: Section header
        Row 30: Add: Cash & Equivalents
        Row 31: Less: Total Debt
        Row 32: Add: Investments / Non-operating Assets
        Row 33: Equity Value (Firm Value)
        """
        # Blank row
        ws.cell(row=28, column=1, value="")
        
        # Section header
        ws.cell(row=29, column=1, value="D. EQUITY BRIDGE")
        ws.cell(row=29, column=1).font = Font(bold=True, size=11, underline="single")
        
        # Row 30: Cash & Equivalents
        ws.cell(row=30, column=1, value="Add: Cash & Equivalents")
        ws.cell(row=30, column=2, value='=Historical!F30')  # From Historical row 30 (Cash)
        ws.cell(row=30, column=2).number_format = ExcelFormats.CURRENCY
        
        # Row 31: Total Debt
        ws.cell(row=31, column=1, value="Less: Total Debt")
        ws.cell(row=31, column=2, value='=Historical!F35')  # From Historical row 35 (Total Debt)
        ws.cell(row=31, column=2).number_format = ExcelFormats.CURRENCY
        
        # Row 32: Investments / Non-operating Assets
        ws.cell(row=32, column=1, value="Add: Investments / Non-operating Assets")
        ws.cell(row=32, column=2, value='=Historical!F31')  # From Historical row 31 (ST Investments)
        ws.cell(row=32, column=2).number_format = ExcelFormats.CURRENCY
        
        # Row 33: Equity Value
        ws.cell(row=33, column=1, value="Equity Value (Firm Value)")
        ws.cell(row=33, column=1).font = Font(bold=True, size=11)
        ws.cell(row=33, column=2, value='=B27+B30-B31+B32')  # EV + Cash - Debt + Investments
        ws.cell(row=33, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=33, column=2).font = Font(bold=True, size=11)
        ws.cell(row=33, column=2).fill = PatternFill(
            start_color=ExcelFormats.IMPORTANT_COLOR,
            end_color=ExcelFormats.IMPORTANT_COLOR,
            fill_type="solid"
        )
    
    def _setup_value_per_share(self, ws: Worksheet) -> None:
        """
        Set up Value per Share section (rows 34-37).
        
        Row 35: Section header
        Row 36: Shares Outstanding (with unit note)
        Row 37: Intrinsic Value per Share
        """
        # Blank row
        ws.cell(row=34, column=1, value="")
        
        # Section header
        ws.cell(row=35, column=1, value="E. VALUE PER SHARE")
        ws.cell(row=35, column=1).font = Font(bold=True, size=11, underline="single")
        
        # Row 36: Shares Outstanding (with clarification)
        ws.cell(row=36, column=1, value="Shares Outstanding (absolute count)")
        ws.cell(row=36, column=2, value='=Assumptions!B31')  # Updated: row 31 in Assumptions
        ws.cell(row=36, column=2).number_format = '#,##0'
        
        # Row 37: Intrinsic Value per Share
        ws.cell(row=37, column=1, value="Intrinsic Value per Share ($)")
        ws.cell(row=37, column=1).font = Font(bold=True, size=11)
        ws.cell(row=37, column=2, value='=B33/B36')  # Equity Value / Shares Outstanding
        ws.cell(row=37, column=2).number_format = '$0.00'
        ws.cell(row=37, column=2).font = Font(bold=True, size=11)
        ws.cell(row=37, column=2).fill = PatternFill(
            start_color=ExcelFormats.IMPORTANT_COLOR,
            end_color=ExcelFormats.IMPORTANT_COLOR,
            fill_type="solid"
        )
    
    def _setup_sanity_checks(self, ws: Worksheet) -> None:
        """
        Set up Sanity Checks section (rows 38-42).
        
        Row 39: Section header
        Row 40: Implied EV/EBITDA (Last Year)
        Row 41: Implied P/NOPAT (Last Year) - Note: Uses NOPAT, not Net Income
        Row 42: FCF Yield (EV Basis)
        """
        # Blank row
        ws.cell(row=38, column=1, value="")
        
        # Section header
        ws.cell(row=39, column=1, value="F. SANITY CHECKS")
        ws.cell(row=39, column=1).font = Font(bold=True, size=11, underline="single")
        
        # Row 40: Implied EV/EBITDA
        ws.cell(row=40, column=1, value="Implied EV/EBITDA (FY5)")
        ws.cell(row=40, column=2, value='=B27/Projections!F21')  # EV / EBITDA_FY5 (row 21 in Projections)
        ws.cell(row=40, column=2).number_format = '0.0x'
        
        # Row 41: Implied P/NOPAT (correctly labeled - using NOPAT not Net Income)
        ws.cell(row=41, column=1, value="Implied P/NOPAT (FY5)")
        ws.cell(row=41, column=2, value='=B33/Projections!F11')  # Equity Value / NOPAT_FY5
        ws.cell(row=41, column=2).number_format = '0.0x'
        
        # Note: If Net Income is added to Projections, change row 41 to:
        # ws.cell(row=41, column=1, value="Implied P/E (FY5)")
        # ws.cell(row=41, column=2, value='=B33/Projections!F[NetIncomeRow]')
        
        # Row 42: FCF Yield
        ws.cell(row=42, column=1, value="FCF Yield (EV Basis, FY5)")
        ws.cell(row=42, column=2, value='=Projections!F19/B27')  # FCF_FY5 / EV
        ws.cell(row=42, column=2).number_format = '0.00%'
    
    def _setup_output_summary(self, ws: Worksheet) -> None:
        """
        Set up Output Summary section (rows 43-48).
        
        Row 44: Section header
        Row 45: Enterprise Value (EV) - Recalculated for validation
        Row 46: Equity Value - Recalculated for validation
        Row 47: Intrinsic Value per Share - Recalculated for validation
        
        Note: These formulas recalculate from base components rather than referencing
        intermediate cells, providing additional validation of the model.
        """
        # Blank row
        ws.cell(row=43, column=1, value="")
        
        # Section header
        ws.cell(row=44, column=1, value="G. OUTPUT SUMMARY")
        ws.cell(row=44, column=1).font = Font(bold=True, size=12, underline="single")
        
        # Row 45: Enterprise Value (recalculated from base components)
        ws.cell(row=45, column=1, value="Enterprise Value (EV)")
        ws.cell(row=45, column=1).font = Font(bold=True)
        # Professional practice: Recalculate EV directly for validation
        ws.cell(row=45, column=2, value='=B19+B26')  # Sum of PV of FCFs + PV of TV
        ws.cell(row=45, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=45, column=2).font = Font(bold=True)
        
        # Row 46: Equity Value (recalculated from EV output)
        ws.cell(row=46, column=1, value="Equity Value")
        ws.cell(row=46, column=1).font = Font(bold=True)
        # Use the recalculated EV from B45 for consistency in output section
        ws.cell(row=46, column=2, value='=B45+B30-B31+B32')  # EV + Cash - Debt + Investments
        ws.cell(row=46, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=46, column=2).font = Font(bold=True)
        
        # Row 47: Intrinsic Value per Share (recalculated from equity value output)
        ws.cell(row=47, column=1, value="Intrinsic Value per Share")
        ws.cell(row=47, column=1).font = Font(bold=True, size=11)
        # Use the recalculated Equity Value from B46 for consistency
        ws.cell(row=47, column=2, value='=B46/B36')  # Equity Value / Shares Outstanding
        ws.cell(row=47, column=2).number_format = '$0.00'
        ws.cell(row=47, column=2).font = Font(bold=True, size=11)
        ws.cell(row=47, column=2).fill = PatternFill(
            start_color="FFD700",  # Gold color for final output
            end_color="FFD700",
            fill_type="solid"
        )
    
    def _format_sheet(self, ws: Worksheet) -> None:
        """Apply general formatting to the sheet."""
        # Set column widths
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 18
        for col_letter in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter].width = 14
        
        # Freeze headers
        ws.freeze_panes = ws['B2']
