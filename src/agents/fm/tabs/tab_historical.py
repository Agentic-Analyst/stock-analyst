"""
Tab 3: Historical Tab Builder

This module creates the Historical tab - pulls last 4-5 years of actual
financial data into normalized P&L, Balance Sheet, and Cash Flow sections.

All formulas reference the Raw tab using SUMIFS with wildcard matching on years.
Years are extracted directly from Raw data to ensure only complete years are used.

Key Features:
- Dynamic year detection (only includes years with complete data)
- Exact field name matching with Raw tab
- Calculated metrics (margins, EBITDA, working capital efficiency)
- QA checks (FCF reconciliation, EBIT validation)
"""

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment

from ..financial_model_builder import ExcelFormats


class HistoricalTabBuilder:
    """
    Builds the Historical tab - last 4-5 years of actual financials.
    
    This tab contains:
    - Row 1: Year headers (dynamically extracted from Raw data)
    - Rows 3-21: Income Statement with margins
    - Rows 23-27: Cash Flow with FCF reconciliation
    - Rows 30-42: Balance Sheet / Working Capital metrics
    - Rows 44-45: QA / Tie-outs (EBIT validation)
    
    Note: Only includes years with complete financial data.
    """
    
    def __init__(self):
        """Initialize the Historical builder."""
        self.year_headers = {}  # Map of column -> year
        self.num_years = 0
        self.start_col = 2
    
    def create_tab(self, workbook: openpyxl.Workbook) -> Worksheet:
        """
        Create and format the Historical tab.
        
        Args:
            workbook: The Excel workbook to add the tab to
            
        Returns:
            The created worksheet
        """
        # Extract years from Raw tab first
        self._extract_years_from_raw(workbook)
        
        # Create or get the sheet
        if "Historical" in workbook.sheetnames:
            ws = workbook["Historical"]
            workbook.remove(ws)
        
        ws = workbook.create_sheet("Historical", 3)  # Position 3 (fourth tab)
        
        # Set up all sections
        self._setup_headers(ws)
        self._setup_stock_price_row(ws)
        self._setup_income_statement(ws)
        self._setup_cash_flow(ws)
        self._setup_balance_sheet_wc(ws)
        self._setup_qa_tieouts(ws)
        
        # Format the sheet
        self._format_sheet(ws)
        
        return ws
    
    def _extract_years_from_raw(self, workbook: openpyxl.Workbook) -> None:
        """
        Extract years with complete financial data from Raw tab.
        
        Only includes years that have all key metrics to avoid empty columns.
        """
        raw = workbook['Raw']
        
        # Key fields that must exist for a year to be considered complete
        key_fields = [
            'Total Revenue', 'Cost Of Revenue', 'Operating Income', 'Net Income',
            'Operating Cash Flow', 'Cash And Cash Equivalents'
        ]
        
        years_data = {}
        for row in raw.iter_rows(min_row=2, values_only=True):
            if row[2] and isinstance(row[2], str) and len(row[2]) >= 10:
                year = row[2][:4]
                if year.isdigit():
                    if year not in years_data:
                        years_data[year] = set()
                    years_data[year].add(row[1])
        
        # Filter to only years with all key fields
        complete_years = []
        for year in sorted(years_data.keys()):
            has_all = all(field in years_data[year] for field in key_fields)
            if has_all:
                complete_years.append(int(year))
        
        years = sorted(complete_years, reverse=True)  # Newest to oldest
        
        # Determine column layout (max 5 years)
        self.num_years = min(len(years), 5)
        self.start_col = 7 - self.num_years  # If 4 years, start at C(3); if 5, start at B(2)
        
        # Map columns to years (oldest to newest, left to right)
        for i in range(self.num_years):
            col = self.start_col + i
            year = years[self.num_years - 1 - i]
            self.year_headers[col] = year
    
    def _setup_headers(self, ws: Worksheet) -> None:
        """
        Set up year column headers (row 1).
        
        Uses actual year values extracted from Raw data, not formulas.
        This ensures years are correct even before workbook calculation.
        """
        # A1: Label
        ws.cell(row=1, column=1, value="Metric")
        ws.cell(row=1, column=1).font = Font(bold=True, size=11)
        ws.cell(row=1, column=1).fill = PatternFill(
            start_color=ExcelFormats.HEADER_COLOR,
            end_color=ExcelFormats.HEADER_COLOR,
            fill_type="solid"
        )
        
        # Fill year headers (B1-F1) with actual years
        for col in range(2, 7):
            if col in self.year_headers:
                year = self.year_headers[col]
                ws.cell(row=1, column=col, value=year)
                ws.cell(row=1, column=col).font = Font(bold=True, size=11)
                ws.cell(row=1, column=col).fill = PatternFill(
                    start_color=ExcelFormats.HEADER_COLOR,
                    end_color=ExcelFormats.HEADER_COLOR,
                    fill_type="solid"
                )
                ws.cell(row=1, column=col).number_format = "0000"
                ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
            else:
                # Empty column (not used)
                ws.cell(row=1, column=col, value="")
    
    def _setup_stock_price_row(self, ws: Worksheet) -> None:
        """
        Set up row 2: Current Stock Price
        
        References Market Data from Raw tab if available.
        
        The formula attempts to pull 'current_price' from Raw tab:
        - If Raw tab has market_data with current_price field: displays price
        - If no data in Raw: cell is empty, user must manually enter
        
        Note: Sensitivity tab references Historical!F2 for market price comparison
        """
        # A2: Label
        ws.cell(row=2, column=1, value="Current Stock Price")
        ws.cell(row=2, column=1).font = Font(bold=True)
        
        # F2: Stock price formula
        # Try multiple field names (current_price, price, current_market_price)
        formula = '=IFERROR(SUMIFS(Raw!$D:$D,Raw!$B:$B,"current_price",Raw!$C:$C,"Current"),'
        formula += 'IFERROR(SUMIFS(Raw!$D:$D,Raw!$B:$B,"price",Raw!$C:$C,"Current"),"")'
        formula += ')'
        
        ws.cell(row=2, column=6, value=formula)
        ws.cell(row=2, column=6).number_format = ExcelFormats.CURRENCY
        ws.cell(row=2, column=6).fill = PatternFill(
            start_color=ExcelFormats.INPUT_COLOR,  # Yellow - indicates manual input may be needed
            end_color=ExcelFormats.INPUT_COLOR,
            fill_type="solid"
        )
        ws.cell(row=2, column=6).font = Font(bold=True)
    
    def _setup_income_statement(self, ws: Worksheet) -> None:
        """
        Set up Income Statement section (rows 3-21).
        
        Includes:
        - Revenue, COGS, Gross Profit
        - Operating expenses (R&D, SG&A)
        - Operating Income, Other Income, Tax
        - Net Income, EPS
        - Margins (Gross, EBITDA, Operating)
        """
        # Row definitions: (row, label, raw_field, format, is_bold)
        # raw_field=None means calculated field
        income_rows = [
            (3, "Total Revenue", "Total Revenue", ExcelFormats.CURRENCY, True),
            (4, "Cost Of Revenue", "Cost Of Revenue", ExcelFormats.CURRENCY, False),
            (5, "Gross Profit", None, ExcelFormats.CURRENCY, True),
            (6, "Research And Development", "Research And Development", ExcelFormats.CURRENCY, False),
            (7, "Selling General And Administration", "Selling General And Administration", ExcelFormats.CURRENCY, False),
            (8, "Operating Income", "Operating Income", ExcelFormats.CURRENCY, True),
            (9, "Other Income Expense", "Other Income Expense", ExcelFormats.CURRENCY, False),
            (10, "Pretax Income", "Pretax Income", ExcelFormats.CURRENCY, False),
            (11, "Tax Provision", "Tax Provision", ExcelFormats.CURRENCY, False),
            (12, "Net Income", "Net Income", ExcelFormats.CURRENCY, True),
            (13, "Basic EPS", "Basic EPS", ExcelFormats.NUMBER_DECIMAL, False),
            (14, "Diluted EPS", "Diluted EPS", ExcelFormats.NUMBER_DECIMAL, False),
            (15, "Diluted Average Shares", "Diluted Average Shares", ExcelFormats.NUMBER, False),
            (17, "Gross Margin %", None, ExcelFormats.PERCENTAGE_DECIMAL, True),
            (18, "Depreciation & Amortization", "Depreciation And Amortization", ExcelFormats.CURRENCY, False),
            (19, "EBITDA", None, ExcelFormats.CURRENCY, True),
            (20, "EBITDA Margin %", None, ExcelFormats.PERCENTAGE_DECIMAL, True),
            (21, "Operating Margin %", None, ExcelFormats.PERCENTAGE_DECIMAL, True),
        ]
        
        for row, label, raw_field, num_format, is_bold in income_rows:
            # Set label in column A
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=is_bold)
            
            # Fill formulas for each year column
            for col in range(2, 7):
                if col not in self.year_headers:
                    continue
                
                col_letter = chr(65 + col - 1)
                
                if raw_field is None:
                    # Calculated formulas
                    if row == 5:  # Gross Profit
                        formula = f'={col_letter}3-{col_letter}4'
                    elif row == 17:  # Gross Margin %
                        formula = f'=IFERROR(ROUND({col_letter}5/{col_letter}3,4),"")'
                    elif row == 19:  # EBITDA
                        formula = f'={col_letter}8+{col_letter}18'
                    elif row == 20:  # EBITDA Margin %
                        formula = f'=IFERROR(ROUND({col_letter}19/{col_letter}3,4),"")'
                    elif row == 21:  # Operating Margin %
                        formula = f'=IFERROR(ROUND({col_letter}8/{col_letter}3,4),"")'
                else:
                    # SUMIFS formulas - exact field name matching
                    formula = f'=SUMIFS(Raw!$D:$D,Raw!$B:$B,"{raw_field}",Raw!$C:$C,{col_letter}$1&"*")'
                
                ws.cell(row=row, column=col, value=formula)
                ws.cell(row=row, column=col).number_format = num_format
                
                # Highlight calculated cells
                if raw_field is None:
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color=ExcelFormats.CALCULATED_COLOR,
                        end_color=ExcelFormats.CALCULATED_COLOR,
                        fill_type="solid"
                    )
    
    def _setup_cash_flow(self, ws: Worksheet) -> None:
        """
        Set up Cash Flow section (rows 23-27).
        
        Includes:
        - Operating Cash Flow
        - Capital Expenditure
        - Free Cash Flow (from JSON)
        - Free Cash Flow (recomputed for validation)
        - FCF Check (delta between JSON and calculated)
        """
        cashflow_rows = [
            (23, "Operating Cash Flow", "Operating Cash Flow", True),
            (24, "Capital Expenditure", "Capital Expenditure", False),
            (25, "Free Cash Flow (from JSON)", "Free Cash Flow", False),
            (26, "Free Cash Flow (recomputed)", None, False),
            (27, "FCF Check (Δ)", None, True),
        ]
        
        for row, label, raw_field, is_bold in cashflow_rows:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=is_bold)
            
            for col in range(2, 7):
                if col not in self.year_headers:
                    continue
                
                col_letter = chr(65 + col - 1)
                
                if raw_field is None:
                    if row == 26:  # FCF recomputed (OCF + CapEx, note CapEx is negative)
                        formula = f'={col_letter}23+{col_letter}24'
                    elif row == 27:  # FCF Check
                        formula = f'={col_letter}25-{col_letter}26'
                else:
                    formula = f'=SUMIFS(Raw!$D:$D,Raw!$B:$B,"{raw_field}",Raw!$C:$C,{col_letter}$1&"*")'
                
                ws.cell(row=row, column=col, value=formula)
                ws.cell(row=row, column=col).number_format = ExcelFormats.CURRENCY
                
                if raw_field is None:
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color=ExcelFormats.CALCULATED_COLOR,
                        end_color=ExcelFormats.CALCULATED_COLOR,
                        fill_type="solid"
                    )
    
    def _setup_balance_sheet_wc(self, ws: Worksheet) -> None:
        """
        Set up Balance Sheet / Working Capital section (rows 30-42).
        
        Includes:
        - Cash, ST Investments, AR, Inventory, AP
        - Total Debt, LT Debt
        - Net Debt (improved calculation including ST investments)
        - Working Capital Efficiency: DSO, DIO, DPO, CCC
        """
        balance_rows = [
            (30, "Cash And Cash Equivalents", "Cash And Cash Equivalents", ExcelFormats.CURRENCY, False),
            (31, "Short Term Investments", "Cash Cash Equivalents And Short Term Investments", ExcelFormats.CURRENCY, False),
            (32, "Accounts Receivable", "Accounts Receivable", ExcelFormats.CURRENCY, False),
            (33, "Inventory", "Inventory", ExcelFormats.CURRENCY, False),
            (34, "Accounts Payable", "Accounts Payable", ExcelFormats.CURRENCY, False),
            (35, "Total Debt", "Total Debt", ExcelFormats.CURRENCY, False),
            (36, "Long Term Debt", "Long Term Debt", ExcelFormats.CURRENCY, False),
            (37, "Net Debt", None, ExcelFormats.CURRENCY, True),
            (39, "DSO (days)", None, ExcelFormats.NUMBER_DECIMAL, True),
            (40, "DIO (days)", None, ExcelFormats.NUMBER_DECIMAL, True),
            (41, "DPO (days)", None, ExcelFormats.NUMBER_DECIMAL, True),
            (42, "Cash Conversion Cycle (days)", None, ExcelFormats.NUMBER_DECIMAL, True),
        ]
        
        for row, label, raw_field, num_format, is_bold in balance_rows:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=is_bold)
            
            for col in range(2, 7):
                if col not in self.year_headers:
                    continue
                
                col_letter = chr(65 + col - 1)
                
                if raw_field is None:
                    if row == 37:  # Net Debt (improved: Total Debt - Cash - ST Investments)
                        formula = f'={col_letter}35-{col_letter}30-{col_letter}31'
                    elif row == 39:  # DSO (Days Sales Outstanding)
                        formula = f'=IFERROR(ROUND({col_letter}32/({col_letter}3/365),2),"")'
                    elif row == 40:  # DIO (Days Inventory Outstanding)
                        formula = f'=IFERROR(ROUND({col_letter}33/({col_letter}4/365),2),"")'
                    elif row == 41:  # DPO (Days Payable Outstanding)
                        formula = f'=IFERROR(ROUND({col_letter}34/({col_letter}4/365),2),"")'
                    elif row == 42:  # CCC (Cash Conversion Cycle)
                        formula = f'=IFERROR(ROUND({col_letter}39+{col_letter}40-{col_letter}41,2),"")'
                else:
                    formula = f'=SUMIFS(Raw!$D:$D,Raw!$B:$B,"{raw_field}",Raw!$C:$C,{col_letter}$1&"*")'
                
                ws.cell(row=row, column=col, value=formula)
                ws.cell(row=row, column=col).number_format = num_format
                
                if raw_field is None:
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color=ExcelFormats.CALCULATED_COLOR,
                        end_color=ExcelFormats.CALCULATED_COLOR,
                        fill_type="solid"
                    )
    
    def _setup_qa_tieouts(self, ws: Worksheet) -> None:
        """
        Set up QA / Tie-outs section (rows 44-45).
        
        EBIT validation:
        - EBIT (recalc) = Gross Profit - R&D - SG&A
        - Compare to Operating Income (should be close/identical)
        """
        qa_rows = [
            (44, "EBIT (recalc)", ExcelFormats.CURRENCY),
            (45, "EBIT vs. Operating Income (Δ)", ExcelFormats.CURRENCY),
        ]
        
        for row, label, num_format in qa_rows:
            ws.cell(row=row, column=1, value=label)
            
            for col in range(2, 7):
                if col not in self.year_headers:
                    continue
                
                col_letter = chr(65 + col - 1)
                
                if row == 44:  # EBIT recalc (Gross Profit - R&D - SG&A)
                    formula = f'={col_letter}5-{col_letter}6-{col_letter}7'
                elif row == 45:  # EBIT vs Operating Income delta
                    formula = f'={col_letter}44-{col_letter}8'
                
                ws.cell(row=row, column=col, value=formula)
                ws.cell(row=row, column=col).number_format = num_format
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=ExcelFormats.CALCULATED_COLOR,
                    end_color=ExcelFormats.CALCULATED_COLOR,
                    fill_type="solid"
                )
    
    def _format_sheet(self, ws: Worksheet) -> None:
        """Apply general formatting to the sheet"""
        # Set column widths
        ws.column_dimensions['A'].width = 35
        for col_letter in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col_letter].width = 18
        
        # Freeze headers (freeze row 1 and column A)
        ws.freeze_panes = ws['B2']
