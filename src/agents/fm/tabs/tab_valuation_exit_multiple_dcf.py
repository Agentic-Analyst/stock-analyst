"""
Tab 6: Valuation (Exit Multiple DCF) Tab Builder

This module creates the Valuation tab using the Exit Multiple DCF method.

Professional-grade DCF valuation following investment banking standards:
- Uses terminal EV/EBITDA exit multiple instead of perpetual growth
- Consistent WACC and FCF discounting with Perpetual Growth method
- Terminal value based on EBITDA multiple at end of projection period
- Full equity bridge and per-share valuation
- Sanity checks and cross-references to Perpetual Growth method
"""

from typing import Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..financial_model_builder import ExcelFormats


class ValuationExitMultipleDCFBuilder:
    """
    Builds the Valuation (Exit Multiple DCF) tab.
    
    This tab contains:
    - WACC and Discounting Inputs (rows 2-5): WACC, Exit Multiple, Tax Rate, Periods
    - FCF Forecast (rows 7-10): FCF, Discount Factors, PV of FCFs, Sum
    - Terminal Value (rows 12-15): EBITDA, Exit Multiple, TV, PV of TV
    - Enterprise Value (row 17): Sum of PV FCFs + PV of TV
    - Equity Bridge (rows 19-22): Cash, Debt, Investments, Equity Value
    - Value per Share (rows 24-25): Shares, Intrinsic Value
    - Sanity Checks (rows 27-30): EV/EBITDA, EV/EBIT, FCF Yield, Cross-check
    - Output Summary (rows 32-35): EV, Equity, Per Share, Upside vs Market
    
    All formulas reference Assumptions, Projections, Historical, and Valuation (DCF) tabs.
    
    Professional features:
    - Complementary to Perpetual Growth method
    - Exit multiple from market comparables
    - Cross-validation with perpetual method
    - Market price comparison
    """
    
    def __init__(self, projection_years: int = 5):
        """
        Initialize the Exit Multiple DCF builder.
        
        Args:
            projection_years: Number of projection years (default: 5)
        """
        self.projection_years = projection_years
    
    def create_tab(self, workbook: openpyxl.Workbook) -> Worksheet:
        """
        Create and format the Valuation (Exit Multiple DCF) tab.
        
        Args:
            workbook: The Excel workbook to add the tab to
            
        Returns:
            The created worksheet
        """
        # Create or get the sheet
        if "Valuation (Exit Multiple)" in workbook.sheetnames:
            ws = workbook["Valuation (Exit Multiple)"]
            workbook.remove(ws)
        
        ws = workbook.create_sheet("Valuation (Exit Multiple)", 6)  # Position 6 (seventh tab)
        
        # Set up sections following specification
        self._setup_header(ws)
        self._setup_wacc_inputs(ws)
        self._setup_fcf_forecast(ws)
        self._setup_terminal_value(ws)
        self._setup_enterprise_value(ws)
        self._setup_equity_bridge(ws)
        self._setup_value_per_share(ws)
        self._setup_sanity_checks(ws)
        self._setup_output_summary(ws)
        
        # Format the sheet
        self._format_sheet(ws)
        
        return ws
    
    def _setup_header(self, ws: Worksheet) -> None:
        """Set up main header."""
        ws.cell(row=1, column=1, value="VALUATION - EXIT MULTIPLE DCF")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    def _setup_wacc_inputs(self, ws: Worksheet) -> None:
        """
        Set up WACC and Discounting Inputs section (rows 2-5).
        
        Row 2: WACC (from Perpetual Growth DCF tab)
        Row 3: Terminal EV/EBITDA Multiple (from Assumptions)
        Row 4: Tax Rate (from Assumptions)
        Row 5: Discount Periods (Years) - count of projection years
        """
        # Section header (in row 1, column G for notes)
        ws.cell(row=1, column=7, value="A. WACC AND DISCOUNTING INPUTS")
        ws.cell(row=1, column=7).font = Font(bold=True, size=10, italic=True)
        
        # Row 2: WACC
        ws.cell(row=2, column=1, value="WACC")
        ws.cell(row=2, column=1).font = Font(bold=True)
        ws.cell(row=2, column=2, value="='Valuation (DCF)'!B12")  # Link to Perpetual Growth WACC
        ws.cell(row=2, column=2).number_format = '0.00%'
        ws.cell(row=2, column=2).font = Font(bold=True)
        ws.cell(row=2, column=2).fill = PatternFill(
            start_color=ExcelFormats.IMPORTANT_COLOR,
            end_color=ExcelFormats.IMPORTANT_COLOR,
            fill_type="solid"
        )
        
        # Row 3: Terminal EV/EBITDA Multiple
        ws.cell(row=3, column=1, value="Terminal EV/EBITDA Multiple")
        ws.cell(row=3, column=1).font = Font(bold=True)
        # Note: Markdown specifies Assumptions!B12, but we need to add this to Assumptions tab
        # For now, use a default value and add a note
        ws.cell(row=3, column=2, value=20.0)  # Default 20x multiple
        ws.cell(row=3, column=2).number_format = '0.0"x"'
        ws.cell(row=3, column=2).font = Font(bold=True)
        ws.cell(row=3, column=7, value="User input: exit multiple assumption")
        ws.cell(row=3, column=7).font = Font(italic=True, size=9)
        
        # Row 4: Tax Rate
        ws.cell(row=4, column=1, value="Tax Rate")
        ws.cell(row=4, column=2, value='=Assumptions!B20')  # Row 20 in Assumptions
        ws.cell(row=4, column=2).number_format = '0.00%'
        
        # Row 5: Discount Periods
        ws.cell(row=5, column=1, value="Discount Periods (Years)")
        ws.cell(row=5, column=2, value='=COLUMNS(Projections!B2:F2)')  # Count projection years
        ws.cell(row=5, column=2).number_format = '0'
    
    def _setup_fcf_forecast(self, ws: Worksheet) -> None:
        """
        Set up FCF Forecast section (rows 7-10).
        
        Row 7: Free Cash Flow (from Projections)
        Row 8: Discount Factor = 1/(1+WACC)^n
        Row 9: PV of FCF = FCF * Discount Factor
        Row 10: Sum of PV(FCFs)
        """
        # Blank row
        ws.cell(row=6, column=1, value="")
        
        # Row 7: FCF
        ws.cell(row=7, column=1, value="Free Cash Flow (FCF)")
        ws.cell(row=7, column=1).font = Font(bold=True)
        for i in range(self.projection_years):
            col = 2 + i  # B, C, D, E, F
            col_letter = chr(64 + col)
            formula = f'=Projections!{col_letter}19'  # FCF row in Projections
            ws.cell(row=7, column=col, value=formula)
            ws.cell(row=7, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 8: Discount Factor
        ws.cell(row=8, column=1, value="Discount Factor (DF)")
        for i in range(self.projection_years):
            col = 2 + i  # B, C, D, E, F
            period = i + 1  # 1, 2, 3, 4, 5
            # Formula: 1/(1+WACC)^(n - MYD_adjustment)
            # MYD toggle in Sensitivity!$B$4: 0.5 if mid-year, 0 if year-end
            formula = f'=1/(1+$B$2)^({period}-Sensitivity!$B$4)'
            ws.cell(row=8, column=col, value=formula)
            ws.cell(row=8, column=col).number_format = '0.0000'
        
        # Row 9: PV of FCF
        ws.cell(row=9, column=1, value="PV of FCF")
        ws.cell(row=9, column=1).font = Font(bold=True)
        for i in range(self.projection_years):
            col = 2 + i  # B, C, D, E, F
            col_letter = chr(64 + col)
            formula = f'={col_letter}7*{col_letter}8'  # FCF * DF
            ws.cell(row=9, column=col, value=formula)
            ws.cell(row=9, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 10: Sum of PV(FCFs)
        ws.cell(row=10, column=1, value="Sum of PV(FCFs)")
        ws.cell(row=10, column=1).font = Font(bold=True)
        ws.cell(row=10, column=2, value='=SUM(B9:F9)')
        ws.cell(row=10, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=10, column=2).font = Font(bold=True)
    
    def _setup_terminal_value(self, ws: Worksheet) -> None:
        """
        Set up Terminal Value section (rows 12-15).
        
        Row 12: Terminal Year EBITDA (from Projections F21)
        Row 13: Exit Multiple (reference to B3)
        Row 14: Terminal Value (Un-discounted) = EBITDA × Multiple
        Row 15: PV of Terminal Value = TV / (1+WACC)^5
        """
        # Blank row
        ws.cell(row=11, column=1, value="")
        
        # Row 12: Terminal Year EBITDA
        ws.cell(row=12, column=1, value="Terminal Year EBITDA (FY5)")
        ws.cell(row=12, column=1).font = Font(bold=True)
        ws.cell(row=12, column=2, value='=Projections!F21')  # EBITDA in final year (row 21)
        ws.cell(row=12, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=12, column=2).font = Font(bold=True)
        
        # Row 13: Exit Multiple
        ws.cell(row=13, column=1, value="Exit Multiple (EV/EBITDA)")
        ws.cell(row=13, column=2, value='=$B$3')  # Reference exit multiple input
        ws.cell(row=13, column=2).number_format = '0.0"x"'
        
        # Row 14: Terminal Value (Un-discounted)
        ws.cell(row=14, column=1, value="Terminal Value (Un-discounted)")
        ws.cell(row=14, column=2, value='=B12*B13')  # EBITDA × Multiple
        ws.cell(row=14, column=2).number_format = ExcelFormats.CURRENCY
        
        # Row 15: PV of Terminal Value
        ws.cell(row=15, column=1, value="PV of Terminal Value")
        ws.cell(row=15, column=1).font = Font(bold=True)
        ws.cell(row=15, column=2, value='=B14/(1+$B$2)^$B$5')  # TV discounted at WACC for 5 periods
        ws.cell(row=15, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=15, column=2).font = Font(bold=True)
    
    def _setup_enterprise_value(self, ws: Worksheet) -> None:
        """
        Set up Enterprise Value section (row 17).
        
        Row 17: Enterprise Value (EV) = Sum of PV FCFs + PV of TV
        """
        # Blank row
        ws.cell(row=16, column=1, value="")
        
        # Row 17: Enterprise Value
        ws.cell(row=17, column=1, value="Enterprise Value (EV)")
        ws.cell(row=17, column=1).font = Font(bold=True, size=11)
        ws.cell(row=17, column=2, value='=B10+B15')  # Sum of PV FCFs + PV TV
        ws.cell(row=17, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=17, column=2).font = Font(bold=True, size=11)
        ws.cell(row=17, column=2).fill = PatternFill(
            start_color=ExcelFormats.IMPORTANT_COLOR,
            end_color=ExcelFormats.IMPORTANT_COLOR,
            fill_type="solid"
        )
    
    def _setup_equity_bridge(self, ws: Worksheet) -> None:
        """
        Set up Equity Bridge section (rows 19-22).
        
        Row 19: Add: Cash & Equivalents (from Perpetual Growth DCF)
        Row 20: Less: Total Debt (from Perpetual Growth DCF)
        Row 21: Add: Investments (from Perpetual Growth DCF)
        Row 22: Equity Value
        """
        # Blank row
        ws.cell(row=18, column=1, value="")
        
        # Row 19: Cash & Equivalents
        ws.cell(row=19, column=1, value="Add: Cash & Equivalents")
        ws.cell(row=19, column=2, value="='Valuation (DCF)'!B30")  # Link to Perpetual Growth DCF
        ws.cell(row=19, column=2).number_format = ExcelFormats.CURRENCY
        
        # Row 20: Total Debt
        ws.cell(row=20, column=1, value="Less: Total Debt")
        ws.cell(row=20, column=2, value="='Valuation (DCF)'!B31")  # Link to Perpetual Growth DCF
        ws.cell(row=20, column=2).number_format = ExcelFormats.CURRENCY
        
        # Row 21: Investments
        ws.cell(row=21, column=1, value="Add: Investments / Non-operating Assets")
        ws.cell(row=21, column=2, value="='Valuation (DCF)'!B32")  # Link to Perpetual Growth DCF
        ws.cell(row=21, column=2).number_format = ExcelFormats.CURRENCY
        
        # Row 22: Equity Value
        ws.cell(row=22, column=1, value="Equity Value (Firm Value)")
        ws.cell(row=22, column=1).font = Font(bold=True, size=11)
        ws.cell(row=22, column=2, value='=B17+B19-B20+B21')  # EV + Cash - Debt + Investments
        ws.cell(row=22, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=22, column=2).font = Font(bold=True, size=11)
        ws.cell(row=22, column=2).fill = PatternFill(
            start_color=ExcelFormats.IMPORTANT_COLOR,
            end_color=ExcelFormats.IMPORTANT_COLOR,
            fill_type="solid"
        )
    
    def _setup_value_per_share(self, ws: Worksheet) -> None:
        """
        Set up Value per Share section (rows 24-25).
        
        Row 24: Shares Outstanding (from Perpetual Growth DCF)
        Row 25: Intrinsic Value per Share
        """
        # Blank row
        ws.cell(row=23, column=1, value="")
        
        # Row 24: Shares Outstanding
        ws.cell(row=24, column=1, value="Shares Outstanding (absolute count)")
        ws.cell(row=24, column=2, value="='Valuation (DCF)'!B36")  # Link to Perpetual Growth DCF
        ws.cell(row=24, column=2).number_format = '#,##0'
        
        # Row 25: Intrinsic Value per Share
        ws.cell(row=25, column=1, value="Intrinsic Value per Share ($)")
        ws.cell(row=25, column=1).font = Font(bold=True, size=11)
        ws.cell(row=25, column=2, value='=B22/B24')  # Equity Value / Shares
        ws.cell(row=25, column=2).number_format = '$0.00'
        ws.cell(row=25, column=2).font = Font(bold=True, size=11)
        ws.cell(row=25, column=2).fill = PatternFill(
            start_color=ExcelFormats.IMPORTANT_COLOR,
            end_color=ExcelFormats.IMPORTANT_COLOR,
            fill_type="solid"
        )
    
    def _setup_sanity_checks(self, ws: Worksheet) -> None:
        """
        Set up Sanity Checks section (rows 27-30).
        
        Row 27: Implied EV/EBITDA (FY5)
        Row 28: EV/EBIT (FY5)
        Row 29: FCF Yield (EV Basis)
        Row 30: Equity Value from Perpetual Growth DCF (cross-check)
        """
        # Blank row
        ws.cell(row=26, column=1, value="")
        
        # Row 27: Implied EV/EBITDA
        ws.cell(row=27, column=1, value="Implied EV/EBITDA (FY5)")
        ws.cell(row=27, column=2, value='=B17/Projections!F21')  # EV / EBITDA
        ws.cell(row=27, column=2).number_format = '0.0"x"'
        
        # Row 28: EV/EBIT
        ws.cell(row=28, column=1, value="Implied EV/EBIT (FY5)")
        ws.cell(row=28, column=2, value='=B17/Projections!F9')  # EV / EBIT (row 9 in Projections)
        ws.cell(row=28, column=2).number_format = '0.0"x"'
        
        # Row 29: FCF Yield
        ws.cell(row=29, column=1, value="FCF Yield (EV Basis, FY5)")
        ws.cell(row=29, column=2, value='=Projections!F19/B17')  # FCF / EV
        ws.cell(row=29, column=2).number_format = '0.00%'
        
        # Row 30: Cross-reference to Perpetual Growth Equity Value
        ws.cell(row=30, column=1, value="Equity Value (Perpetual Growth DCF)")
        ws.cell(row=30, column=1).font = Font(italic=True)
        ws.cell(row=30, column=2, value="='Valuation (DCF)'!B33")  # Perpetual Growth equity value
        ws.cell(row=30, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=30, column=2).font = Font(italic=True)
    
    def _setup_output_summary(self, ws: Worksheet) -> None:
        """
        Set up Output Summary section (rows 32-35).
        
        Row 32: Enterprise Value (Exit Multiple DCF)
        Row 33: Equity Value (Exit Multiple DCF)
        Row 34: Intrinsic Value per Share (Exit Multiple DCF)
        Row 35: Implied Upside vs Current Price
        """
        # Blank row
        ws.cell(row=31, column=1, value="")
        
        # Row 32: Enterprise Value
        ws.cell(row=32, column=1, value="Enterprise Value (Exit Multiple DCF)")
        ws.cell(row=32, column=1).font = Font(bold=True)
        ws.cell(row=32, column=2, value='=B17')
        ws.cell(row=32, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=32, column=2).font = Font(bold=True)
        
        # Row 33: Equity Value
        ws.cell(row=33, column=1, value="Equity Value (Exit Multiple DCF)")
        ws.cell(row=33, column=1).font = Font(bold=True)
        ws.cell(row=33, column=2, value='=B22')
        ws.cell(row=33, column=2).number_format = ExcelFormats.CURRENCY
        ws.cell(row=33, column=2).font = Font(bold=True)
        
        # Row 34: Intrinsic Value per Share
        ws.cell(row=34, column=1, value="Intrinsic Value per Share (Exit Multiple DCF)")
        ws.cell(row=34, column=1).font = Font(bold=True, size=11)
        ws.cell(row=34, column=2, value='=B25')
        ws.cell(row=34, column=2).number_format = '$0.00'
        ws.cell(row=34, column=2).font = Font(bold=True, size=11)
        ws.cell(row=34, column=2).fill = PatternFill(
            start_color="FFD700",  # Gold color
            end_color="FFD700",
            fill_type="solid"
        )
        
        # Row 35: Implied Upside vs Current Price
        ws.cell(row=35, column=1, value="Implied Upside vs Current Price")
        ws.cell(row=35, column=1).font = Font(italic=True)
        # Note: Historical!F2 should contain current stock price
        # If not available, this will show #REF! until price is added
        ws.cell(row=35, column=2, value='=B34/Historical!F2-1')  # (Intrinsic / Current) - 1
        ws.cell(row=35, column=2).number_format = '0.0%'
        ws.cell(row=35, column=2).font = Font(italic=True)
    
    def _format_sheet(self, ws: Worksheet) -> None:
        """Apply general formatting to the sheet."""
        # Set column widths
        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 18
        for col_letter in ['C', 'D', 'E', 'F']:
            ws.column_dimensions[col_letter].width = 14
        ws.column_dimensions['G'].width = 35
        
        # Freeze headers
        ws.freeze_panes = ws['B2']
