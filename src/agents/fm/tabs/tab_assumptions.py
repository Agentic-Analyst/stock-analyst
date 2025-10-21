"""
Tab 3: Assumptions Tab Builder with Excel Formulas

This module creates the Assumptions tab with embedded Excel formulas that
reference the Raw tab for calculations. LLM is used to infer forward projections.

Key Design:
- FY0 values use formulas referencing Raw tab
- FY1-FY5 values use formulas referencing hidden LLM_Inferred tab
- LLM integration uses proper LLMProvider from llms.config
"""

from typing import Dict, Optional, Any
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
import json
from pathlib import Path
import sys

# Add parent directories to path for imports
current_dir = Path(__file__).resolve().parent
sys.path.insert(0, str(current_dir.parent.parent.parent))  # Add 'src' to path

from llms.config import get_llm


class AssumptionsTabBuilder:
    """
    Builds the Assumptions tab with Excel formulas referencing Raw tab.
    
    Layout:
    - Column A: Labels
    - Column B: FY0 (latest actual, calculated from Raw using formulas)
    - Columns C-G: FY1-FY5 (formulas referencing LLM_Inferred tab)
    """
    
    def __init__(self, llm_assumptions: Optional[Dict[str, Any]] = None):
        """Initialize with optional LLM-inferred assumptions."""
        self.llm_assumptions = llm_assumptions or {}
    
    def create_tab(self, workbook: openpyxl.Workbook) -> Worksheet:
        """Create and format the Assumptions tab with Excel formulas."""
        # Remove existing tabs if they exist
        if "Assumptions" in workbook.sheetnames:
            ws = workbook["Assumptions"]
            workbook.remove(ws)
        if "LLM_Inferred" in workbook.sheetnames:
            ws = workbook["LLM_Inferred"]
            workbook.remove(ws)
        
        # Create hidden LLM_Inferred tab first
        self._create_llm_inferred_tab(workbook)
        
        # Create Assumptions tab
        ws = workbook.create_sheet("Assumptions", 2)
        
        self._setup_headers(ws)
        self._setup_fy0_year(ws)
        self._setup_valuation_params(ws)
        self._setup_revenue_growth(ws)
        self._setup_operating_margins(ws)
        self._setup_working_capital(ws)
        self._setup_capital_structure(ws)
        self._setup_dcf_parameters(ws)
        self._format_sheet(ws)
        
        return ws
    
    def _create_llm_inferred_tab(self, workbook: openpyxl.Workbook) -> None:
        """Create hidden tab with LLM-inferred values."""
        ws = workbook.create_sheet("LLM_Inferred", 3)
        
        # Headers
        ws.cell(row=1, column=1, value="Metric").font = Font(bold=True)
        for i in range(5):
            ws.cell(row=1, column=2 + i, value=f"FY{i+1}").font = Font(bold=True)
        
        # WACC and Terminal Growth (same for all years, but we'll put in B column)
        ws.cell(row=2, column=1, value="WACC")
        ws.cell(row=2, column=2, value=self.llm_assumptions.get('wacc', 0.09))
        
        ws.cell(row=3, column=1, value="Terminal Growth Rate")
        ws.cell(row=3, column=2, value=self.llm_assumptions.get('terminal_growth_rate', 0.025))
        
        # Revenue Growth Rates (FY1-FY5)
        ws.cell(row=4, column=1, value="Revenue Growth Rate")
        rates = self.llm_assumptions.get('revenue_growth_rates', [0.05] * 5)
        for i, rate in enumerate(rates):
            ws.cell(row=4, column=2 + i, value=rate)
        
        # Gross Margins (FY1-FY5)
        ws.cell(row=5, column=1, value="Gross Margin")
        margins = self.llm_assumptions.get('gross_margins', [0.46] * 5)
        for i, margin in enumerate(margins):
            ws.cell(row=5, column=2 + i, value=margin)
        
        # EBITDA Margins (FY1-FY5)
        ws.cell(row=6, column=1, value="EBITDA Margin")
        ebitda = self.llm_assumptions.get('ebitda_margins', [0.33] * 5)
        for i, margin in enumerate(ebitda):
            ws.cell(row=6, column=2 + i, value=margin)
        
        # Operating Margins (FY1-FY5)
        ws.cell(row=7, column=1, value="Operating Margin")
        operating = self.llm_assumptions.get('operating_margins', [0.31] * 5)
        for i, margin in enumerate(operating):
            ws.cell(row=7, column=2 + i, value=margin)
        
        # DSO Days (FY1-FY5)
        ws.cell(row=8, column=1, value="DSO Days")
        dso = self.llm_assumptions.get('dso_days', [45] * 5)
        for i, days in enumerate(dso):
            ws.cell(row=8, column=2 + i, value=days)
        
        # DIO Days (FY1-FY5)
        ws.cell(row=9, column=1, value="DIO Days")
        dio = self.llm_assumptions.get('dio_days', [10] * 5)
        for i, days in enumerate(dio):
            ws.cell(row=9, column=2 + i, value=days)
        
        # DPO Days (FY1-FY5)
        ws.cell(row=10, column=1, value="DPO Days")
        dpo = self.llm_assumptions.get('dpo_days', [90] * 5)
        for i, days in enumerate(dpo):
            ws.cell(row=10, column=2 + i, value=days)
        
        # Make this tab visible (not hidden) for transparency
        # ws.sheet_state = 'hidden'  # User requested this be visible
    
    def _setup_headers(self, ws: Worksheet) -> None:
        """Set up column headers."""
        ws.cell(row=1, column=1, value="Metric").font = Font(bold=True)
        ws.cell(row=1, column=2, value="FY0 (Actual)").font = Font(bold=True)
        
        for i in range(5):
            col = 3 + i
            ws.cell(row=1, column=col, value=f"FY{i+1}").font = Font(bold=True)
    
    def _setup_fy0_year(self, ws: Worksheet) -> None:
        """Set up FY0 year extraction using simpler approach."""
        ws.cell(row=2, column=1, value="Latest Fiscal Year (FY0)").font = Font(bold=True, italic=True)
        # Use INDEX/MATCH to find the latest year from Raw tab
        # Simpler formula: just show latest year as a number
        ws.cell(row=2, column=2, value='=VALUE(LEFT(INDEX(Raw!$C:$C,MATCH("Total Revenue",Raw!$B:$B,0)),4))').number_format = '0'
    
    def _setup_valuation_params(self, ws: Worksheet) -> None:
        """Set up valuation parameters."""
        ws.cell(row=3, column=1, value="VALUATION PARAMETERS").font = Font(bold=True, size=11)
        
        # WACC
        ws.cell(row=4, column=1, value="WACC").font = Font(bold=True)
        ws.cell(row=4, column=2, value='=LLM_Inferred!B2').number_format = '0.00%'
        ws.cell(row=4, column=3, value="[LLM]").font = Font(italic=True, size=9)
        
        # Terminal Growth Rate
        ws.cell(row=5, column=1, value="Terminal Growth Rate (g)").font = Font(bold=True)
        ws.cell(row=5, column=2, value='=LLM_Inferred!B3').number_format = '0.00%'
        ws.cell(row=5, column=3, value="[LLM]").font = Font(italic=True, size=9)
    
    def _setup_revenue_growth(self, ws: Worksheet) -> None:
        """Set up revenue growth with formulas."""
        ws.cell(row=6, column=1, value="REVENUE GROWTH ASSUMPTIONS").font = Font(bold=True, size=11)
        
        ws.cell(row=7, column=1, value="Revenue Growth (YoY)").font = Font(bold=True)
        
        # FY0: Calculate from Raw tab - use INDIRECT to avoid complex date matching
        # Simplified: Use helper cells or direct reference
        # For now, use a simpler SUMIFS approach
        formula_fy0 = (
            '=IFERROR('
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Total Revenue",Raw!$C:$C,$B$2&"*")/'
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Total Revenue",Raw!$C:$C,($B$2-1)&"*")-1,'
            '"")'
        )
        ws.cell(row=7, column=2, value=formula_fy0).number_format = '0.00%'
        
        # FY1-FY5: Reference LLM_Inferred tab (columns B-F for FY1-FY5)
        for i in range(5):
            col_letter = chr(66 + i)  # B, C, D, E, F (FY1-FY5 in LLM_Inferred)
            # Use simple IFERROR for compatibility with all Excel versions
            if i == 0:
                # FY1: fallback to FY0 (B7) if blank
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}4,"")<>"",LLM_Inferred!{col_letter}4,B7)'
            else:
                # FY2-FY5: fallback to previous FY
                prev_col = chr(66 + i)  # C, D, E, F (previous column in Assumptions tab)
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}4,"")<>"",LLM_Inferred!{col_letter}4,{prev_col}7)'
            ws.cell(row=7, column=3 + i, value=formula).number_format = '0.00%'
    
    def _setup_operating_margins(self, ws: Worksheet) -> None:
        """Set up operating margins with formulas."""
        ws.cell(row=8, column=1, value="OPERATING MARGIN ASSUMPTIONS").font = Font(bold=True, size=11)
        
        # Gross Margin
        ws.cell(row=9, column=1, value="Gross Margin").font = Font(bold=True)
        formula_fy0 = (
            '=IFERROR('
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Gross Profit",Raw!$C:$C,$B$2&"*")/'
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Total Revenue",Raw!$C:$C,$B$2&"*"),'
            '"")'
        )
        ws.cell(row=9, column=2, value=formula_fy0).number_format = '0.00%'
        
        # FY1-FY5: Reference LLM_Inferred (columns B-F)
        for i in range(5):
            col_letter = chr(66 + i)  # B, C, D, E, F
            if i == 0:
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}5,"")<>"",LLM_Inferred!{col_letter}5,B9)'
            else:
                prev_col = chr(66 + i)  # C, D, E, F
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}5,"")<>"",LLM_Inferred!{col_letter}5,{prev_col}9)'
            ws.cell(row=9, column=3 + i, value=formula).number_format = '0.00%'
        
        # EBITDA Margin
        ws.cell(row=10, column=1, value="EBITDA Margin").font = Font(bold=True)
        formula_fy0 = (
            '=IFERROR('
            '(SUMIFS(Raw!$D:$D,Raw!$B:$B,"Operating Income",Raw!$C:$C,$B$2&"*")+'
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Depreciation And Amortization",Raw!$C:$C,$B$2&"*"))/'
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Total Revenue",Raw!$C:$C,$B$2&"*"),'
            '"")'
        )
        ws.cell(row=10, column=2, value=formula_fy0).number_format = '0.00%'
        
        # FY1-FY5: Reference LLM_Inferred (columns B-F)
        for i in range(5):
            col_letter = chr(66 + i)  # B, C, D, E, F
            if i == 0:
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}6,"")<>"",LLM_Inferred!{col_letter}6,B10)'
            else:
                prev_col = chr(66 + i)  # C, D, E, F
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}6,"")<>"",LLM_Inferred!{col_letter}6,{prev_col}10)'
            ws.cell(row=10, column=3 + i, value=formula).number_format = '0.00%'
        
        # Operating Margin
        ws.cell(row=11, column=1, value="Operating Margin").font = Font(bold=True)
        formula_fy0 = (
            '=IFERROR('
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Operating Income",Raw!$C:$C,$B$2&"*")/'
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Total Revenue",Raw!$C:$C,$B$2&"*"),'
            '"")'
        )
        ws.cell(row=11, column=2, value=formula_fy0).number_format = '0.00%'
        
        # FY1-FY5: Reference LLM_Inferred (columns B-F)
        for i in range(5):
            col_letter = chr(66 + i)  # B, C, D, E, F
            if i == 0:
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}7,"")<>"",LLM_Inferred!{col_letter}7,B11)'
            else:
                prev_col = chr(66 + i)  # C, D, E, F
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}7,"")<>"",LLM_Inferred!{col_letter}7,{prev_col}11)'
            ws.cell(row=11, column=3 + i, value=formula).number_format = '0.00%'
    
    def _setup_working_capital(self, ws: Worksheet) -> None:
        """Set up working capital with formulas."""
        ws.cell(row=12, column=1, value="WORKING CAPITAL ASSUMPTIONS").font = Font(bold=True, size=11)
        
        # DSO
        ws.cell(row=13, column=1, value="DSO (Days)").font = Font(bold=True)
        formula_fy0 = (
            '=IFERROR('
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Accounts Receivable",Raw!$C:$C,$B$2&"*")/'
            '(SUMIFS(Raw!$D:$D,Raw!$B:$B,"Total Revenue",Raw!$C:$C,$B$2&"*")/365),'
            '"")'
        )
        ws.cell(row=13, column=2, value=formula_fy0).number_format = '0.0'
        
        # FY1-FY5: Reference LLM_Inferred (columns B-F)
        for i in range(5):
            col_letter = chr(66 + i)  # B, C, D, E, F
            if i == 0:
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}8,"")<>"",LLM_Inferred!{col_letter}8,B13)'
            else:
                prev_col = chr(66 + i)  # C, D, E, F
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}8,"")<>"",LLM_Inferred!{col_letter}8,{prev_col}13)'
            ws.cell(row=13, column=3 + i, value=formula).number_format = '0.0'
        
        # DIO
        ws.cell(row=14, column=1, value="DIO (Days)").font = Font(bold=True)
        formula_fy0 = (
            '=IFERROR('
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Inventory",Raw!$C:$C,$B$2&"*")/'
            '(SUMIFS(Raw!$D:$D,Raw!$B:$B,"Cost Of Revenue",Raw!$C:$C,$B$2&"*")/365),'
            '"")'
        )
        ws.cell(row=14, column=2, value=formula_fy0).number_format = '0.0'
        
        # FY1-FY5: Reference LLM_Inferred (columns B-F)
        for i in range(5):
            col_letter = chr(66 + i)  # B, C, D, E, F
            if i == 0:
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}9,"")<>"",LLM_Inferred!{col_letter}9,B14)'
            else:
                prev_col = chr(66 + i)  # C, D, E, F
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}9,"")<>"",LLM_Inferred!{col_letter}9,{prev_col}14)'
            ws.cell(row=14, column=3 + i, value=formula).number_format = '0.0'
        
        # DPO
        ws.cell(row=15, column=1, value="DPO (Days)").font = Font(bold=True)
        formula_fy0 = (
            '=IFERROR('
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Accounts Payable",Raw!$C:$C,$B$2&"*")/'
            '(SUMIFS(Raw!$D:$D,Raw!$B:$B,"Cost Of Revenue",Raw!$C:$C,$B$2&"*")/365),'
            '"")'
        )
        ws.cell(row=15, column=2, value=formula_fy0).number_format = '0.0'
        
        # FY1-FY5: Reference LLM_Inferred (columns B-F)
        for i in range(5):
            col_letter = chr(66 + i)  # B, C, D, E, F
            if i == 0:
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}10,"")<>"",LLM_Inferred!{col_letter}10,B15)'
            else:
                prev_col = chr(66 + i)  # C, D, E, F
                formula = f'=IF(IFERROR(LLM_Inferred!{col_letter}10,"")<>"",LLM_Inferred!{col_letter}10,{prev_col}15)'
            ws.cell(row=15, column=3 + i, value=formula).number_format = '0.0'
        
        # CCC - ALL columns use formulas
        ws.cell(row=16, column=1, value="Cash Conversion Cycle (Days)").font = Font(bold=True)
        ws.cell(row=16, column=2, value='=IFERROR(B13+B14-B15,"")').number_format = '0.0'
        
        for i in range(5):
            col_letter = chr(67 + i)
            ws.cell(row=16, column=3 + i, value=f'=IFERROR({col_letter}13+{col_letter}14-{col_letter}15,"")').number_format = '0.0'
    
    def _setup_capital_structure(self, ws: Worksheet) -> None:
        """Set up capital structure."""
        ws.cell(row=17, column=1, value="CAPITAL STRUCTURE").font = Font(bold=True, size=11)
        
        # Shares Outstanding
        ws.cell(row=18, column=1, value="Shares Outstanding (latest)").font = Font(bold=True)
        ws.cell(row=18, column=2, value='=SUMIFS(Raw!$D:$D,Raw!$B:$B,"Diluted Average Shares",Raw!$C:$C,$B$2&"*")').number_format = '#,##0'
        ws.cell(row=18, column=3, value="[From JSON]").font = Font(italic=True, size=9)
        
        # Net Debt
        ws.cell(row=19, column=1, value="Net Debt (latest)").font = Font(bold=True)
        formula = (
            '=SUMIFS(Raw!$D:$D,Raw!$B:$B,"Total Debt",Raw!$C:$C,$B$2&"*")-'
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Cash And Cash Equivalents",Raw!$C:$C,$B$2&"*")'
        )
        ws.cell(row=19, column=2, value=formula).number_format = '#,##0'
        ws.cell(row=19, column=3, value="[From JSON]").font = Font(italic=True, size=9)
        
        # Effective Tax Rate
        ws.cell(row=20, column=1, value="Effective Tax Rate (FY0)").font = Font(bold=True)
        formula = (
            '=IFERROR('
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Tax Provision",Raw!$C:$C,$B$2&"*")/'
            'SUMIFS(Raw!$D:$D,Raw!$B:$B,"Pretax Income",Raw!$C:$C,$B$2&"*"),'
            '"")'
        )
        ws.cell(row=20, column=2, value=formula).number_format = '0.00%'
        ws.cell(row=20, column=3, value="[From JSON]").font = Font(italic=True, size=9)
    
    def _setup_dcf_parameters(self, ws: Worksheet) -> None:
        """Set up DCF valuation parameters (rows 21-36)."""
        # Section header
        ws.cell(row=21, column=1, value="DCF VALUATION PARAMETERS").font = Font(bold=True, size=11)
        
        # Subsection: Cost of Equity Inputs
        ws.cell(row=22, column=1, value="Cost of Equity Inputs:")
        ws.cell(row=22, column=1).font = Font(bold=True, italic=True, size=10)
        
        # Risk-Free Rate (row 30 in markdown, but we'll use row 23 here)
        ws.cell(row=23, column=1, value="Risk-Free Rate (Rf)").font = Font(bold=True)
        ws.cell(row=23, column=2, value=0.045).number_format = '0.00%'  # Default 4.5%
        ws.cell(row=23, column=3, value="[10Y Treasury]").font = Font(italic=True, size=9)
        
        # Equity Risk Premium (row 31 in markdown, row 24 here)
        ws.cell(row=24, column=1, value="Equity Risk Premium (ERP)").font = Font(bold=True)
        ws.cell(row=24, column=2, value=0.065).number_format = '0.00%'  # Default 6.5%
        ws.cell(row=24, column=3, value="[Historical ERP]").font = Font(italic=True, size=9)
        
        # Levered Beta (row 32 in markdown, row 25 here)
        ws.cell(row=25, column=1, value="Levered Beta (β)").font = Font(bold=True)
        ws.cell(row=25, column=2, value=1.2).number_format = '0.00'  # Default 1.2
        ws.cell(row=25, column=3, value="[From Bloomberg/Yahoo]").font = Font(italic=True, size=9)
        
        # Subsection: Cost of Debt Inputs
        ws.cell(row=26, column=1, value="Cost of Debt Inputs:").font = Font(bold=True, italic=True, size=10)
        
        # Pre-Tax Cost of Debt (row 33 in markdown, row 27 here)
        ws.cell(row=27, column=1, value="Pre-Tax Cost of Debt (Kd)").font = Font(bold=True)
        ws.cell(row=27, column=2, value=0.055).number_format = '0.00%'  # Default 5.5%
        ws.cell(row=27, column=3, value="[Corporate bonds yield]").font = Font(italic=True, size=9)
        
        # Subsection: Capital Structure
        ws.cell(row=28, column=1, value="Capital Structure Weights:").font = Font(bold=True, italic=True, size=10)
        
        # Equity Weight (row 34 in markdown, row 29 here)
        ws.cell(row=29, column=1, value="Equity Weight (E/V)").font = Font(bold=True)
        ws.cell(row=29, column=2, value=0.85).number_format = '0.00%'  # Default 85%
        ws.cell(row=29, column=3, value="[Target capital structure]").font = Font(italic=True, size=9)
        
        # Terminal Growth Rate (row 35 in markdown, row 30 here)
        ws.cell(row=30, column=1, value="Terminal Growth Rate (g)").font = Font(bold=True)
        ws.cell(row=30, column=2, value='=LLM_Inferred!B3').number_format = '0.00%'  # From LLM
        ws.cell(row=30, column=3, value="[LLM]").font = Font(italic=True, size=9)
        
        # Shares Outstanding (row 36 in markdown, row 31 here)
        ws.cell(row=31, column=1, value="Shares Outstanding (for valuation)").font = Font(bold=True)
        ws.cell(row=31, column=2, value='=B18').number_format = '#,##0'  # Reference row 18
        ws.cell(row=31, column=3, value="[From row 18]").font = Font(italic=True, size=9)
    
    def _format_sheet(self, ws: Worksheet) -> None:
        """Apply formatting."""
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        for col in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 12
        ws.freeze_panes = ws['A2']
    
    def get_summary(self) -> Dict[str, Any]:
        """Get summary."""
        return {
            'wacc': self.llm_assumptions.get('wacc', 0.09),
            'terminal_growth_rate': self.llm_assumptions.get('terminal_growth_rate', 0.025),
            'projection_years': 5
        }


def infer_assumptions_with_llm(json_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Use LLM to infer forward assumptions based on historical data.
    
    Args:
        json_data: The financial JSON data from Raw tab
        
    Returns:
        Dict with inferred assumptions
    """
    # Use proper LLM provider
    llm = get_llm()
    
    # Load prompt template
    current_file = Path(__file__).resolve()
    src_dir = current_file.parent.parent.parent.parent
    project_root = src_dir.parent
    prompt_path = project_root / "prompts" / "assumptions_inference.md"
    
    if not prompt_path.exists():
        raise FileNotFoundError(f"Prompt file not found: {prompt_path}")
    
    with open(prompt_path, 'r') as f:
        prompt_template = f.read()
    
    # Extract historical metrics from JSON
    fs = json_data.get('financial_statements', {})
    company_data = json_data.get('company_data', {})
    
    # Get latest year
    years = sorted([y for y in fs.get('income_statement', {}).keys() if y != 'date'])
    latest_year = years[-1] if years else '2024-09-30'
    prev_year = years[-2] if len(years) > 1 else None
    
    # Extract metrics
    latest_income = fs.get('income_statement', {}).get(latest_year, {})
    prev_income = fs.get('income_statement', {}).get(prev_year, {}) if prev_year else {}
    latest_balance = fs.get('balance_sheet', {}).get(latest_year, {})
    
    # Calculate FY0 metrics
    revenue_fy0 = latest_income.get('Total Revenue', 0)
    revenue_prev = prev_income.get('Total Revenue', 0)
    revenue_growth_fy0 = ((revenue_fy0 / revenue_prev - 1) * 100) if revenue_prev else 0
    
    gross_profit = latest_income.get('Gross Profit', 0)
    gross_margin_fy0 = (gross_profit / revenue_fy0 * 100) if revenue_fy0 else 0
    
    operating_income = latest_income.get('Operating Income', 0)
    operating_margin_fy0 = (operating_income / revenue_fy0 * 100) if revenue_fy0 else 0
    
    da = latest_income.get('Depreciation And Amortization', 0)
    ebitda = operating_income + da
    ebitda_margin_fy0 = (ebitda / revenue_fy0 * 100) if revenue_fy0 else 0
    
    # Working capital metrics
    ar = latest_balance.get('Accounts Receivable', 0)
    dso_fy0 = (ar / revenue_fy0 * 365) if revenue_fy0 else 45
    
    inventory = latest_balance.get('Inventory', 0)
    cogs = latest_income.get('Cost Of Revenue', 0)
    dio_fy0 = (inventory / cogs * 365) if cogs else 10
    
    ap = latest_balance.get('Accounts Payable', 0)
    dpo_fy0 = (ap / cogs * 365) if cogs else 90
    
    # Tax rate
    tax_provision = latest_income.get('Tax Provision', 0)
    pretax_income = latest_income.get('Pretax Income', 0)
    tax_rate_fy0 = (tax_provision / pretax_income * 100) if pretax_income else 21
    
    # Company info
    ticker = company_data.get('basic_info', {}).get('symbol', 'UNKNOWN')
    company_name = company_data.get('basic_info', {}).get('long_name', 'Unknown Company')
    sector = company_data.get('basic_info', {}).get('sector', 'Unknown')
    
    # Fill prompt
    prompt = prompt_template.format(
        company_name=company_name,
        ticker=ticker,
        sector=sector,
        latest_fy=latest_year[:4],
        revenue_growth_fy0=f"{revenue_growth_fy0:.1f}",
        gross_margin_fy0=f"{gross_margin_fy0:.1f}",
        ebitda_margin_fy0=f"{ebitda_margin_fy0:.1f}",
        operating_margin_fy0=f"{operating_margin_fy0:.1f}",
        dso_fy0=f"{dso_fy0:.1f}",
        dio_fy0=f"{dio_fy0:.1f}",
        dpo_fy0=f"{dpo_fy0:.1f}",
        tax_rate_fy0=f"{tax_rate_fy0:.1f}"
    )
    
    # Call LLM using proper provider
    print("\n🤖 Calling LLM to infer assumptions...")
    messages = [{"role": "user", "content": prompt}]
    response, cost = llm(messages, temperature=0.3)
    print(f"   💰 LLM cost: ${cost:.4f}")
    
    # Parse JSON response
    try:
        # Clean response if it has markdown code blocks
        response_clean = response.strip()
        if response_clean.startswith('```'):
            # Remove ```json or ``` markers
            lines = response_clean.split('\n')
            response_clean = '\n'.join(lines[1:-1]) if len(lines) > 2 else response_clean
        
        assumptions = json.loads(response_clean)
        print(f"   ✅ LLM inference successful")
        return assumptions
    except json.JSONDecodeError as e:
        # Fallback to defaults if parsing fails
        print(f"   ⚠️  Failed to parse LLM response: {e}")
        print(f"   Using default assumptions based on FY0")
        return {
            'wacc': 0.09,
            'terminal_growth_rate': 0.025,
            'revenue_growth_rates': [0.05, 0.05, 0.04, 0.04, 0.03],
            'gross_margins': [gross_margin_fy0/100] * 5 if gross_margin_fy0 else [0.46] * 5,
            'ebitda_margins': [ebitda_margin_fy0/100] * 5 if ebitda_margin_fy0 else [0.33] * 5,
            'operating_margins': [operating_margin_fy0/100] * 5 if operating_margin_fy0 else [0.31] * 5,
            'dso_days': [dso_fy0] * 5 if dso_fy0 else [45] * 5,
            'dio_days': [dio_fy0] * 5 if dio_fy0 else [10] * 5,
            'dpo_days': [dpo_fy0] * 5 if dpo_fy0 else [90] * 5
        }
