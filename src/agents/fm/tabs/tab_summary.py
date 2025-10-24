"""
Tab 7: Summary Tab Builder

Professional banker-grade Summary tab with comprehensive valuation dashboard.

This module creates the Summary tab following investment banking standards:
- Key snapshot & assumptions (WACC, g, shares, current price)
- Perpetual Growth DCF outputs (EV, Equity Value, Value/Share)
- Exit Multiple DCF outputs (EV, Equity Value, Value/Share)
- Blended valuation & market comparison (average, upside, premium)
- Sanity metrics (terminal year revenue, EBITDA, FCF, multiples, yields)
- Quality assurance flags (WACC > g, DFs ≤ 1, shares > 0, MYD toggle wired)

All formulas reference the correct cells from existing tabs:
- Valuation (DCF): B12 (WACC), B23 (g), B27 (EV), B30 (Cash), B31 (Debt), 
                   B32 (Investments), B33 (Equity Value), B36 (Shares), B37 (Value/Share)
- Valuation (Exit Multiple): B2 (WACC), B3 (Exit Multiple), B17 (EV), 
                             B22 (Equity Value), B25 (Value/Share)
- Projections: F3 (Revenue FY5), F21 (EBITDA FY5), F19 (FCF FY5)
- Sensitivity: B2 (MYD Toggle)
- Historical: F2 (Current Stock Price)
"""

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..financial_model_builder import ExcelFormats


class SummaryTabBuilder:
    """
    Builds the Summary tab - comprehensive valuation dashboard.
    
    This tab provides a clean, professional summary of the entire valuation
    with all key metrics, checks, and comparisons in one place.
    """
    
    def __init__(self):
        """Initialize the Summary builder."""
        pass
    
    def create_tab(self, workbook: openpyxl.Workbook) -> Worksheet:
        """
        Create and format the Summary tab.
        
        Args:
            workbook: The Excel workbook to add the tab to
            
        Returns:
            The created worksheet
        """
        # Create or get the sheet
        if "Summary" in workbook.sheetnames:
            ws = workbook["Summary"]
            workbook.remove(ws)
        
        ws = workbook.create_sheet("Summary", 9)  # Position 9 (tenth tab)
        
        # Set up sections following banker's specification
        self._setup_header(ws)
        self._setup_key_snapshot(ws)
        self._setup_perpetual_dcf_outputs(ws)
        self._setup_exit_multiple_outputs(ws)
        self._setup_blended_valuation(ws)
        self._setup_sanity_metrics(ws)
        self._setup_qa_flags(ws)
        
        # Format the sheet
        self._format_sheet(ws)
        
        return ws
    
    def _setup_header(self, ws: Worksheet) -> None:
        """Set up header (row 1)."""
        ws.cell(row=1, column=1, value="SUMMARY — DUAL DCF (PERPETUAL & EXIT MULTIPLE)")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    def _setup_key_snapshot(self, ws: Worksheet) -> None:
        """
        Set up Key Snapshot & Assumptions section (rows 3-10).
        
        Rows:
        3: Mid-Year Discount Toggle = Sensitivity!B2
        4: WACC (Perpetual DCF) = Valuation (DCF)!B12
        5: Terminal Growth g = Valuation (DCF)!B23
        6: WACC (Exit Multiple DCF) = Valuation (Exit Multiple)!B2
        7: Exit Multiple (EV/EBITDA) = Valuation (Exit Multiple)!B3
        8: Shares Outstanding = Valuation (DCF)!B36
        9: Current Market Price = Historical!F2
        10: Market Capitalization = B8*B9
        """
        # Blank row
        ws.cell(row=2, column=1, value="")
        
        # Row 3: Mid-Year Discount Toggle
        ws.cell(row=3, column=1, value="Mid-Year Discount Toggle")
        ws.cell(row=3, column=2, value="='Sensitivity'!$B$2")
        ws.cell(row=3, column=2).alignment = Alignment(horizontal="center")
        
        # Row 4: WACC (Perpetual DCF)
        ws.cell(row=4, column=1, value="WACC (Perpetual DCF)")
        ws.cell(row=4, column=2, value="='Valuation (DCF)'!$B$12")
        ws.cell(row=4, column=2).number_format = '0.00%'
        
        # Row 5: Terminal Growth g
        ws.cell(row=5, column=1, value="Terminal Growth g")
        ws.cell(row=5, column=2, value="='Valuation (DCF)'!$B$23")
        ws.cell(row=5, column=2).number_format = '0.00%'
        
        # Row 6: WACC (Exit Multiple DCF)
        ws.cell(row=6, column=1, value="WACC (Exit Multiple DCF)")
        ws.cell(row=6, column=2, value="='Valuation (Exit Multiple)'!$B$2")
        ws.cell(row=6, column=2).number_format = '0.00%'
        
        # Row 7: Exit Multiple
        ws.cell(row=7, column=1, value="Exit Multiple (EV/EBITDA)")
        ws.cell(row=7, column=2, value="='Valuation (Exit Multiple)'!$B$3")
        ws.cell(row=7, column=2).number_format = '0.0"x"'
        
        # Row 8: Shares Outstanding
        ws.cell(row=8, column=1, value="Shares Outstanding (diluted)")
        ws.cell(row=8, column=2, value="='Valuation (DCF)'!$B$36")
        ws.cell(row=8, column=2).number_format = '#,##0'
        
        # Row 9: Current Market Price
        ws.cell(row=9, column=1, value="Current Market Price")
        ws.cell(row=9, column=2, value="='Historical'!$F$2")
        ws.cell(row=9, column=2).number_format = '$0.00'
        
        # Row 10: Market Capitalization
        ws.cell(row=10, column=1, value="Market Capitalization")
        ws.cell(row=10, column=2, value="=B8*B9")
        ws.cell(row=10, column=2).number_format = '[$$-409]#,,0.0,," B"'
    
    def _setup_perpetual_dcf_outputs(self, ws: Worksheet) -> None:
        """
        Set up Perpetual Growth DCF Outputs section (rows 13-18).
        
        Rows:
        13: Equity Value = Valuation (DCF)!B33
        14: Cash & Equivalents = Valuation (DCF)!B30
        15: Total Debt = Valuation (DCF)!B31
        16: Investments = Valuation (DCF)!B32
        17: Enterprise Value = B13 - B14 + B15 - B16
        18: Value per Share = Valuation (DCF)!B37
        """
        # Blank rows
        ws.cell(row=11, column=1, value="")
        ws.cell(row=12, column=1, value="")
        
        # Row 13: Equity Value (Perpetual DCF)
        ws.cell(row=13, column=1, value="Equity Value (Perpetual DCF)")
        ws.cell(row=13, column=1).font = Font(bold=True)
        ws.cell(row=13, column=2, value="='Valuation (DCF)'!$B$33")
        ws.cell(row=13, column=2).number_format = '[$$-409]#,,0.0,," B"'
        
        # Row 14: Cash & Equivalents
        ws.cell(row=14, column=1, value="Cash & Equivalents (from DCF)")
        ws.cell(row=14, column=2, value="='Valuation (DCF)'!$B$30")
        ws.cell(row=14, column=2).number_format = '[$$-409]#,,0.0,," B"'
        
        # Row 15: Total Debt
        ws.cell(row=15, column=1, value="Total Debt (from DCF)")
        ws.cell(row=15, column=2, value="='Valuation (DCF)'!$B$31")
        ws.cell(row=15, column=2).number_format = '[$$-409]#,,0.0,," B"'
        
        # Row 16: Investments
        ws.cell(row=16, column=1, value="Investments / Non-operating")
        ws.cell(row=16, column=2, value="='Valuation (DCF)'!$B$32")
        ws.cell(row=16, column=2).number_format = '[$$-409]#,,0.0,," B"'
        
        # Row 17: Enterprise Value (Perpetual DCF)
        ws.cell(row=17, column=1, value="Enterprise Value (Perpetual DCF)")
        ws.cell(row=17, column=1).font = Font(bold=True)
        ws.cell(row=17, column=2, value="=B13-B14+B15-B16")
        ws.cell(row=17, column=2).number_format = '[$$-409]#,,0.0,," B"'
        ws.cell(row=17, column=2).font = Font(bold=True)
        
        # Row 18: Value per Share (Perpetual DCF)
        ws.cell(row=18, column=1, value="Value per Share (Perpetual DCF)")
        ws.cell(row=18, column=1).font = Font(bold=True, size=11)
        ws.cell(row=18, column=2, value="='Valuation (DCF)'!$B$37")
        ws.cell(row=18, column=2).number_format = '$0.00'
        ws.cell(row=18, column=2).font = Font(bold=True, size=11)
        ws.cell(row=18, column=2).fill = PatternFill(
            start_color=ExcelFormats.IMPORTANT_COLOR,
            end_color=ExcelFormats.IMPORTANT_COLOR,
            fill_type="solid"
        )
    
    def _setup_exit_multiple_outputs(self, ws: Worksheet) -> None:
        """
        Set up Exit Multiple DCF Outputs section (rows 20-22).
        
        Rows:
        20: Enterprise Value = Valuation (Exit Multiple)!B17
        21: Equity Value = Valuation (Exit Multiple)!B22
        22: Value per Share = Valuation (Exit Multiple)!B25
        """
        # Blank row
        ws.cell(row=19, column=1, value="")
        
        # Row 20: Enterprise Value (Exit Multiple DCF)
        ws.cell(row=20, column=1, value="Enterprise Value (Exit Multiple DCF)")
        ws.cell(row=20, column=1).font = Font(bold=True)
        ws.cell(row=20, column=2, value="='Valuation (Exit Multiple)'!$B$17")
        ws.cell(row=20, column=2).number_format = '[$$-409]#,,0.0,," B"'
        ws.cell(row=20, column=2).font = Font(bold=True)
        
        # Row 21: Equity Value (Exit Multiple DCF)
        ws.cell(row=21, column=1, value="Equity Value (Exit Multiple DCF)")
        ws.cell(row=21, column=2, value="='Valuation (Exit Multiple)'!$B$22")
        ws.cell(row=21, column=2).number_format = '[$$-409]#,,0.0,," B"'
        
        # Row 22: Value per Share (Exit Multiple DCF)
        ws.cell(row=22, column=1, value="Value per Share (Exit Multiple DCF)")
        ws.cell(row=22, column=1).font = Font(bold=True, size=11)
        ws.cell(row=22, column=2, value="='Valuation (Exit Multiple)'!$B$25")
        ws.cell(row=22, column=2).number_format = '$0.00'
        ws.cell(row=22, column=2).font = Font(bold=True, size=11)
        ws.cell(row=22, column=2).fill = PatternFill(
            start_color=ExcelFormats.IMPORTANT_COLOR,
            end_color=ExcelFormats.IMPORTANT_COLOR,
            fill_type="solid"
        )
    
    def _setup_blended_valuation(self, ws: Worksheet) -> None:
        """
        Set up Blended Valuation & Market Comparison section (rows 26-29).
        
        Rows:
        26: Average of Methods = AVERAGE(B18, B22)
        27: Upside vs Market = IFERROR(B26/B9-1, "")
        28: Premium (Exit vs Perpetual) = IFERROR(B22/B18-1, "")
        29: Market Enterprise Value = B10 - B14 + B15 - B16
        """
        # Blank rows
        ws.cell(row=23, column=1, value="")
        ws.cell(row=24, column=1, value="")
        ws.cell(row=25, column=1, value="")
        
        # Row 26: Average of Methods
        ws.cell(row=26, column=1, value="Average of Methods (Per-Share)")
        ws.cell(row=26, column=1).font = Font(bold=True, size=12)
        ws.cell(row=26, column=2, value="=AVERAGE($B$18,$B$22)")
        ws.cell(row=26, column=2).number_format = '$0.00'
        ws.cell(row=26, column=2).font = Font(bold=True, size=12)
        ws.cell(row=26, column=2).fill = PatternFill(
            start_color="FFD966",  # Gold color for highlight
            end_color="FFD966",
            fill_type="solid"
        )
        
        # Row 27: Upside vs Market
        ws.cell(row=27, column=1, value="Upside vs Market")
        ws.cell(row=27, column=1).font = Font(bold=True, size=11)
        ws.cell(row=27, column=2, value="=IFERROR($B$26/$B$9-1,\"\")")
        ws.cell(row=27, column=2).number_format = '0.0%'
        ws.cell(row=27, column=2).font = Font(bold=True, size=11)
        ws.cell(row=27, column=2).fill = PatternFill(
            start_color="FFD966",  # Gold color for highlight
            end_color="FFD966",
            fill_type="solid"
        )
        
        # Row 28: Premium (Exit vs Perpetual)
        ws.cell(row=28, column=1, value="Premium (Exit vs Perpetual)")
        ws.cell(row=28, column=2, value="=IFERROR($B$22/$B$18-1,\"\")")
        ws.cell(row=28, column=2).number_format = '0.0%'
        
        # Row 29: Market Enterprise Value
        ws.cell(row=29, column=1, value="Market Enterprise Value")
        ws.cell(row=29, column=2, value="=$B$10-$B$14+$B$15-$B$16")
        ws.cell(row=29, column=2).number_format = '[$$-409]#,,0.0,," B"'
    
    def _setup_sanity_metrics(self, ws: Worksheet) -> None:
        """
        Set up Sanity Metrics (Terminal Year) section (rows 33-39).
        
        Rows:
        33: Revenue (FY5) = Projections!F3
        34: EBITDA (FY5) = Projections!F21
        35: FCF (FY5) = Projections!F19
        36: EV/EBITDA - Perpetual = IFERROR(B17/B34, "")
        37: EV/EBITDA - Exit Multiple = IFERROR(B20/B34, "")
        38: FCF Yield on EV - Perpetual = IFERROR(B35/B17, "")
        39: FCF Yield on EV - Exit Multiple = IFERROR(B35/B20, "")
        """
        # Blank rows
        ws.cell(row=30, column=1, value="")
        ws.cell(row=31, column=1, value="")
        ws.cell(row=32, column=1, value="")
        
        # Row 33: Revenue (FY5)
        ws.cell(row=33, column=1, value="Revenue (FY5)")
        ws.cell(row=33, column=2, value="='Projections'!$F$3")
        ws.cell(row=33, column=2).number_format = '[$$-409]#,,0.0,," B"'
        
        # Row 34: EBITDA (FY5)
        ws.cell(row=34, column=1, value="EBITDA (FY5)")
        ws.cell(row=34, column=2, value="='Projections'!$F$21")
        ws.cell(row=34, column=2).number_format = '[$$-409]#,,0.0,," B"'
        
        # Row 35: FCF (FY5)
        ws.cell(row=35, column=1, value="FCF (FY5)")
        ws.cell(row=35, column=2, value="='Projections'!$F$19")
        ws.cell(row=35, column=2).number_format = '[$$-409]#,,0.0,," B"'
        
        # Row 36: EV/EBITDA - Perpetual
        ws.cell(row=36, column=1, value="EV/EBITDA — Perpetual")
        ws.cell(row=36, column=2, value="=IFERROR($B$17/$B$34,\"\")")
        ws.cell(row=36, column=2).number_format = '0.0"x"'
        
        # Row 37: EV/EBITDA - Exit Multiple
        ws.cell(row=37, column=1, value="EV/EBITDA — Exit Multiple")
        ws.cell(row=37, column=2, value="=IFERROR($B$20/$B$34,\"\")")
        ws.cell(row=37, column=2).number_format = '0.0"x"'
        
        # Row 38: FCF Yield on EV - Perpetual
        ws.cell(row=38, column=1, value="FCF Yield on EV — Perpetual")
        ws.cell(row=38, column=2, value="=IFERROR($B$35/$B$17,\"\")")
        ws.cell(row=38, column=2).number_format = '0.0%'
        
        # Row 39: FCF Yield on EV - Exit Multiple
        ws.cell(row=39, column=1, value="FCF Yield on EV — Exit Multiple")
        ws.cell(row=39, column=2, value="=IFERROR($B$35/$B$20,\"\")")
        ws.cell(row=39, column=2).number_format = '0.0%'
    
    def _setup_qa_flags(self, ws: Worksheet) -> None:
        """
        Set up Quality Assurance Flags section (rows 41-46).
        
        These return TRUE when checks pass, FALSE when they fail.
        Add conditional formatting to highlight FALSE values in red.
        
        Rows:
        41: Check: E/V + D/V = 1.0 (capital structure adds to 100%)
        42: Check: WACC > g (perpetuity validity)
        43: Check: DF ≤ 1 (Perpetual) (all discount factors valid)
        44: Check: DF ≤ 1 (Exit Multiple) (all discount factors valid)
        45: Check: Shares > 0 & Price > 0 (basic data validity)
        46: Check: Mid-Year toggle wired (dropdown working)
        """
        # Blank row
        ws.cell(row=40, column=1, value="")
        
        # Row 41: Check E/V + D/V = 1.0
        ws.cell(row=41, column=1, value="Check: E/V + D/V = 1.0")
        ws.cell(row=41, column=2, value="=ABS('Valuation (DCF)'!$B$10+'Valuation (DCF)'!$B$11-1)<=0.001")
        ws.cell(row=41, column=2).alignment = Alignment(horizontal="center")
        
        # Row 42: Check WACC > g
        ws.cell(row=42, column=1, value="Check: WACC > g")
        ws.cell(row=42, column=2, value="='Valuation (DCF)'!$B$12>'Valuation (DCF)'!$B$23")
        ws.cell(row=42, column=2).alignment = Alignment(horizontal="center")
        
        # Row 43: Check DF ≤ 1 (Perpetual)
        ws.cell(row=43, column=1, value="Check: DF ≤ 1 (Perpetual)")
        ws.cell(row=43, column=2, value="=MAX('Valuation (DCF)'!$B$17:'Valuation (DCF)'!$F$17)<=1")
        ws.cell(row=43, column=2).alignment = Alignment(horizontal="center")
        
        # Row 44: Check DF ≤ 1 (Exit Multiple)
        ws.cell(row=44, column=1, value="Check: DF ≤ 1 (Exit Multiple)")
        ws.cell(row=44, column=2, value="=MAX('Valuation (Exit Multiple)'!$B$8:'Valuation (Exit Multiple)'!$F$8)<=1")
        ws.cell(row=44, column=2).alignment = Alignment(horizontal="center")
        
        # Row 45: Check Shares > 0 & Price > 0
        ws.cell(row=45, column=1, value="Check: Shares > 0 & Price > 0")
        ws.cell(row=45, column=2, value="=AND($B$8>0,$B$9>0)")
        ws.cell(row=45, column=2).alignment = Alignment(horizontal="center")
        
        # Row 46: Check Mid-Year toggle wired
        ws.cell(row=46, column=1, value="Check: Mid-Year toggle wired")
        ws.cell(row=46, column=2, value="=OR('Sensitivity'!$B$2=\"No\",'Sensitivity'!$B$2=\"Yes\")")
        ws.cell(row=46, column=2).alignment = Alignment(horizontal="center")
    
    def _format_sheet(self, ws: Worksheet) -> None:
        """Apply general formatting to the sheet."""
        # Set column widths
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 30
        
        # Turn off gridlines for cleaner look
        ws.sheet_view.showGridLines = False
        
        # Add a note about QA flags
        ws.cell(row=48, column=1, value="Note: QA flags return TRUE when checks pass. Add conditional formatting to highlight FALSE values.")
        ws.cell(row=48, column=1).font = Font(italic=True, size=9, color="808080")
