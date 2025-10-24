"""
Tab: Lever_Map - Parameter Adjustment Governance

This tab shows base → adjusted parameter transitions with:
- Base values from original LLM_Inferred
- Adjusted values from LLM_Inferred_Adjusted  
- Deltas, caps, decay factors, and effective adjustments
- Apply flags for analyst approval
- Traceability to news factors

Layout per price_adjustor_tab.md:
Columns: Lever_Name, Sheet_Target, Row_Label, Col_Label, Unit,
         Base_Value, Adjusted_Value, Delta_Raw, Cap_Pos, Cap_Neg, Delta_Capped,
         Half_Life_Days, Days_Since_AsOf, Decay_Factor, Delta_Effective,
         Final_Value, Apply_Flag, Value_To_Use, Source_URL, Notes
"""

from typing import Dict, Any, Optional, List
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class LeverMapTabBuilder:
    """
    Builds the Lever_Map tab showing parameter adjustments with governance.
    
    This tab provides full transparency into how news affects model parameters:
    - Shows base vs adjusted values
    - Applies caps and decay
    - Allows analyst override via Apply_Flag
    - Traces back to specific news factors
    """
    
    def __init__(self, adjustment_metadata: Optional[Dict[str, Any]] = None):
        """
        Initialize with adjustment metadata from LLM.
        
        Args:
            adjustment_metadata: Dict with adjustments_summary and factor_mappings
        """
        self.adjustment_metadata = adjustment_metadata or {}
    
    def create_tab(self, workbook: openpyxl.Workbook) -> Worksheet:
        """Create and format the Lever_Map tab."""
        # Remove existing tab if it exists
        if "Lever_Map" in workbook.sheetnames:
            ws = workbook["Lever_Map"]
            workbook.remove(ws)
        
        # Create new tab (insert after LLM_Inferred_Adjusted)
        if "LLM_Inferred_Adjusted" in workbook.sheetnames:
            idx = workbook.sheetnames.index("LLM_Inferred_Adjusted")
            ws = workbook.create_sheet("Lever_Map", idx + 1)
        else:
            ws = workbook.create_sheet("Lever_Map")
        
        self._setup_headers(ws)
        self._populate_lever_rows(ws)
        self._format_sheet(ws)
        
        return ws
    
    def _setup_headers(self, ws: Worksheet) -> None:
        """Set up column headers."""
        headers = [
            ("A", "Lever_Name"),
            ("B", "Sheet_Target"),
            ("C", "Row_Label"),
            ("D", "Col_Label"),
            ("E", "Unit"),
            ("F", "Base_Value"),
            ("G", "Adjusted_Value"),
            ("H", "Delta_Raw"),
            ("I", "Cap_Pos"),
            ("J", "Cap_Neg"),
            ("K", "Delta_Capped"),
            ("L", "Half_Life_Days"),
            ("M", "Days_Since_AsOf"),
            ("N", "Decay_Factor"),
            ("O", "Delta_Effective"),
            ("P", "Final_Value"),
            ("Q", "Apply_Flag"),
            ("R", "Value_To_Use"),
            ("S", "Source_URL"),
            ("T", "Notes")
        ]
        
        # Write headers
        for col, header in headers:
            cell = ws[f"{col}1"]
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Freeze header row
        ws.freeze_panes = ws['A2']
    
    def _populate_lever_rows(self, ws: Worksheet) -> None:
        """Populate lever rows with data and formulas."""
        # Define lever configurations
        # Each lever: (name, sheet_target, row_label, col_label, unit, cap_pos, cap_neg, half_life, notes)
        levers = [
            # Revenue Growth (FY1-FY5)
            ("RevenueGrowth_FY1", "Assumptions", "Revenue Growth (YoY)", "FY1", "pct", 0.0300, 0.0300, 0, "From catalysts/risks"),
            ("RevenueGrowth_FY2", "Assumptions", "Revenue Growth (YoY)", "FY2", "pct", 0.0200, 0.0200, 0, "Taper from FY1"),
            ("RevenueGrowth_FY3", "Assumptions", "Revenue Growth (YoY)", "FY3", "pct", 0.0150, 0.0150, 0, "Taper"),
            ("RevenueGrowth_FY4", "Assumptions", "Revenue Growth (YoY)", "FY4", "pct", 0.0100, 0.0100, 0, "Taper"),
            ("RevenueGrowth_FY5", "Assumptions", "Revenue Growth (YoY)", "FY5", "pct", 0.0100, 0.0100, 0, "Taper"),
            
            # Gross Margin (FY1-FY5)
            ("GrossMargin_FY1", "Assumptions", "Gross Margin", "FY1", "pct", 0.0075, 0.0075, 0, "From cost/pricing factors"),
            ("GrossMargin_FY2", "Assumptions", "Gross Margin", "FY2", "pct", 0.0050, 0.0050, 0, "Taper"),
            ("GrossMargin_FY3", "Assumptions", "Gross Margin", "FY3", "pct", 0.0025, 0.0025, 0, "Taper"),
            
            # Operating Margin (FY1-FY5)
            ("OperatingMargin_FY1", "Assumptions", "Operating Margin", "FY1", "pct", 0.0050, 0.0050, 0, "From OpEx/efficiency"),
            ("OperatingMargin_FY2", "Assumptions", "Operating Margin", "FY2", "pct", 0.0040, 0.0040, 0, "Taper"),
            ("OperatingMargin_FY3", "Assumptions", "Operating Margin", "FY3", "pct", 0.0030, 0.0030, 0, "Taper"),
            
            # EBITDA Margin (FY1-FY3)
            ("EBITDAMargin_FY1", "Assumptions", "EBITDA Margin", "FY1", "pct", 0.0050, 0.0050, 0, "From OpEx/efficiency"),
            ("EBITDAMargin_FY2", "Assumptions", "EBITDA Margin", "FY2", "pct", 0.0040, 0.0040, 0, "Taper"),
            
            # Working Capital (FY1-FY3)
            ("DSO_FY1", "Assumptions", "DSO (Days)", "FY1", "days", 5, 5, 60, "Decays if unconfirmed"),
            ("DSO_FY2", "Assumptions", "DSO (Days)", "FY2", "days", 5, 5, 60, ""),
            ("DIO_FY1", "Assumptions", "DIO (Days)", "FY1", "days", 5, 5, 60, "Inventory efficiency"),
            ("DIO_FY2", "Assumptions", "DIO (Days)", "FY2", "days", 5, 5, 60, ""),
            ("DPO_FY1", "Assumptions", "DPO (Days)", "FY1", "days", 5, 5, 60, "Payables terms"),
            ("DPO_FY2", "Assumptions", "DPO (Days)", "FY2", "days", 5, 5, 60, ""),
            
            # Valuation Parameters
            ("WACC", "Assumptions", "WACC", "FY0 (Actual)", "pct", 0.0025, 0.0025, 30, "Only for credit events"),
            ("TerminalGrowth", "Assumptions", "Terminal Growth Rate (g)", "FY0 (Actual)", "pct", 0.0010, 0.0010, 0, "Structural shifts only"),
        ]
        
        row = 2
        for lever_config in levers:
            self._write_lever_row(ws, row, lever_config)
            row += 1
    
    def _write_lever_row(self, ws: Worksheet, row: int, config: tuple) -> None:
        """Write a single lever row with formulas."""
        (lever_name, sheet_target, row_label, col_label, unit, 
         cap_pos, cap_neg, half_life, notes) = config
        
        # A: Lever_Name
        ws[f"A{row}"] = lever_name
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # B: Sheet_Target
        ws[f"B{row}"] = sheet_target
        
        # C: Row_Label
        ws[f"C{row}"] = row_label
        
        # D: Col_Label
        ws[f"D{row}"] = col_label
        
        # E: Unit
        ws[f"E{row}"] = unit
        
        # F: Base_Value (lookup from LLM_Inferred)
        # Map FY1-FY5 to columns B-F in LLM_Inferred
        fy_to_col = {"FY1": "B", "FY2": "C", "FY3": "D", "FY4": "E", "FY5": "F", "FY0 (Actual)": "B"}
        llm_col = fy_to_col.get(col_label, "B")
        
        # Map row labels to LLM_Inferred rows
        label_to_row = {
            "Revenue Growth (YoY)": 4,
            "Gross Margin": 5,
            "EBITDA Margin": 6,
            "Operating Margin": 7,
            "DSO (Days)": 8,
            "DIO (Days)": 9,
            "DPO (Days)": 10,
            "WACC": 2,
            "Terminal Growth Rate (g)": 3
        }
        llm_row = label_to_row.get(row_label, 2)
        
        ws[f"F{row}"] = f"=LLM_Inferred!{llm_col}{llm_row}"
        
        # G: Adjusted_Value (lookup from LLM_Inferred_Adjusted)
        ws[f"G{row}"] = f"=LLM_Inferred_Adjusted!{llm_col}{llm_row}"
        
        # H: Delta_Raw (Adjusted - Base)
        ws[f"H{row}"] = f"=G{row}-F{row}"
        
        # I: Cap_Pos
        ws[f"I{row}"] = cap_pos
        
        # J: Cap_Neg
        ws[f"J{row}"] = cap_neg
        
        # K: Delta_Capped (apply caps)
        ws[f"K{row}"] = f"=IF(H{row}>0,MIN(H{row},I{row}),-MIN(ABS(H{row}),J{row}))"
        
        # L: Half_Life_Days
        ws[f"L{row}"] = half_life
        
        # M: Days_Since_AsOf (default 0 for now)
        ws[f"M{row}"] = 0
        
        # N: Decay_Factor
        ws[f"N{row}"] = f"=IF(L{row}=0,1,EXP(-LN(2)*M{row}/L{row}))"
        
        # O: Delta_Effective (capped delta × decay)
        ws[f"O{row}"] = f"=K{row}*N{row}"
        
        # P: Final_Value (base + effective delta)
        ws[f"P{row}"] = f"=F{row}+O{row}"
        
        # Q: Apply_Flag (default 1 = enabled)
        ws[f"Q{row}"] = 1
        
        # R: Value_To_Use (if apply_flag=1, use final, else use base)
        ws[f"R{row}"] = f"=IF(Q{row}=1,P{row},F{row})"
        
        # S: Source_URL (empty for now, could link to news)
        ws[f"S{row}"] = ""
        
        # T: Notes
        ws[f"T{row}"] = notes
    
    def _format_sheet(self, ws: Worksheet) -> None:
        """Apply formatting to the sheet."""
        # Column widths
        widths = {
            'A': 20,  # Lever_Name
            'B': 15,  # Sheet_Target
            'C': 25,  # Row_Label
            'D': 12,  # Col_Label
            'E': 8,   # Unit
            'F': 12,  # Base_Value
            'G': 12,  # Adjusted_Value
            'H': 12,  # Delta_Raw
            'I': 10,  # Cap_Pos
            'J': 10,  # Cap_Neg
            'K': 12,  # Delta_Capped
            'L': 12,  # Half_Life_Days
            'M': 12,  # Days_Since_AsOf
            'N': 12,  # Decay_Factor
            'O': 12,  # Delta_Effective
            'P': 12,  # Final_Value
            'Q': 10,  # Apply_Flag
            'R': 12,  # Value_To_Use
            'S': 15,  # Source_URL
            'T': 30   # Notes
        }
        
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        # Number formats
        # Percentages: F-H, K, O-P, R (for pct unit rows)
        # Days: same columns but for days unit rows
        # We'll apply formats conditionally based on row
        
        # For simplicity, apply percentage format to known percentage columns
        pct_cols = ['F', 'G', 'H', 'K', 'O', 'P', 'R']
        
        # Find max row
        max_row = ws.max_row
        
        for row in range(2, max_row + 1):
            unit = ws[f"E{row}"].value
            
            if unit == "pct":
                # Apply percentage format
                for col in pct_cols:
                    ws[f"{col}{row}"].number_format = '0.00%'
            elif unit == "days":
                # Apply number format
                for col in pct_cols:
                    ws[f"{col}{row}"].number_format = '0.0'
            
            # Caps (always absolute values)
            ws[f"I{row}"].number_format = '0.0000' if unit == "pct" else '0.0'
            ws[f"J{row}"].number_format = '0.0000' if unit == "pct" else '0.0'
            
            # Decay factor (always decimal)
            ws[f"N{row}"].number_format = '0.000'
            
            # Apply_Flag (integer)
            ws[f"Q{row}"].number_format = '0'
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, max_row + 1):
            for col in range(1, 21):  # A-T
                ws.cell(row=row, column=col).border = thin_border
        
        # Color coding
        # Base/Adjusted columns: light blue
        base_fill = PatternFill(start_color="E0E7F1", end_color="E0E7F1", fill_type="solid")
        for row in range(2, max_row + 1):
            ws[f"F{row}"].fill = base_fill
            ws[f"G{row}"].fill = base_fill
        
        # Delta columns: light yellow
        delta_fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
        for row in range(2, max_row + 1):
            ws[f"H{row}"].fill = delta_fill
            ws[f"K{row}"].fill = delta_fill
            ws[f"O{row}"].fill = delta_fill
        
        # Final columns: light green
        final_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        for row in range(2, max_row + 1):
            ws[f"P{row}"].fill = final_fill
            ws[f"R{row}"].fill = final_fill
        
        # Apply_Flag: light orange (input)
        flag_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        for row in range(2, max_row + 1):
            ws[f"Q{row}"].fill = flag_fill
