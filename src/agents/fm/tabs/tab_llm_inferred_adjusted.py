"""
Tab: LLM_Inferred_Adjusted - News-Adjusted Assumptions

This tab stores LLM-inferred parameter values AFTER adjusting for news analysis.
It mirrors the structure of LLM_Inferred but contains news-driven adjustments.

Layout (same as LLM_Inferred):
- Row 1: Headers (Metric, FY1-FY5)
- Row 2: WACC (B2)
- Row 3: Terminal Growth Rate (B3)
- Row 4: Revenue Growth Rates (B4:F4)
- Row 5: Gross Margins (B5:F5)
- Row 6: EBITDA Margins (B6:F6)
- Row 7: Operating Margins (B7:F7)
- Row 8: DSO Days (B8:F8)
- Row 9: DIO Days (B9:F9)
- Row 10: DPO Days (B10:F10)
"""

from typing import Dict, Any, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment


class LLMInferredAdjustedTabBuilder:
    """
    Builds the LLM_Inferred_Adjusted tab with news-adjusted values.
    
    This tab is similar to LLM_Inferred but contains parameter values
    adjusted based on catalysts, risks, and mitigations from news screening.
    """
    
    def __init__(self, adjusted_assumptions: Optional[Dict[str, Any]] = None):
        """
        Initialize with adjusted assumptions from LLM.
        
        Args:
            adjusted_assumptions: Dict with adjusted parameter values
        """
        self.adjusted_assumptions = adjusted_assumptions or {}
    
    def create_tab(self, workbook: openpyxl.Workbook) -> Worksheet:
        """Create and format the LLM_Inferred_Adjusted tab."""
        # Remove existing tab if it exists
        if "LLM_Inferred_Adjusted" in workbook.sheetnames:
            ws = workbook["LLM_Inferred_Adjusted"]
            workbook.remove(ws)
        
        # Create new tab (insert after LLM_Inferred if it exists)
        if "LLM_Inferred" in workbook.sheetnames:
            llm_idx = workbook.sheetnames.index("LLM_Inferred")
            ws = workbook.create_sheet("LLM_Inferred_Adjusted", llm_idx + 1)
        else:
            ws = workbook.create_sheet("LLM_Inferred_Adjusted")
        
        self._setup_headers(ws)
        self._populate_data(ws)
        self._format_sheet(ws)
        
        return ws
    
    def _setup_headers(self, ws: Worksheet) -> None:
        """Set up column headers."""
        # Header row
        ws.cell(row=1, column=1, value="Metric").font = Font(bold=True)
        for i in range(5):
            ws.cell(row=1, column=2 + i, value=f"FY{i+1}").font = Font(bold=True)
        
        # Style headers with light blue background
        header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        for col in range(1, 7):
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
    
    def _populate_data(self, ws: Worksheet) -> None:
        """Populate adjusted parameter values."""
        # Get adjusted values (with fallback to defaults if missing)
        wacc = self.adjusted_assumptions.get('wacc', 0.09)
        terminal_growth = self.adjusted_assumptions.get('terminal_growth_rate', 0.025)
        revenue_growth = self.adjusted_assumptions.get('revenue_growth_rates', [0.05] * 5)
        gross_margins = self.adjusted_assumptions.get('gross_margins', [0.46] * 5)
        ebitda_margins = self.adjusted_assumptions.get('ebitda_margins', [0.33] * 5)
        operating_margins = self.adjusted_assumptions.get('operating_margins', [0.31] * 5)
        dso_days = self.adjusted_assumptions.get('dso_days', [45] * 5)
        dio_days = self.adjusted_assumptions.get('dio_days', [10] * 5)
        dpo_days = self.adjusted_assumptions.get('dpo_days', [90] * 5)
        
        # Ensure all lists have exactly 5 elements
        def ensure_length(lst, default, length=5):
            if len(lst) < length:
                return lst + [default] * (length - len(lst))
            return lst[:length]
        
        revenue_growth = ensure_length(revenue_growth, 0.05)
        gross_margins = ensure_length(gross_margins, 0.46)
        ebitda_margins = ensure_length(ebitda_margins, 0.33)
        operating_margins = ensure_length(operating_margins, 0.31)
        dso_days = ensure_length(dso_days, 45)
        dio_days = ensure_length(dio_days, 10)
        dpo_days = ensure_length(dpo_days, 90)
        
        # Row 2: WACC
        ws.cell(row=2, column=1, value="WACC").font = Font(bold=True)
        ws.cell(row=2, column=2, value=wacc)
        
        # Row 3: Terminal Growth Rate
        ws.cell(row=3, column=1, value="Terminal Growth Rate").font = Font(bold=True)
        ws.cell(row=3, column=2, value=terminal_growth)
        
        # Row 4: Revenue Growth Rates (FY1-FY5)
        ws.cell(row=4, column=1, value="Revenue Growth Rate").font = Font(bold=True)
        for i, rate in enumerate(revenue_growth):
            ws.cell(row=4, column=2 + i, value=rate)
        
        # Row 5: Gross Margins (FY1-FY5)
        ws.cell(row=5, column=1, value="Gross Margin").font = Font(bold=True)
        for i, margin in enumerate(gross_margins):
            ws.cell(row=5, column=2 + i, value=margin)
        
        # Row 6: EBITDA Margins (FY1-FY5)
        ws.cell(row=6, column=1, value="EBITDA Margin").font = Font(bold=True)
        for i, margin in enumerate(ebitda_margins):
            ws.cell(row=6, column=2 + i, value=margin)
        
        # Row 7: Operating Margins (FY1-FY5)
        ws.cell(row=7, column=1, value="Operating Margin").font = Font(bold=True)
        for i, margin in enumerate(operating_margins):
            ws.cell(row=7, column=2 + i, value=margin)
        
        # Row 8: DSO Days (FY1-FY5)
        ws.cell(row=8, column=1, value="DSO Days").font = Font(bold=True)
        for i, days in enumerate(dso_days):
            ws.cell(row=8, column=2 + i, value=days)
        
        # Row 9: DIO Days (FY1-FY5)
        ws.cell(row=9, column=1, value="DIO Days").font = Font(bold=True)
        for i, days in enumerate(dio_days):
            ws.cell(row=9, column=2 + i, value=days)
        
        # Row 10: DPO Days (FY1-FY5)
        ws.cell(row=10, column=1, value="DPO Days").font = Font(bold=True)
        for i, days in enumerate(dpo_days):
            ws.cell(row=10, column=2 + i, value=days)
    
    def _format_sheet(self, ws: Worksheet) -> None:
        """Apply formatting to the sheet."""
        # Column widths
        ws.column_dimensions['A'].width = 25
        for col in 'BCDEF':
            ws.column_dimensions[col].width = 12
        
        # Number formats
        # WACC and Terminal Growth (percentages)
        ws['B2'].number_format = '0.00%'
        ws['B3'].number_format = '0.00%'
        
        # Revenue Growth Rates (row 4) - percentages
        for col in 'BCDEF':
            ws[f'{col}4'].number_format = '0.00%'
        
        # Margins (rows 5-7) - percentages
        for row in [5, 6, 7]:
            for col in 'BCDEF':
                ws[f'{col}{row}'].number_format = '0.00%'
        
        # Days (rows 8-10) - numbers
        for row in [8, 9, 10]:
            for col in 'BCDEF':
                ws[f'{col}{row}'].number_format = '0.0'
        
        # Freeze top row
        ws.freeze_panes = ws['A2']
        
        # Add light yellow fill to data cells (news-adjusted)
        data_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for row in range(2, 11):
            for col in range(2, 7):
                ws.cell(row=row, column=col).fill = data_fill
