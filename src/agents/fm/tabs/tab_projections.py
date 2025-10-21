"""
Tab 4: Projections Tab Builder

This module creates the Projections tab - 5-year forward P&L and FCF
projections based on assumptions from the Assumptions tab.

Professional-grade financial model following investment banking standards:
- All formulas reference Historical (FY0 base) and Assumptions (drivers)
- Revenue driven by growth rates from Assumptions!C7:G7
- Gross Margin driven by Assumptions!C9:G9
- R&D/SG&A scaled by revenue intensity
- Working capital modeled using DSO/DIO/DPO explicit drivers
- Unlevered FCF = NOPAT + D&A + Capex - ΔNWC
"""

from typing import List
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment

from ..financial_model_builder import ExcelFormats


class ProjectionsTabBuilder:
    """
    Builds the Projections tab - 5-year forward projections with professional analytics.
    
    This tab contains:
    - Year headers (row 2): FY1-FY5 auto-increment from Assumptions!B2
    - Revenue & COGS (rows 3-6)
    - Operating expenses → EBIT (rows 7-9)
    - Taxes → NOPAT (rows 10-11)
    - D&A, Capex (rows 12-13)
    - Working Capital components (rows 14-18): AR, Inv, AP, NWC, ΔNWC
    - Unlevered FCF (row 19)
    - Analytics & Diagnostics (rows 20-41):
      * EBITDA & EBITDA Margin
      * Operational metrics (R&D%, SG&A%, EBIT%, D&A%, Capex% as % of Revenue)
      * Working capital drivers (DSO, DIO, DPO)
      * Working capital changes (ΔAR, ΔInventory, ΔAP, NWC % of Revenue)
      * PP&E roll-forward (Beginning PP&E, Capex, D&A, Ending PP&E)
      * Key ratios (FCF Margin %, Reinvestment Rate)
    
    All formulas exactly match professional investment banking standards.
    """
    
    def __init__(self, projection_years: int = 5):
        """
        Initialize the Projections builder.
        
        Args:
            projection_years: Number of projection years (default: 5)
        """
        self.projection_years = projection_years
    
    def create_tab(self, workbook: openpyxl.Workbook) -> Worksheet:
        """
        Create and format the Projections tab.
        
        Args:
            workbook: The Excel workbook to add the tab to
            
        Returns:
            The created worksheet
        """
        # Create or get the sheet
        if "Projections" in workbook.sheetnames:
            ws = workbook["Projections"]
            workbook.remove(ws)
        
        ws = workbook.create_sheet("Projections", 4)  # Position 4 (fifth tab)
        
        # Set up sections following exact specification
        self._setup_headers(ws)
        self._setup_revenue_section(ws)
        self._setup_opex_section(ws)
        self._setup_tax_section(ws)
        self._setup_capex_section(ws)
        self._setup_nwc_section(ws)
        self._setup_fcf_row(ws)
        self._setup_analytics_section(ws)  # Add professional analytics & diagnostics
        
        # Format the sheet
        self._format_sheet(ws)
        
        return ws
    
    def _setup_headers(self, ws: Worksheet) -> None:
        """
        Set up projection year headers (Row 2).
        
        Row 2 structure:
        - A2: "Metric"
        - B2: =Assumptions!$B$2+1 (FY1)
        - C2: =B2+1 (FY2)
        - D2: =C2+1 (FY3)
        - E2: =D2+1 (FY4)
        - F2: =E2+1 (FY5)
        """
        # Row 1: Main header
        ws.cell(row=1, column=1, value="PROJECTIONS (FY1-FY5)")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Row 2: Year headers
        ws.cell(row=2, column=1, value="Metric")
        ws.cell(row=2, column=1).font = Font(bold=True, size=11)
        ws.cell(row=2, column=1).fill = PatternFill(
            start_color=ExcelFormats.HEADER_COLOR,
            end_color=ExcelFormats.HEADER_COLOR,
            fill_type="solid"
        )
        
        # B2: FY1 = Assumptions!B2 + 1
        ws.cell(row=2, column=2, value='=Assumptions!$B$2+1')
        ws.cell(row=2, column=2).number_format = '0'
        ws.cell(row=2, column=2).font = Font(bold=True, size=11)
        ws.cell(row=2, column=2).fill = PatternFill(
            start_color=ExcelFormats.HEADER_COLOR,
            end_color=ExcelFormats.HEADER_COLOR,
            fill_type="solid"
        )
        ws.cell(row=2, column=2).alignment = Alignment(horizontal="center", vertical="center")
        
        # C2-F2: Auto-increment from previous year
        for i in range(1, self.projection_years):
            col = 2 + i  # columns 3, 4, 5, 6 = C, D, E, F
            col_letter = chr(64 + col)  # C, D, E, F
            prev_col_letter = chr(64 + col - 1)  # B, C, D, E (the previous column)
            ws.cell(row=2, column=col, value=f'={prev_col_letter}2+1')
            ws.cell(row=2, column=col).number_format = '0'
            ws.cell(row=2, column=col).font = Font(bold=True, size=11)
            ws.cell(row=2, column=col).fill = PatternFill(
                start_color=ExcelFormats.HEADER_COLOR,
                end_color=ExcelFormats.HEADER_COLOR,
                fill_type="solid"
            )
            ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")
    
    def _setup_revenue_section(self, ws: Worksheet) -> None:
        """
        Set up Revenue & COGS section (rows 3-6).
        
        Row 3: Revenue
          - B3: =Historical!$F$3*(1+Assumptions!C7)
          - C3: =B3*(1+Assumptions!D7)
          - etc.
        Row 4: Cost Of Revenue
          - B4: =B3*(1-Assumptions!C9)  [using Gross Margin %]
          - C4: =C3*(1-Assumptions!D9)
          - etc.
        Row 5: Gross Profit = Revenue - COGS
        Row 6: Gross Margin % = Gross Profit / Revenue
        """
        # Row 3: Revenue
        ws.cell(row=3, column=1, value="Revenue")
        ws.cell(row=3, column=1).font = Font(bold=True)
        
        # B3: FY1 Revenue = Historical F3 * (1 + Assumptions C7)
        ws.cell(row=3, column=2, value='=Historical!$F$3*(1+Assumptions!C7)')
        ws.cell(row=3, column=2).number_format = ExcelFormats.CURRENCY
        
        # C3-F3: FY2-FY5 Revenue = Prior Year * (1 + Growth Rate)
        for i in range(1, self.projection_years):
            col = 2 + i  # columns 3, 4, 5, 6 = C, D, E, F
            col_letter = chr(64 + col)  # C, D, E, F (64+3=67=C, 64+4=68=D, etc.)
            prior_col_letter = chr(64 + col - 1)  # B, C, D, E
            assumptions_col = chr(67 + i)  # D, E, F, G (for FY2-FY5 in Assumptions row 7)
            
            formula = f'={prior_col_letter}3*(1+Assumptions!{assumptions_col}7)'
            ws.cell(row=3, column=col, value=formula)
            ws.cell(row=3, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 4: Cost Of Revenue = Revenue * (1 - Gross Margin%)
        ws.cell(row=4, column=1, value="Cost Of Revenue")
        for i in range(self.projection_years):
            col = 2 + i  # columns 2, 3, 4, 5, 6 = B, C, D, E, F
            col_letter = chr(64 + col)  # B, C, D, E, F (64+2=66=B, 64+3=67=C, etc.)
            assumptions_col = chr(67 + i)  # C, D, E, F, G (for FY1-FY5 in Assumptions row 9)
            
            formula = f'={col_letter}3*(1-Assumptions!{assumptions_col}9)'
            ws.cell(row=4, column=col, value=formula)
            ws.cell(row=4, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 5: Gross Profit = Revenue - COGS
        ws.cell(row=5, column=1, value="Gross Profit")
        ws.cell(row=5, column=1).font = Font(bold=True)
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            
            formula = f'={col_letter}3-{col_letter}4'
            ws.cell(row=5, column=col, value=formula)
            ws.cell(row=5, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 6: Gross Margin % = Gross Profit / Revenue
        ws.cell(row=6, column=1, value="Gross Margin %")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            
            formula = f'=IFERROR({col_letter}5/{col_letter}3,"")'
            ws.cell(row=6, column=col, value=formula)
            ws.cell(row=6, column=col).number_format = '0.00%'
    
    def _setup_opex_section(self, ws: Worksheet) -> None:
        """
        Set up Operating Expenses section (rows 7-9).
        
        Row 7: R&D = Historical!$F$6 * (Revenue / Historical!$F$3)
          - Scales R&D by revenue intensity from Historical FY0
        Row 8: SG&A = Historical!$F$7 * (Revenue / Historical!$F$3)
          - Scales SG&A by revenue intensity from Historical FY0
        Row 9: Operating Income (EBIT) = Gross Profit - R&D - SG&A
        """
        # Row 7: R&D
        ws.cell(row=7, column=1, value="R&D")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            
            formula = f'=Historical!$F$6*({col_letter}3/Historical!$F$3)'
            ws.cell(row=7, column=col, value=formula)
            ws.cell(row=7, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 8: SG&A
        ws.cell(row=8, column=1, value="SG&A")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            
            formula = f'=Historical!$F$7*({col_letter}3/Historical!$F$3)'
            ws.cell(row=8, column=col, value=formula)
            ws.cell(row=8, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 9: Operating Income (EBIT)
        ws.cell(row=9, column=1, value="Operating Income (EBIT)")
        ws.cell(row=9, column=1).font = Font(bold=True)
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            
            formula = f'={col_letter}5-{col_letter}7-{col_letter}8'
            ws.cell(row=9, column=col, value=formula)
            ws.cell(row=9, column=col).number_format = ExcelFormats.CURRENCY
    
    def _setup_tax_section(self, ws: Worksheet) -> None:
        """
        Set up Tax and NOPAT section (rows 10-11).
        
        Row 10: Tax Expense = MAX(0, EBIT) * Assumptions!$B$20
          - Uses effective tax rate from Assumptions (FY0 basis)
        Row 11: NOPAT = EBIT - Tax Expense
        """
        # Row 10: Tax Expense
        ws.cell(row=10, column=1, value="Tax Expense")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            
            formula = f'=MAX(0,{col_letter}9)*Assumptions!$B$20'
            ws.cell(row=10, column=col, value=formula)
            ws.cell(row=10, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 11: NOPAT
        ws.cell(row=11, column=1, value="NOPAT")
        ws.cell(row=11, column=1).font = Font(bold=True)
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            
            formula = f'={col_letter}9-{col_letter}10'
            ws.cell(row=11, column=col, value=formula)
            ws.cell(row=11, column=col).number_format = ExcelFormats.CURRENCY
    
    def _setup_capex_section(self, ws: Worksheet) -> None:
        """
        Set up D&A and Capex section (rows 12-13).
        
        Row 12: Depreciation & Amortization = Historical!$F$18 * (Revenue / Historical!$F$3)
          - Scales D&A by revenue intensity
        Row 13: Capital Expenditure = Historical!$F$24 * (Revenue / Historical!$F$3)
          - Scales Capex by revenue intensity (keeps negative sign)
        """
        # Row 12: Depreciation & Amortization
        ws.cell(row=12, column=1, value="Depreciation & Amortization")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            
            formula = f'=Historical!$F$18*({col_letter}3/Historical!$F$3)'
            ws.cell(row=12, column=col, value=formula)
            ws.cell(row=12, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 13: Capital Expenditure (negative)
        ws.cell(row=13, column=1, value="Capital Expenditure")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            
            formula = f'=Historical!$F$24*({col_letter}3/Historical!$F$3)'
            ws.cell(row=13, column=col, value=formula)
            ws.cell(row=13, column=col).number_format = ExcelFormats.CURRENCY
    
    def _setup_nwc_section(self, ws: Worksheet) -> None:
        """
        Set up Working Capital section (rows 14-18).
        
        Row 14: Accounts Receivable = Revenue / 365 * DSO
          - B14: =B3/365*Assumptions!C13
          - C14: =C3/365*Assumptions!D13, etc.
        Row 15: Inventory = COGS / 365 * DIO
          - B15: =B4/365*Assumptions!C14
          - C15: =C4/365*Assumptions!D14, etc.
        Row 16: Accounts Payable = COGS / 365 * DPO
          - B16: =B4/365*Assumptions!C15
          - C16: =C4/365*Assumptions!D15, etc.
        Row 17: Net Working Capital (NWC) = AR + Inventory - AP
        Row 18: ΔNWC = Change in NWC
          - B18: =B17 - (Historical!$F$32 + Historical!$F$33 - Historical!$F$34)
          - C18: =C17 - B17, etc.
        """
        # Row 14: Accounts Receivable (AR)
        ws.cell(row=14, column=1, value="Accounts Receivable (AR)")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            assumptions_col = chr(67 + i)  # C, D, E, F, G (for FY1-FY5 DSO in Assumptions row 13)
            
            formula = f'={col_letter}3/365*Assumptions!{assumptions_col}13'
            ws.cell(row=14, column=col, value=formula)
            ws.cell(row=14, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 15: Inventory (Inv)
        ws.cell(row=15, column=1, value="Inventory (Inv)")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            assumptions_col = chr(67 + i)  # C, D, E, F, G (for FY1-FY5 DIO in Assumptions row 14)
            
            formula = f'={col_letter}4/365*Assumptions!{assumptions_col}14'
            ws.cell(row=15, column=col, value=formula)
            ws.cell(row=15, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 16: Accounts Payable (AP)
        ws.cell(row=16, column=1, value="Accounts Payable (AP)")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            assumptions_col = chr(67 + i)  # C, D, E, F, G (for FY1-FY5 DPO in Assumptions row 15)
            
            formula = f'={col_letter}4/365*Assumptions!{assumptions_col}15'
            ws.cell(row=16, column=col, value=formula)
            ws.cell(row=16, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 17: Net Working Capital (NWC)
        ws.cell(row=17, column=1, value="Net Working Capital (NWC)")
        ws.cell(row=17, column=1).font = Font(bold=True)
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            
            formula = f'={col_letter}14+{col_letter}15-{col_letter}16'
            ws.cell(row=17, column=col, value=formula)
            ws.cell(row=17, column=col).number_format = ExcelFormats.CURRENCY
        
        # Row 18: ΔNWC (Change in NWC)
        ws.cell(row=18, column=1, value="ΔNWC")
        ws.cell(row=18, column=1).font = Font(bold=True)
        
        # B18: FY1 ΔNWC = Current NWC - Historical NWC (AR + Inv - AP from Historical F32, F33, F34)
        ws.cell(row=18, column=2, value='=B17-(Historical!$F$32+Historical!$F$33-Historical!$F$34)')
        ws.cell(row=18, column=2).number_format = ExcelFormats.CURRENCY
        
        # C18-F18: FY2-FY5 ΔNWC = Current NWC - Prior Year NWC
        for i in range(1, self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # C, D, E, F
            prior_col_letter = chr(64 + col - 1)  # B, C, D, E
            
            formula = f'={col_letter}17-{prior_col_letter}17'
            ws.cell(row=18, column=col, value=formula)
            ws.cell(row=18, column=col).number_format = ExcelFormats.CURRENCY
    
    def _setup_fcf_row(self, ws: Worksheet) -> None:
        """
        Set up FCF calculation row (row 19).
        
        Row 19: Free Cash Flow = NOPAT + D&A + Capex - ΔNWC
          - B19: =B11 + B12 + B13 - B18
          - C19: =C11 + C12 + C13 - C18
          - etc.
        
        This is unlevered FCF ready for DCF valuation.
        """
        # Row 19: Free Cash Flow
        ws.cell(row=19, column=1, value="Free Cash Flow")
        ws.cell(row=19, column=1).font = Font(bold=True, size=11)
        
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)  # B, C, D, E, F
            
            # FCF = NOPAT + D&A + Capex - ΔNWC
            formula = f'={col_letter}11+{col_letter}12+{col_letter}13-{col_letter}18'
            cell = ws.cell(row=19, column=col, value=formula)
            cell.number_format = ExcelFormats.CURRENCY
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=ExcelFormats.IMPORTANT_COLOR,
                end_color=ExcelFormats.IMPORTANT_COLOR,
                fill_type="solid"
            )
    
    def _setup_analytics_section(self, ws: Worksheet) -> None:
        """
        Set up professional analytics and diagnostic rows (rows 20-41).
        
        These rows enhance model auditability and provide key metrics:
        - EBITDA & EBITDA Margin
        - Operational metrics (R&D%, SG&A%, D&A%, Capex% as % of Revenue)
        - Working capital transparency (ΔAR, ΔInventory, ΔAP)
        - Explicit drivers (DSO, DIO, DPO)
        - PP&E roll-forward
        - Key ratios (FCF Margin, Reinvestment Rate, NWC % of Revenue)
        """
        current_row = 20
        
        # Section header
        ws.cell(row=current_row, column=1, value="ANALYTICS & DIAGNOSTICS")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=11, underline="single")
        current_row += 1
        
        # EBITDA Section
        ws.cell(row=current_row, column=1, value="EBITDA")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            formula = f'={col_letter}9+{col_letter}12'  # EBIT + D&A
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = ExcelFormats.CURRENCY
        current_row += 1
        
        # EBITDA Margin %
        ws.cell(row=current_row, column=1, value="EBITDA Margin %")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            formula = f'=IFERROR({col_letter}{current_row-1}/{col_letter}3,"")'  # EBITDA / Revenue
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '0.00%'
        current_row += 1
        
        # Blank row separator
        current_row += 1
        
        # Operational Metrics Section
        ws.cell(row=current_row, column=1, value="Operational Metrics (% of Revenue)")
        ws.cell(row=current_row, column=1).font = Font(bold=True, italic=True)
        current_row += 1
        
        # R&D % of Revenue
        ws.cell(row=current_row, column=1, value="R&D % of Revenue")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            formula = f'=IFERROR({col_letter}7/{col_letter}3,"")'  # R&D / Revenue
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '0.00%'
        current_row += 1
        
        # SG&A % of Revenue
        ws.cell(row=current_row, column=1, value="SG&A % of Revenue")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            formula = f'=IFERROR({col_letter}8/{col_letter}3,"")'  # SG&A / Revenue
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '0.00%'
        current_row += 1
        
        # EBIT Margin %
        ws.cell(row=current_row, column=1, value="EBIT Margin %")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            formula = f'=IFERROR({col_letter}9/{col_letter}3,"")'  # EBIT / Revenue
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '0.00%'
        current_row += 1
        
        # D&A % of Revenue
        ws.cell(row=current_row, column=1, value="D&A % of Revenue")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            formula = f'=IFERROR({col_letter}12/{col_letter}3,"")'  # D&A / Revenue
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '0.00%'
        current_row += 1
        
        # Capex % of Revenue
        ws.cell(row=current_row, column=1, value="Capex % of Revenue")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            formula = f'=IFERROR({col_letter}13/{col_letter}3,"")'  # Capex / Revenue
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '0.00%'
        current_row += 1
        
        # Blank row separator
        current_row += 1
        
        # Working Capital Drivers Section
        ws.cell(row=current_row, column=1, value="Working Capital Drivers")
        ws.cell(row=current_row, column=1).font = Font(bold=True, italic=True)
        current_row += 1
        
        # DSO (Days Sales Outstanding)
        ws.cell(row=current_row, column=1, value="DSO (days)")
        for i in range(self.projection_years):
            col = 2 + i
            assumptions_col = chr(67 + i)  # C, D, E, F, G
            formula = f'=Assumptions!{assumptions_col}13'
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '0.0'
        current_row += 1
        
        # DIO (Days Inventory Outstanding)
        ws.cell(row=current_row, column=1, value="DIO (days)")
        for i in range(self.projection_years):
            col = 2 + i
            assumptions_col = chr(67 + i)  # C, D, E, F, G
            formula = f'=Assumptions!{assumptions_col}14'
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '0.0'
        current_row += 1
        
        # DPO (Days Payable Outstanding)
        ws.cell(row=current_row, column=1, value="DPO (days)")
        for i in range(self.projection_years):
            col = 2 + i
            assumptions_col = chr(67 + i)  # C, D, E, F, G
            formula = f'=Assumptions!{assumptions_col}15'
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '0.0'
        current_row += 1
        
        # Blank row separator
        current_row += 1
        
        # Working Capital Changes Section
        ws.cell(row=current_row, column=1, value="Working Capital Changes (Transparency)")
        ws.cell(row=current_row, column=1).font = Font(bold=True, italic=True)
        current_row += 1
        
        # ΔAR (Change in Accounts Receivable)
        ws.cell(row=current_row, column=1, value="ΔAR")
        # B column: empty (first year, no prior period)
        ws.cell(row=current_row, column=2, value='=""')
        # C-F columns: current year AR - prior year AR
        for i in range(1, self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            prior_col_letter = chr(64 + col - 1)
            formula = f'={col_letter}14-{prior_col_letter}14'
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = ExcelFormats.CURRENCY
        current_row += 1
        
        # ΔInventory (Change in Inventory)
        ws.cell(row=current_row, column=1, value="ΔInventory")
        ws.cell(row=current_row, column=2, value='=""')
        for i in range(1, self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            prior_col_letter = chr(64 + col - 1)
            formula = f'={col_letter}15-{prior_col_letter}15'
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = ExcelFormats.CURRENCY
        current_row += 1
        
        # ΔAP (Change in Accounts Payable)
        ws.cell(row=current_row, column=1, value="ΔAP")
        ws.cell(row=current_row, column=2, value='=""')
        for i in range(1, self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            prior_col_letter = chr(64 + col - 1)
            formula = f'={col_letter}16-{prior_col_letter}16'
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = ExcelFormats.CURRENCY
        current_row += 1
        
        # NWC % of Revenue
        ws.cell(row=current_row, column=1, value="NWC % of Revenue")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            formula = f'=IFERROR({col_letter}17/{col_letter}3,"")'  # NWC / Revenue
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '0.00%'
        current_row += 1
        
        # Blank row separator
        current_row += 1
        
        # PP&E Roll-forward Section
        ws.cell(row=current_row, column=1, value="PP&E Roll-forward")
        ws.cell(row=current_row, column=1).font = Font(bold=True, italic=True)
        current_row += 1
        
        # Beginning Net PP&E
        ws.cell(row=current_row, column=1, value="Beginning Net PP&E")
        # B column: from Historical F25 (Net PP&E row in Historical tab)
        ws.cell(row=current_row, column=2, value='=Historical!$F$25')
        ws.cell(row=current_row, column=2).number_format = ExcelFormats.CURRENCY
        # C-F columns: prior year's Ending PP&E
        for i in range(1, self.projection_years):
            col = 2 + i
            prior_col_letter = chr(64 + col - 1)
            formula = f'={prior_col_letter}{current_row+3}'  # Reference Ending PP&E from 3 rows below
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = ExcelFormats.CURRENCY
        current_row += 1
        
        # + Capex
        ws.cell(row=current_row, column=1, value="+ Capex")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            formula = f'={col_letter}13'  # Reference Capex from row 13
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = ExcelFormats.CURRENCY
        current_row += 1
        
        # - D&A
        ws.cell(row=current_row, column=1, value="- D&A")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            formula = f'=-{col_letter}12'  # Negative of D&A from row 12
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = ExcelFormats.CURRENCY
        current_row += 1
        
        # Ending Net PP&E
        ws.cell(row=current_row, column=1, value="Ending Net PP&E")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            # Beginning + Capex - D&A
            formula = f'={col_letter}{current_row-3}+{col_letter}{current_row-2}+{col_letter}{current_row-1}'
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = ExcelFormats.CURRENCY
        current_row += 1
        
        # Blank row separator
        current_row += 1
        
        # Key Ratios Section
        ws.cell(row=current_row, column=1, value="Key Ratios")
        ws.cell(row=current_row, column=1).font = Font(bold=True, italic=True)
        current_row += 1
        
        # FCF Margin %
        ws.cell(row=current_row, column=1, value="FCF Margin %")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            formula = f'=IFERROR({col_letter}19/{col_letter}3,"")'  # FCF / Revenue
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '0.00%'
        current_row += 1
        
        # Reinvestment Rate
        ws.cell(row=current_row, column=1, value="Reinvestment Rate")
        for i in range(self.projection_years):
            col = 2 + i
            col_letter = chr(64 + col)
            formula = f'=IFERROR(({col_letter}13+{col_letter}18)/{col_letter}11,"")'  # (Capex + ΔNWC) / NOPAT
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '0.00%'
        current_row += 1
    
    def _format_sheet(self, ws: Worksheet) -> None:
        """Apply general formatting to the sheet"""
        # Set column widths
        ws.column_dimensions['A'].width = 25
        for col_letter in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col_letter].width = 16
        
        # Freeze headers
        ws.freeze_panes = ws['B2']
