"""
Tab 0: Raw Data Tab Builder

This module handles the creation of the Raw tab - a flat database structure
that stores all financial data from JSON in (Statement, Field, Year, Value) format.
"""

from typing import Dict, Any, List, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment


class RawTabBuilder:
    """
    Builds the Raw tab - a comprehensive flat database of all JSON data.
    
    Structure: Statement | Field | Year | Value
    
    This tab serves as the source of truth for all financial data,
    populated from JSON and queryable by other tabs.
    """
    
    def __init__(self):
        self.data_rows: List[Dict[str, Any]] = []
    
    def create_tab(self, workbook: openpyxl.Workbook) -> Worksheet:
        """
        Create and format the Raw tab.
        
        Args:
            workbook: The Excel workbook to add the tab to
            
        Returns:
            The created worksheet
        """
        # Create or get the sheet
        if "Raw" in workbook.sheetnames:
            ws = workbook["Raw"]
            # Clear existing data
            workbook.remove(ws)
        
        ws = workbook.create_sheet("Raw", 0)  # Position 0 (first tab)
        
        # Set up headers
        self._setup_headers(ws)
        
        # If we have data, populate it
        if self.data_rows:
            self._populate_data(ws)
        
        # Format the sheet
        self._format_sheet(ws)
        
        return ws
    
    def _setup_headers(self, ws: Worksheet) -> None:
        """Set up column headers for the Raw tab"""
        headers = ["Statement", "Field", "Year", "Value"]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 35  # Statement
        ws.column_dimensions['B'].width = 50  # Field
        ws.column_dimensions['C'].width = 15  # Year
        ws.column_dimensions['D'].width = 20  # Value
    
    def _populate_data(self, ws: Worksheet) -> None:
        """Populate the Raw tab with data rows"""
        for row_idx, data_row in enumerate(self.data_rows, start=2):
            ws.cell(row=row_idx, column=1, value=data_row['statement'])
            ws.cell(row=row_idx, column=2, value=data_row['field'])
            ws.cell(row=row_idx, column=3, value=data_row['year'])
            
            value = data_row['value']
            
            # Convert non-primitive values to strings
            if isinstance(value, (dict, list)):
                value = str(value)
            elif value is None:
                value = ''
            
            ws.cell(row=row_idx, column=4, value=value)
            
            # Format numbers
            if isinstance(value, (int, float)):
                ws.cell(row=row_idx, column=4).number_format = '#,##0.00'
    
    def _format_sheet(self, ws: Worksheet) -> None:
        """Apply general formatting to the sheet"""
        # Freeze the header row
        ws.freeze_panes = ws['A2']
    
    def add_data_from_json(self, json_data: Dict[str, Any]) -> None:
        """
        Parse JSON data and convert to (Statement, Field, Year, Value) rows.
        
        Args:
            json_data: The complete financial data JSON
        """
        self.data_rows = []
        
        # Process financial statements
        self.data_rows.extend(self._flatten_financial_statements(json_data))
        
        # Process company data
        self.data_rows.extend(self._flatten_company_data(json_data))
        
        # Process modeling metrics
        self.data_rows.extend(self._flatten_modeling_metrics(json_data))
        
        # Process market data
        self.data_rows.extend(self._flatten_market_data(json_data))
    
    def _flatten_financial_statements(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Flatten financial statements into rows.
        Each row: {statement, field, year, value}
        """
        rows = []
        
        if 'financial_statements' not in data:
            return rows
        
        fs = data['financial_statements']
        
        # Map statement types to friendly names
        statement_names = {
            'income_statement': 'Income Statement',
            'balance_sheet': 'Balance Sheet',
            'cash_flow': 'Cash Flow Statement'
        }
        
        for stmt_type, stmt_name in statement_names.items():
            if stmt_type not in fs:
                continue
                
            stmt_data = fs[stmt_type]
            
            if not isinstance(stmt_data, dict):
                continue
            
            # Get all years (most recent first)
            years = sorted(stmt_data.keys(), reverse=True)
            
            if not years:
                continue
            
            # Get all unique fields across all years
            all_fields = set()
            for year in years:
                if isinstance(stmt_data[year], dict):
                    all_fields.update(stmt_data[year].keys())
            
            # Create a row for each field-year combination
            for field in sorted(all_fields):
                for year in years:
                    value = None
                    if isinstance(stmt_data[year], dict):
                        value = stmt_data[year].get(field)
                    
                    # Only add if value exists
                    if value is not None:
                        rows.append({
                            'statement': stmt_name,
                            'field': field,
                            'year': year,
                            'value': value
                        })
        
        return rows
    
    def _flatten_company_data(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Flatten company data into rows.
        """
        rows = []
        
        if 'company_data' not in data:
            return rows
        
        company_data = data['company_data']
        
        for section, section_data in company_data.items():
            if not isinstance(section_data, dict):
                continue
            
            section_name = section.replace('_', ' ').title()
            
            for field, value in section_data.items():
                if value is not None:
                    rows.append({
                        'statement': f'Company Data - {section_name}',
                        'field': field,
                        'year': 'Current',
                        'value': value
                    })
        
        return rows
    
    def _flatten_modeling_metrics(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Flatten modeling metrics into rows.
        """
        rows = []
        
        if 'modeling_metrics' not in data:
            return rows
        
        mm = data['modeling_metrics']
        
        for metric_type, metric_data in mm.items():
            if not isinstance(metric_data, dict):
                continue
            
            metric_name = metric_type.replace('_', ' ').title()
            
            for field, value in metric_data.items():
                if value is not None:
                    rows.append({
                        'statement': f'Modeling Metrics - {metric_name}',
                        'field': field,
                        'year': 'Calculated',
                        'value': value
                    })
        
        return rows
    
    def _flatten_market_data(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Flatten market data (skip historical prices - too large).
        """
        rows = []
        
        if 'market_data' not in data:
            return rows
        
        md = data['market_data']
        
        for key, value in md.items():
            if key == 'historical_prices':
                # Skip historical prices - too much data
                continue
            
            if value is not None:
                rows.append({
                    'statement': 'Market Data',
                    'field': key,
                    'year': 'Current',
                    'value': value
                })
        
        return rows
    
    def get_data_summary(self) -> Dict[str, Any]:
        """
        Get a summary of the data in the Raw tab.
        
        Returns:
            Dictionary with statistics about the data
        """
        if not self.data_rows:
            return {
                'total_rows': 0,
                'statements': [],
                'years': [],
                'non_null_values': 0
            }
        
        statements = sorted(set(row['statement'] for row in self.data_rows))
        years = sorted(set(row['year'] for row in self.data_rows if row['year'] != 'Current' and row['year'] != 'Calculated'))
        non_null = sum(1 for row in self.data_rows if row['value'] is not None and row['value'] != '')
        
        return {
            'total_rows': len(self.data_rows),
            'statements': statements,
            'years': years,
            'non_null_values': non_null
        }
