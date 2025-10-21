"""
Tab 1: Keys_Map Tab Builder

This module creates the Keys_Map tab - a reference list of all fields
parsed from the JSON and stored in the Raw tab.

Structure: Statement | Field | Description
"""

from typing import Dict, Any, List, Set, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment


class KeysMapTabBuilder:
    """
    Builds the Keys_Map tab - a reference list of all available fields.
    
    This tab shows which fields are available in each statement type,
    serving as a data dictionary for the model.
    """
    
    def __init__(self):
        self.field_mappings: List[Dict[str, str]] = []
    
    def create_tab(self, workbook: openpyxl.Workbook) -> Worksheet:
        """
        Create and format the Keys_Map tab.
        
        Args:
            workbook: The Excel workbook to add the tab to
            
        Returns:
            The created worksheet
        """
        # Create or get the sheet
        if "Keys_Map" in workbook.sheetnames:
            ws = workbook["Keys_Map"]
            workbook.remove(ws)
        
        ws = workbook.create_sheet("Keys_Map", 1)  # Position 1 (second tab)
        
        # Set up headers
        self._setup_headers(ws)
        
        # Populate data
        if self.field_mappings:
            self._populate_data(ws)
        
        # Format the sheet
        self._format_sheet(ws)
        
        return ws
    
    def _setup_headers(self, ws: Worksheet) -> None:
        """Set up column headers for the Keys_Map tab"""
        headers = ["Statement", "Field", "Data Type"]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 35  # Statement
        ws.column_dimensions['B'].width = 50  # Field
        ws.column_dimensions['C'].width = 15  # Data Type
    
    def _populate_data(self, ws: Worksheet) -> None:
        """Populate the Keys_Map tab with field mappings"""
        for row_idx, mapping in enumerate(self.field_mappings, start=2):
            ws.cell(row=row_idx, column=1, value=mapping['statement'])
            ws.cell(row=row_idx, column=2, value=mapping['field'])
            ws.cell(row=row_idx, column=3, value=mapping.get('data_type', 'Numeric'))
    
    def _format_sheet(self, ws: Worksheet) -> None:
        """Apply general formatting to the sheet"""
        # Freeze the header row
        ws.freeze_panes = 'A2'
    
    def build_from_raw_data(self, raw_data_rows: List[Dict[str, Any]]) -> None:
        """
        Build the Keys_Map from the Raw tab data rows.
        
        This creates a unique list of all (Statement, Field) combinations.
        
        Args:
            raw_data_rows: List of rows from RawTabBuilder.data_rows
        """
        # Extract unique (statement, field) combinations
        unique_combinations: Set[Tuple[str, str]] = set()
        
        for row in raw_data_rows:
            statement = row['statement']
            field = row['field']
            unique_combinations.add((statement, field))
        
        # Convert to list and sort
        sorted_combinations = sorted(unique_combinations)
        
        # Build field mappings
        self.field_mappings = []
        
        for statement, field in sorted_combinations:
            # Infer data type based on field name
            data_type = self._infer_data_type(field)
            
            self.field_mappings.append({
                'statement': statement,
                'field': field,
                'data_type': data_type
            })
    
    def _infer_data_type(self, field_name: str) -> str:
        """
        Infer the data type based on field name patterns.
        
        Args:
            field_name: The field name
            
        Returns:
            Inferred data type ('Numeric', 'Text', 'Percent', etc.)
        """
        field_lower = field_name.lower()
        
        # Text fields
        if any(keyword in field_lower for keyword in ['name', 'sector', 'industry', 'symbol', 'ticker', 'exchange']):
            return 'Text'
        
        # Percentage fields
        if any(keyword in field_lower for keyword in ['rate', 'margin', 'ratio', 'growth', 'yield', 'percent']):
            return 'Percent'
        
        # Date fields
        if any(keyword in field_lower for keyword in ['date', 'year', 'period']):
            return 'Date'
        
        # Default to numeric
        return 'Numeric'
    
    def get_summary(self) -> Dict[str, Any]:
        """
        Get a summary of the Keys_Map tab.
        
        Returns:
            Dictionary with statistics
        """
        if not self.field_mappings:
            return {
                'total_fields': 0,
                'by_statement': {},
                'by_data_type': {}
            }
        
        # Count by statement
        by_statement: Dict[str, int] = {}
        for mapping in self.field_mappings:
            stmt = mapping['statement']
            by_statement[stmt] = by_statement.get(stmt, 0) + 1
        
        # Count by data type
        by_data_type: Dict[str, int] = {}
        for mapping in self.field_mappings:
            dtype = mapping['data_type']
            by_data_type[dtype] = by_data_type.get(dtype, 0) + 1
        
        return {
            'total_fields': len(self.field_mappings),
            'by_statement': by_statement,
            'by_data_type': by_data_type
        }
