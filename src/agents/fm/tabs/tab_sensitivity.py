"""
Tab 8: Sensitivity Tab Builder

This module creates the Sensitivity tab - professional banker-grade sensitivity analysis
with Mid-Year Discounting (MYD) toggle and dual DCF method sensitivity tables.

Professional features:
- Mid-Year Discounting toggle (Yes/No dropdown)
- Perpetual Growth DCF sensitivity: WACC × Terminal Growth (g)
- Exit Multiple DCF sensitivity: WACC × Exit Multiple
- Summary comparison of both methods
- Cross-references to both Valuation tabs
"""

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

from ..financial_model_builder import ExcelFormats


class SensitivityTabBuilder:
    """
    Builds the Sensitivity tab - dual DCF sensitivity analysis with MYD toggle.
    
    This tab contains:
    - Section A: Mid-Year Discounting Toggle & Inputs (rows 2-9)
    - Section B: Perpetual Growth DCF Sensitivity Table - WACC × g (rows 12-17)
    - Section C: Exit Multiple DCF Sensitivity Table - WACC × Exit Multiple (rows 22-27)
    - Section D: Summary Block (rows 32-38)
    
    The MYD toggle affects discount factor exponents in both DCF tabs.
    """
    
    def __init__(self):
        """Initialize the Sensitivity builder."""
        pass
    
    def create_tab(self, workbook: openpyxl.Workbook) -> Worksheet:
        """
        Create and format the Sensitivity tab.
        
        Args:
            workbook: The Excel workbook to add the tab to
            
        Returns:
            The created worksheet
        """
        # Create or get the sheet
        if "Sensitivity" in workbook.sheetnames:
            ws = workbook["Sensitivity"]
            workbook.remove(ws)
        
        ws = workbook.create_sheet("Sensitivity", 8)  # Position 8 (ninth tab)
        
        # Build all sections following specification
        self._setup_header(ws)
        self._setup_myd_toggle_section(ws)
        self._setup_perpetual_growth_sensitivity(ws)
        self._setup_exit_multiple_sensitivity(ws)
        self._setup_summary_block(ws)
        
        # Format the sheet
        self._format_sheet(ws)
        
        return ws
    
    def _setup_header(self, ws: Worksheet) -> None:
        """Set up main header."""
        ws.cell(row=1, column=1, value="SENSITIVITY ANALYSIS - DUAL DCF METHODS")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
    
    def _setup_myd_toggle_section(self, ws: Worksheet) -> None:
        """
        Set up Section A: Mid-Year Discounting Toggle & Inputs (rows 2-9).
        
        Row 2: MYD Toggle (Yes/No dropdown)
        Row 3: Current Selection
        Row 4: MYD Factor (0.5 if Yes, 0 if No)
        Row 6: Base WACC
        Row 7: Terminal Growth g
        Row 8: Exit Multiple
        Row 9: Documentation note
        """
        # Row 2: MYD Toggle
        ws.cell(row=2, column=1, value="Mid-Year Discount Toggle")
        ws.cell(row=2, column=1).font = Font(bold=True)
        ws.cell(row=2, column=2, value="No")  # Default value
        ws.cell(row=2, column=2).fill = PatternFill(
            start_color=ExcelFormats.INPUT_COLOR,
            end_color=ExcelFormats.INPUT_COLOR,
            fill_type="solid"
        )
        
        # Add data validation (dropdown: Yes,No)
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
        dv.add(ws['B2'])
        ws.add_data_validation(dv)
        
        # Row 3: Current Selection
        ws.cell(row=3, column=1, value="Current Selection")
        ws.cell(row=3, column=2, value="=B2")
        
        # Row 4: MYD Factor
        ws.cell(row=4, column=1, value="Mid-Year Factor (years)")
        ws.cell(row=4, column=1).font = Font(bold=True)
        ws.cell(row=4, column=2, value='=IF(B3="Yes",0.5,0)')
        ws.cell(row=4, column=2).number_format = '0.0'
        ws.cell(row=4, column=2).font = Font(bold=True)
        ws.cell(row=4, column=2).fill = PatternFill(
            start_color="FFD700",  # Gold color for important calculated value
            end_color="FFD700",
            fill_type="solid"
        )
        
        # Blank row
        ws.cell(row=5, column=1, value="")
        
        # Row 6: Base WACC
        ws.cell(row=6, column=1, value="Base WACC (from DCF)")
        ws.cell(row=6, column=2, value="='Valuation (DCF)'!$B$12")
        ws.cell(row=6, column=2).number_format = '0.00%'
        
        # Row 7: Terminal Growth g
        ws.cell(row=7, column=1, value="Terminal Growth g (from DCF)")
        ws.cell(row=7, column=2, value="='Valuation (DCF)'!$B$23")
        ws.cell(row=7, column=2).number_format = '0.00%'
        
        # Row 8: Exit Multiple
        ws.cell(row=8, column=1, value="Exit Multiple (from Exit DCF)")
        ws.cell(row=8, column=2, value="='Valuation (Exit Multiple)'!$B$3")
        ws.cell(row=8, column=2).number_format = '0.0"x"'
        
        # Row 9: Documentation note
        ws.cell(row=9, column=1, value="Note (how MYD is used)")
        ws.cell(row=9, column=1).font = Font(italic=True, size=9)
        ws.cell(row=9, column=2, value='DF exponents in both DCF tabs use (... - Sensitivity!$B$4)')
        ws.cell(row=9, column=2).font = Font(italic=True, size=9)
        
        # Blank rows
        ws.cell(row=10, column=1, value="")
        ws.cell(row=11, column=1, value="")
    
    def _setup_perpetual_growth_sensitivity(self, ws: Worksheet) -> None:
        """
        Set up Section B: Perpetual Growth DCF Sensitivity Table (rows 12-17).
        
        WACC (rows) × Terminal Growth g (columns)
        Creates a 5×4 data table structure.
        """
        # Row 12: Headers
        ws.cell(row=12, column=1, value="WACC ↓ / g →")
        ws.cell(row=12, column=1).font = Font(bold=True, size=10)
        ws.cell(row=12, column=1).fill = PatternFill(
            start_color=ExcelFormats.HEADER_COLOR,
            end_color=ExcelFormats.HEADER_COLOR,
            fill_type="solid"
        )
        
        # B12: Anchor (value per share from Perpetual Growth DCF)
        ws.cell(row=12, column=2, value="='Valuation (DCF)'!$B$37")
        ws.cell(row=12, column=2).number_format = '$0.00'
        ws.cell(row=12, column=2).font = Font(bold=True)
        ws.cell(row=12, column=2).fill = PatternFill(
            start_color="FFD700",  # Gold
            end_color="FFD700",
            fill_type="solid"
        )
        
        # C12-E12: g scenarios (−0.5%, Base, +0.5%)
        ws.cell(row=12, column=3, value="=$B$7-0.005")  # g - 0.5%
        ws.cell(row=12, column=3).number_format = '0.00%'
        ws.cell(row=12, column=3).font = Font(bold=True, size=10)
        ws.cell(row=12, column=3).fill = PatternFill(
            start_color=ExcelFormats.HEADER_COLOR,
            end_color=ExcelFormats.HEADER_COLOR,
            fill_type="solid"
        )
        
        ws.cell(row=12, column=4, value="=$B$7")  # g Base
        ws.cell(row=12, column=4).number_format = '0.00%'
        ws.cell(row=12, column=4).font = Font(bold=True, size=10)
        ws.cell(row=12, column=4).fill = PatternFill(
            start_color=ExcelFormats.HEADER_COLOR,
            end_color=ExcelFormats.HEADER_COLOR,
            fill_type="solid"
        )
        
        ws.cell(row=12, column=5, value="=$B$7+0.005")  # g + 0.5%
        ws.cell(row=12, column=5).number_format = '0.00%'
        ws.cell(row=12, column=5).font = Font(bold=True, size=10)
        ws.cell(row=12, column=5).fill = PatternFill(
            start_color=ExcelFormats.HEADER_COLOR,
            end_color=ExcelFormats.HEADER_COLOR,
            fill_type="solid"
        )
        
        # Rows 13-17: WACC scenarios
        wacc_scenarios = [
            (13, "=$B$6-0.01"),   # WACC - 1.0%
            (14, "=$B$6-0.005"),  # WACC - 0.5%
            (15, "=$B$6"),        # WACC Base
            (16, "=$B$6+0.005"),  # WACC + 0.5%
            (17, "=$B$6+0.01"),   # WACC + 1.0%
        ]
        
        for row, formula in wacc_scenarios:
            ws.cell(row=row, column=1, value=formula)
            ws.cell(row=row, column=1).number_format = '0.00%'
            ws.cell(row=row, column=1).font = Font(bold=True, size=10)
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color=ExcelFormats.HEADER_COLOR,
                end_color=ExcelFormats.HEADER_COLOR,
                fill_type="solid"
            )
        
        # Data table area B13:E17 - will be filled by Excel Data Table
        # Add note about how to create the data table
        for row in range(13, 18):
            for col in range(2, 6):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '$0.00'
                cell.fill = PatternFill(
                    start_color="F0F0F0",
                    end_color="F0F0F0",
                    fill_type="solid"
                )
        
        # Add instruction note
        ws.cell(row=19, column=1, value="NOTE: Select B12:E17, Data > What-If Analysis > Data Table")
        ws.cell(row=19, column=1).font = Font(italic=True, size=9, color="666666")
        ws.cell(row=20, column=1, value="Row input: 'Valuation (DCF)'!$B$23, Column input: 'Valuation (DCF)'!$B$12")
        ws.cell(row=20, column=1).font = Font(italic=True, size=9, color="666666")
        
        # Blank row
        ws.cell(row=21, column=1, value="")
    
    def _setup_exit_multiple_sensitivity(self, ws: Worksheet) -> None:
        """
        Set up Section C: Exit Multiple DCF Sensitivity Table (rows 22-27).
        
        WACC (rows) × Exit Multiple (columns)
        Creates a 5×6 data table structure.
        """
        # Row 22: Headers
        ws.cell(row=22, column=1, value="WACC ↓ / Exit Multiple →")
        ws.cell(row=22, column=1).font = Font(bold=True, size=10)
        ws.cell(row=22, column=1).fill = PatternFill(
            start_color=ExcelFormats.HEADER_COLOR,
            end_color=ExcelFormats.HEADER_COLOR,
            fill_type="solid"
        )
        
        # B22: Anchor (value per share from Exit Multiple DCF)
        ws.cell(row=22, column=2, value="='Valuation (Exit Multiple)'!$B$25")
        ws.cell(row=22, column=2).number_format = '$0.00'
        ws.cell(row=22, column=2).font = Font(bold=True)
        ws.cell(row=22, column=2).fill = PatternFill(
            start_color="FFD700",  # Gold
            end_color="FFD700",
            fill_type="solid"
        )
        
        # C22-G22: Exit Multiple scenarios (−2×, −1×, Base, +1×, +2×)
        multiple_scenarios = [
            (3, "=$B$8-2"),   # Mult - 2x
            (4, "=$B$8-1"),   # Mult - 1x
            (5, "=$B$8"),     # Mult Base
            (6, "=$B$8+1"),   # Mult + 1x
            (7, "=$B$8+2"),   # Mult + 2x
        ]
        
        for col, formula in multiple_scenarios:
            ws.cell(row=22, column=col, value=formula)
            ws.cell(row=22, column=col).number_format = '0.0"x"'
            ws.cell(row=22, column=col).font = Font(bold=True, size=10)
            ws.cell(row=22, column=col).fill = PatternFill(
                start_color=ExcelFormats.HEADER_COLOR,
                end_color=ExcelFormats.HEADER_COLOR,
                fill_type="solid"
            )
        
        # Rows 23-27: WACC scenarios (same as perpetual growth table)
        wacc_scenarios = [
            (23, "=$B$6-0.01"),   # WACC - 1.0%
            (24, "=$B$6-0.005"),  # WACC - 0.5%
            (25, "=$B$6"),        # WACC Base
            (26, "=$B$6+0.005"),  # WACC + 0.5%
            (27, "=$B$6+0.01"),   # WACC + 1.0%
        ]
        
        for row, formula in wacc_scenarios:
            ws.cell(row=row, column=1, value=formula)
            ws.cell(row=row, column=1).number_format = '0.00%'
            ws.cell(row=row, column=1).font = Font(bold=True, size=10)
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color=ExcelFormats.HEADER_COLOR,
                end_color=ExcelFormats.HEADER_COLOR,
                fill_type="solid"
            )
        
        # Data table area B23:G27 - will be filled by Excel Data Table
        for row in range(23, 28):
            for col in range(2, 8):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '$0.00'
                cell.fill = PatternFill(
                    start_color="F0F0F0",
                    end_color="F0F0F0",
                    fill_type="solid"
                )
        
        # Add instruction note
        ws.cell(row=29, column=1, value="NOTE: Select B22:G27, Data > What-If Analysis > Data Table")
        ws.cell(row=29, column=1).font = Font(italic=True, size=9, color="666666")
        ws.cell(row=30, column=1, value="Row input: 'Valuation (Exit Multiple)'!$B$3, Column input: 'Valuation (Exit Multiple)'!$B$2")
        ws.cell(row=30, column=1).font = Font(italic=True, size=9, color="666666")
        
        # Blank row
        ws.cell(row=31, column=1, value="")
    
    def _setup_summary_block(self, ws: Worksheet) -> None:
        """
        Set up Section D: Summary Block (rows 32-38).
        
        Compares both DCF methods and shows upside vs market price.
        """
        # Row 32: Perpetual DCF Value/Share
        ws.cell(row=32, column=1, value="Perpetual DCF — Value / Share")
        ws.cell(row=32, column=1).font = Font(bold=True)
        ws.cell(row=32, column=2, value="='Valuation (DCF)'!$B$37")
        ws.cell(row=32, column=2).number_format = '$0.00'
        ws.cell(row=32, column=2).font = Font(bold=True)
        
        # Row 33: Exit Multiple DCF Value/Share
        ws.cell(row=33, column=1, value="Exit Multiple DCF — Value / Share")
        ws.cell(row=33, column=1).font = Font(bold=True)
        ws.cell(row=33, column=2, value="='Valuation (Exit Multiple)'!$B$25")
        ws.cell(row=33, column=2).number_format = '$0.00'
        ws.cell(row=33, column=2).font = Font(bold=True)
        
        # Row 34: Average of Methods
        ws.cell(row=34, column=1, value="Average of Methods")
        ws.cell(row=34, column=1).font = Font(bold=True, size=11)
        ws.cell(row=34, column=2, value="=AVERAGE(B32:B33)")
        ws.cell(row=34, column=2).number_format = '$0.00'
        ws.cell(row=34, column=2).font = Font(bold=True, size=11)
        ws.cell(row=34, column=2).fill = PatternFill(
            start_color="FFD700",  # Gold
            end_color="FFD700",
            fill_type="solid"
        )
        
        # Row 35: Current Market Price
        ws.cell(row=35, column=1, value="Current Market Price")
        ws.cell(row=35, column=2, value="=Historical!$F$2")
        ws.cell(row=35, column=2).number_format = '$0.00'
        
        # Row 36: Upside vs Market
        ws.cell(row=36, column=1, value="Upside vs Market")
        ws.cell(row=36, column=1).font = Font(bold=True, size=11)
        ws.cell(row=36, column=2, value="=B34/B35-1")
        ws.cell(row=36, column=2).number_format = '0.0%'
        ws.cell(row=36, column=2).font = Font(bold=True, size=11)
        ws.cell(row=36, column=2).fill = PatternFill(
            start_color="90EE90",  # Light green
            end_color="90EE90",
            fill_type="solid"
        )
        
        # Blank row
        ws.cell(row=37, column=1, value="")
        
        # Row 38: Mid-Year Toggle Active?
        ws.cell(row=38, column=1, value="Mid-Year Toggle Active?")
        ws.cell(row=38, column=1).font = Font(italic=True)
        ws.cell(row=38, column=2, value="=B2")
        ws.cell(row=38, column=2).font = Font(italic=True, bold=True)
    

    
    def _format_sheet(self, ws: Worksheet) -> None:
        """Apply general formatting to the sheet."""
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        for col_letter in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter].width = 14
        
        # Freeze panes at B2
        ws.freeze_panes = ws['B2']
