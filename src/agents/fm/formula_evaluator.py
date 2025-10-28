"""
Formula Evaluator - Compute Excel formulas without opening Excel

This module evaluates Excel formulas dynamically and generates a JSON file
containing all computed values for every tab and cell in the financial model.

Architecture:
- Respects tab dependency order (Raw → Keys_Map → ... → Summary)
- Evaluates formulas recursively using previously computed tab values
- Handles cell references, arithmetic operations, and Excel functions
- Outputs structured JSON with all computed values

Usage:
    from src.agents.fm.formula_evaluator import FormulaEvaluator
    
    evaluator = FormulaEvaluator(workbook)
    results = evaluator.evaluate_all_tabs()
    evaluator.save_to_json(results, "output_path.json")
"""
import re
import json
from typing import Dict, Any, List, Tuple, Optional, Union
from pathlib import Path
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
import math
from src.logger import get_logger

import json
import re
from typing import Dict, Any, List, Tuple, Optional, Union
from pathlib import Path
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
import math


class FormulaEvaluator:
    """
    Evaluates all Excel formulas in a financial model workbook.
    
    This class:
    1. Processes tabs in dependency order
    2. Evaluates formulas using previously computed values
    3. Handles cell references, ranges, and Excel functions
    4. Stores all computed values in JSON format
    """
    
    # Tab processing order (must match dependency chain)
    TAB_ORDER = [
        "Raw",
        "Keys_Map",
        "Assumptions",
        "LLM_Inferred",
        "Historical",
        "Projections",
        "Valuation (DCF)",
        "Valuation (Exit Multiple)",
        "Sensitivity",
        "Summary",
    ]
    
    def __init__(self, workbook: Workbook):
        """
        Initialize the Formula Evaluator.
        
        Args:
            workbook: The openpyxl Workbook to evaluate
        """
        self.workbook = workbook
        
        # Storage for computed values
        # Structure: {tab_name: {(row, col): value}}
        self.computed_values: Dict[str, Dict[Tuple[int, int], Any]] = {}
        
        # Cache for formula evaluation to avoid re-computing
        self.eval_cache: Dict[str, Any] = {}
        
        # Logger - will be set by pipeline if available
        self.logger = None

    def set_logger(self, logger):
        """Set the logger instance."""
        self.logger = logger
    
    def _log(self, level: str, message: str):
        """Log message using logger if available, otherwise print."""
        getattr(self.logger, level)(message)
    
    def evaluate_all_tabs(self) -> Dict[str, Dict[str, Any]]:
        """
        Evaluate all tabs in dependency order.
        
        Returns:
            Dictionary with structure:
            {
                "tab_name": {
                    "cells": {
                        "(row, col)": value,
                        ...
                    },
                    "metadata": {
                        "total_cells": int,
                        "formula_cells": int,
                        "value_cells": int
                    }
                }
            }
        """
        self._log("info", "="*70)
        self._log("info", "Starting Formula Evaluation")
        self._log("info", "="*70)

        results = {}
        
        for tab_name in self.TAB_ORDER:
            if tab_name not in self.workbook.sheetnames:
                self._log("warning", f"Skipping {tab_name} (not found in workbook)")
                continue

            self._log("info", f"[{self.TAB_ORDER.index(tab_name) + 1}/{len(self.TAB_ORDER)}] Evaluating {tab_name}...")

            tab_results = self._evaluate_tab(tab_name)
            results[tab_name] = tab_results
            
            # Store computed values for next tabs to reference
            self.computed_values[tab_name] = tab_results["cells"]

            self._log("info", f"      ✅ {tab_results['metadata']['total_cells']} cells evaluated")
            self._log("info", f"         • Formulas: {tab_results['metadata']['formula_cells']}")
            self._log("info", f"         • Values: {tab_results['metadata']['value_cells']}")

        self._log("info", "="*70)
        self._log("info", "✅ Formula Evaluation Complete!")
        self._log("info", "="*70)

        return results
    
    def _evaluate_tab(self, tab_name: str) -> Dict[str, Any]:
        """
        Evaluate all cells in a single tab.
        
        Args:
            tab_name: Name of the tab to evaluate
            
        Returns:
            Dictionary with cells and metadata
        """
        ws = self.workbook[tab_name]
        
        cells = {}
        formula_count = 0
        value_count = 0
        
        # Initialize computed_values for this tab if not exists
        if tab_name not in self.computed_values:
            self.computed_values[tab_name] = {}
        
        # Iterate through all cells with data
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                
                row_idx = cell.row
                col_idx = cell.column
                key = f"({row_idx}, {col_idx})"
                
                # Check if it's a formula
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    # Evaluate the formula
                    try:
                        evaluated_value = self._evaluate_formula(
                            cell.value[1:],  # Remove '=' prefix
                            tab_name,
                            row_idx,
                            col_idx
                        )
                        cells[key] = evaluated_value
                        # Store immediately in computed_values so subsequent formulas can reference it
                        self.computed_values[tab_name][key] = evaluated_value
                        formula_count += 1
                    except Exception as e:
                        # If formula evaluation fails, store error info
                        error_value = {
                            "error": str(e),
                            "formula": cell.value
                        }
                        cells[key] = error_value
                        self.computed_values[tab_name][key] = error_value
                        formula_count += 1
                else:
                    # It's a direct value
                    serialized_value = self._serialize_value(cell.value)
                    cells[key] = serialized_value
                    # Store immediately in computed_values so formulas can reference it
                    self.computed_values[tab_name][key] = serialized_value
                    value_count += 1
        
        return {
            "cells": cells,
            "metadata": {
                "total_cells": len(cells),
                "formula_cells": formula_count,
                "value_cells": value_count,
                "tab_name": tab_name
            }
        }
    
    def _evaluate_formula(
        self,
        formula: str,
        current_tab: str,
        current_row: int,
        current_col: int
    ) -> Any:
        """
        Recursively evaluate an Excel formula.
        
        Args:
            formula: The formula string (without '=' prefix)
            current_tab: The tab containing this formula
            current_row: Row of the cell with this formula
            current_col: Column of the cell with this formula
            
        Returns:
            The evaluated result
        """
        # Create cache key
        cache_key = f"{current_tab}!{get_column_letter(current_col)}{current_row}:{formula}"
        if cache_key in self.eval_cache:
            return self.eval_cache[cache_key]
        
        try:
            result = self._eval_expression(formula, current_tab, current_row, current_col)
            self.eval_cache[cache_key] = result
            return result
        except Exception as e:
            # Return error information
            return {
                "error": f"Evaluation error: {str(e)}",
                "formula": formula
            }
    
    def _eval_expression(
        self,
        expr: str,
        current_tab: str,
        current_row: int,
        current_col: int
    ) -> Any:
        """
        Evaluate a formula expression.
        
        Handles:
        - Cell references (A1, 'Tab'!B2, $A$1)
        - Arithmetic operations (+, -, *, /, ^)
        - Excel functions (SUM, IF, IFERROR, SUMIFS, etc.)
        - Ranges (A1:A10)
        """
        expr = expr.strip()
        
        # Handle comparison operators (must come before function check)
        # But only if the operator is at the top level, not inside parentheses
        has_top_level_comparison = False
        paren_depth = 0
        in_quotes = False
        
        for i, char in enumerate(expr):
            if char == '"':
                in_quotes = not in_quotes
            elif not in_quotes:
                if char == '(':
                    paren_depth += 1
                elif char == ')':
                    paren_depth -= 1
                elif paren_depth == 0:
                    # Check for comparison operators at this position
                    for op in ['<>', '>=', '<=', '>', '<', '=']:
                        if expr[i:i+len(op)] == op:
                            has_top_level_comparison = True
                            break
            if has_top_level_comparison:
                break
        
        if has_top_level_comparison:
            return self._evaluate_comparison(expr, current_tab, current_row, current_col)
        
        # Handle Excel functions
        if self._is_function_call(expr):
            return self._evaluate_function(expr, current_tab, current_row, current_col)
        
        # Handle cell references
        if self._is_cell_reference(expr):
            return self._get_cell_value(expr, current_tab)
        
        # Handle arithmetic expressions
        if any(op in expr for op in ['+', '-', '*', '/', '^', '(', ')', '&']):
            return self._evaluate_arithmetic(expr, current_tab, current_row, current_col)
        
        # Handle numeric literals
        try:
            # Try to parse as number
            if '.' in expr or 'E' in expr.upper():
                return float(expr)
            return int(expr)
        except ValueError:
            pass
        
        # Handle string literals
        if expr.startswith('"') and expr.endswith('"'):
            return expr[1:-1]
        
        # Handle boolean literals
        if expr.upper() == 'TRUE':
            return True
        if expr.upper() == 'FALSE':
            return False
        
        # If we can't parse it, return as string
        return expr
    
    def _is_function_call(self, expr: str) -> bool:
        """
        Check if expression is a PURE function call (not a composite expression).
        
        Examples:
        - SUM(A1:A10) -> True
        - IF(A1>0,1,0) -> True
        - MAX(0,B9)*Assumptions!$B$20 -> False (contains arithmetic after function)
        - A1+SUM(B1:B10) -> False (contains arithmetic before function)
        """
        if '(' not in expr:
            return False
        
        func_name = expr.split('(')[0].strip()
        if not func_name.replace('_', '').isalpha():
            return False
        
        # Check if there are any operators outside the function call
        # Find the matching closing parenthesis for the function
        paren_count = 0
        start_idx = expr.find('(')
        for i in range(start_idx, len(expr)):
            if expr[i] == '(':
                paren_count += 1
            elif expr[i] == ')':
                paren_count -= 1
                if paren_count == 0:
                    # Check if there's anything after the closing parenthesis
                    remaining = expr[i+1:].strip()
                    if remaining and remaining not in ['', ')']:
                        # There's content after the function call
                        return False
                    break
        
        return True
    
    def _is_cell_reference(self, expr: str) -> bool:
        """Check if expression is a cell reference"""
        # Remove $ signs first
        cleaned_expr = expr.replace('$', '').strip()
        
        # Pattern: [Tab!]A1 or [Tab!]A1:B5 (single cell or range)
        # Handle both with and without quotes around tab name
        # Tab name can be: 'Tab Name'! or TabName! or no tab
        # Changed [^!]+! to [A-Za-z0-9_ ]+! to only match valid tab name characters
        pattern = r"^(?:'[^']+'!|[A-Za-z0-9_ ]+!)?[A-Z]+\d+(?::[A-Z]+\d+)?$"
        return bool(re.match(pattern, cleaned_expr))
    
    def _get_cell_value(self, cell_ref: str, current_tab: str) -> Any:
        """
        Get the value of a cell reference.
        
        Args:
            cell_ref: Cell reference (e.g., "A1", "'Tab'!B2", "$A$1")
            current_tab: The current tab (used for relative references)
            
        Returns:
            The cell's computed value
        """
        # Remove $ signs (absolute references)
        cell_ref = cell_ref.replace('$', '')
        
        # Parse tab and cell
        if '!' in cell_ref:
            # Cross-tab reference
            parts = cell_ref.split('!')
            tab_name = parts[0].strip("'")
            cell_addr = parts[1]
        else:
            # Same-tab reference
            tab_name = current_tab
            cell_addr = cell_ref
        
        # Parse cell address
        match = re.match(r'([A-Z]+)(\d+)', cell_addr)
        if not match:
            raise ValueError(f"Invalid cell reference: {cell_ref}")
        
        col_letter = match.group(1)
        row_num = int(match.group(2))
        col_num = column_index_from_string(col_letter)
        
        # Look up in computed values
        if tab_name in self.computed_values:
            key = f"({row_num}, {col_num})"
            if key in self.computed_values[tab_name]:
                return self.computed_values[tab_name][key]
        
        # If not found in computed values, try to get from workbook
        if tab_name in self.workbook.sheetnames:
            ws = self.workbook[tab_name]
            cell = ws.cell(row=row_num, column=col_num)
            
            if cell.value is None:
                return 0  # Empty cell = 0 in Excel
            
            if isinstance(cell.value, str) and cell.value.startswith('='):
                # It's a formula - evaluate it recursively
                return self._evaluate_formula(
                    cell.value[1:],
                    tab_name,
                    row_num,
                    col_num
                )
            
            return cell.value
        
        # Default to 0 if not found
        return 0
    
    def _evaluate_arithmetic(
        self,
        expr: str,
        current_tab: str,
        current_row: int,
        current_col: int
    ) -> float:
        """
        Evaluate arithmetic expressions with cell references.
        
        Strategy:
        1. Replace all function calls with their results
        2. Replace all cell references with their values
        3. Handle string concatenation (&)
        4. Evaluate the resulting arithmetic expression
        """
        # Check if there's a & operator at the top level (not inside parentheses)
        # This handles string concatenation
        has_top_level_concat = False
        paren_depth = 0
        in_quotes = False
        for char in expr:
            if char == '"':
                in_quotes = not in_quotes
            elif not in_quotes:
                if char == '(':
                    paren_depth += 1
                elif char == ')':
                    paren_depth -= 1
                elif char == '&' and paren_depth == 0:
                    has_top_level_concat = True
                    break
        
        if has_top_level_concat:
            return self._evaluate_concatenation(expr, current_tab, current_row, current_col)
        
        # First, find and evaluate any function calls
        # Pattern to match function calls: FUNCTION_NAME(...)
        func_pattern = r'([A-Z_]+)\('
        matches = list(re.finditer(func_pattern, expr))
        
        # Process function calls from right to left to avoid index issues
        for match in reversed(matches):
            func_name = match.group(1)
            start_pos = match.start()
            
            # Find the matching closing parenthesis
            paren_count = 1
            pos = match.end()
            while pos < len(expr) and paren_count > 0:
                if expr[pos] == '(':
                    paren_count += 1
                elif expr[pos] == ')':
                    paren_count -= 1
                pos += 1
            
            if paren_count == 0:
                # Found complete function call
                func_call = expr[start_pos:pos]
                try:
                    # Evaluate the function
                    result = self._evaluate_function(func_call, current_tab, current_row, current_col)
                    # Replace function call with its result in the expression
                    expr = expr[:start_pos] + f'({result})' + expr[pos:]
                except Exception:
                    # If function evaluation fails, replace with 0
                    expr = expr[:start_pos] + '(0)' + expr[pos:]
        
        # Find all cell references in the expression
        # Pattern must handle: Tab!A1, 'Tab'!A1, $A$1, Tab!$A$1, 'Tab Name'!A1
        # Use a more specific pattern that captures the entire reference including tab name
        pattern = r"(?:'[^']+'!|[A-Za-z_][A-Za-z0-9_]*!)?\$?[A-Z]+\$?\d+"
        
        def replace_cell_ref(match):
            cell_ref = match.group(0)
            try:
                value = self._get_cell_value(cell_ref, current_tab)
                # Handle errors and non-numeric values
                if isinstance(value, dict) and 'error' in value:
                    return '0'
                if value is None or value == '':
                    return '0'
                # Convert boolean to int
                if isinstance(value, bool):
                    return '1' if value else '0'
                # Wrap in parentheses to maintain operator precedence
                return f'({value})'
            except Exception as e:
                return '0'
        
        # Replace all cell references with their values
        evaluated_expr = re.sub(pattern, replace_cell_ref, expr)
        
        # Replace Excel's ^ with Python's **
        evaluated_expr = evaluated_expr.replace('^', '**')
        
        try:
            # Evaluate the arithmetic expression
            # Use safe evaluation with limited builtins
            result = eval(evaluated_expr, {"__builtins__": {"abs": abs, "round": round}}, {})
            return float(result) if isinstance(result, (int, float)) else result
        except ZeroDivisionError:
            # Handle division by zero
            return 0
        except Exception as e:
            raise ValueError(f"Cannot evaluate arithmetic: {expr} -> {evaluated_expr}: {e}")
    
    def _evaluate_concatenation(
        self,
        expr: str,
        current_tab: str,
        current_row: int,
        current_col: int
    ) -> str:
        """
        Evaluate string concatenation expressions (Excel's & operator).
        
        Example: C$1&"*" where C$1 is 2024 -> "2024*"
        """
        # Split by & but preserve quoted strings
        parts = []
        current_part = []
        in_quotes = False
        
        for char in expr:
            if char == '"':
                in_quotes = not in_quotes
                current_part.append(char)
            elif char == '&' and not in_quotes:
                parts.append(''.join(current_part).strip())
                current_part = []
            else:
                current_part.append(char)
        
        if current_part:
            parts.append(''.join(current_part).strip())
        
        # Evaluate each part
        result_parts = []
        for part in parts:
            # Remove quotes if it's a string literal
            if part.startswith('"') and part.endswith('"'):
                result_parts.append(part[1:-1])
            else:
                # It's a cell reference or expression
                value = self._eval_expression(part, current_tab, current_row, current_col)
                # Format numbers without unnecessary decimal points
                # Excel's & operator converts 2024.0 to "2024", not "2024.0"
                if isinstance(value, float) and value == int(value):
                    result_parts.append(str(int(value)))
                else:
                    result_parts.append(str(value))
        
        return ''.join(result_parts)
    
    def _evaluate_function(
        self,
        func_expr: str,
        current_tab: str,
        current_row: int,
        current_col: int
    ) -> Any:
        """
        Evaluate Excel functions.
        
        Supported functions:
        - SUM, AVERAGE, MIN, MAX, COUNT
        - IF, IFERROR, IFNA
        - SUMIFS, COUNTIFS, AVERAGEIFS
        - INDEX, MATCH
        - AND, OR, NOT
        - ROUND, ABS, SQRT
        - LEFT, RIGHT, MID, LEN
        - VALUE, TEXT
        - COLUMNS, ROWS
        """
        # Parse function name and arguments
        match = re.match(r'([A-Z_]+)\((.*)\)$', func_expr.strip(), re.IGNORECASE)
        if not match:
            raise ValueError(f"Invalid function: {func_expr}")
        
        func_name = match.group(1).upper()
        args_str = match.group(2)
        
        # Parse arguments (respecting nested parentheses and commas)
        args = self._parse_function_args(args_str)
        
        # Dispatch to appropriate handler
        if func_name == 'SUM':
            return self._func_sum(args, current_tab, current_row, current_col)
        elif func_name == 'AVERAGE' or func_name == 'AVG':
            return self._func_average(args, current_tab, current_row, current_col)
        elif func_name == 'MIN':
            return self._func_min(args, current_tab, current_row, current_col)
        elif func_name == 'MAX':
            return self._func_max(args, current_tab, current_row, current_col)
        elif func_name == 'COUNT':
            return self._func_count(args, current_tab, current_row, current_col)
        elif func_name == 'IF':
            return self._func_if(args, current_tab, current_row, current_col)
        elif func_name == 'IFERROR':
            return self._func_iferror(args, current_tab, current_row, current_col)
        elif func_name == 'SUMIFS':
            return self._func_sumifs(args, current_tab, current_row, current_col)
        elif func_name == 'INDEX':
            return self._func_index(args, current_tab, current_row, current_col)
        elif func_name == 'MATCH':
            return self._func_match(args, current_tab, current_row, current_col)
        elif func_name == 'AND':
            return self._func_and(args, current_tab, current_row, current_col)
        elif func_name == 'OR':
            return self._func_or(args, current_tab, current_row, current_col)
        elif func_name == 'NOT':
            return self._func_not(args, current_tab, current_row, current_col)
        elif func_name == 'ABS':
            return self._func_abs(args, current_tab, current_row, current_col)
        elif func_name == 'MAX':
            return self._func_max(args, current_tab, current_row, current_col)
        elif func_name == 'MIN':
            return self._func_min(args, current_tab, current_row, current_col)
        elif func_name == 'ROUND':
            return self._func_round(args, current_tab, current_row, current_col)
        elif func_name == 'LEFT':
            return self._func_left(args, current_tab, current_row, current_col)
        elif func_name == 'VALUE':
            return self._func_value(args, current_tab, current_row, current_col)
        elif func_name == 'COLUMNS':
            return self._func_columns(args, current_tab, current_row, current_col)
        elif func_name == 'ROWS':
            return self._func_rows(args, current_tab, current_row, current_col)
        else:
            # Unsupported function - return placeholder
            return f"[{func_name}({args_str})]"
    
    def _parse_function_args(self, args_str: str) -> List[str]:
        """
        Parse function arguments, respecting nested parentheses and quoted strings.
        
        Example:
            "A1, SUM(B1:B5), 'Text, with comma'" 
            -> ["A1", "SUM(B1:B5)", "'Text, with comma'"]
        """
        args = []
        current_arg = []
        paren_depth = 0
        in_quotes = False
        quote_char = None
        
        for char in args_str:
            if char in ('"', "'") and (not in_quotes or char == quote_char):
                in_quotes = not in_quotes
                quote_char = char if in_quotes else None
                current_arg.append(char)
            elif char == '(' and not in_quotes:
                paren_depth += 1
                current_arg.append(char)
            elif char == ')' and not in_quotes:
                paren_depth -= 1
                current_arg.append(char)
            elif char == ',' and paren_depth == 0 and not in_quotes:
                # End of argument
                args.append(''.join(current_arg).strip())
                current_arg = []
            else:
                current_arg.append(char)
        
        # Add last argument
        if current_arg:
            args.append(''.join(current_arg).strip())
        
        return args
    
    def _parse_range(self, range_ref: str, current_tab: str) -> List[Tuple[str, int, int]]:
        """
        Parse a range reference into list of (tab, row, col) tuples.
        
        Example: "A1:A5" -> [(tab, 1, 1), (tab, 2, 1), ..., (tab, 5, 1)]
        Example: "D:D" -> all cells in column D
        """
        # Remove $ signs
        range_ref = range_ref.replace('$', '')
        
        # Parse tab
        if '!' in range_ref:
            parts = range_ref.split('!')
            tab_name = parts[0].strip("'")
            range_part = parts[1]
        else:
            tab_name = current_tab
            range_part = range_ref
        
        # Parse range
        if ':' in range_part:
            start_cell, end_cell = range_part.split(':')
            
            # Check for full column reference (e.g., "D:D")
            if re.match(r'^[A-Z]+$', start_cell) and re.match(r'^[A-Z]+$', end_cell):
                # Full column reference - get all non-empty cells in these columns
                start_col = column_index_from_string(start_cell)
                end_col = column_index_from_string(end_cell)
                
                cells = []
                if tab_name in self.computed_values:
                    # Scan all cells in this tab for matching columns
                    for cell_key, cell_value in self.computed_values[tab_name].items():
                        if cell_key.startswith('(') and cell_key.endswith(')'):
                            # Parse cell key like "(3, 4)"
                            try:
                                parts = cell_key[1:-1].split(',')
                                row = int(parts[0].strip())
                                col = int(parts[1].strip())
                                if start_col <= col <= end_col:
                                    cells.append((tab_name, row, col))
                            except:
                                continue
                return cells
            
            # Check for full row reference (e.g., "1:5")
            elif re.match(r'^\d+$', start_cell) and re.match(r'^\d+$', end_cell):
                # Full row reference
                start_row = int(start_cell)
                end_row = int(end_cell)
                
                cells = []
                if tab_name in self.computed_values:
                    for cell_key, cell_value in self.computed_values[tab_name].items():
                        if cell_key.startswith('(') and cell_key.endswith(')'):
                            try:
                                parts = cell_key[1:-1].split(',')
                                row = int(parts[0].strip())
                                col = int(parts[1].strip())
                                if start_row <= row <= end_row:
                                    cells.append((tab_name, row, col))
                            except:
                                continue
                return cells
            
            # Regular cell range (e.g., "A1:B5")
            match_start = re.match(r'([A-Z]+)(\d+)', start_cell)
            if not match_start:
                return []
            start_col = column_index_from_string(match_start.group(1))
            start_row = int(match_start.group(2))
            
            match_end = re.match(r'([A-Z]+)(\d+)', end_cell)
            if not match_end:
                return []
            end_col = column_index_from_string(match_end.group(1))
            end_row = int(match_end.group(2))
            
            # Generate all cells in range
            cells = []
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cells.append((tab_name, row, col))
            
            return cells
        else:
            # Single cell
            match = re.match(r'([A-Z]+)(\d+)', range_part)
            if not match:
                return []
            col = column_index_from_string(match.group(1))
            row = int(match.group(2))
            return [(tab_name, row, col)]
    
    def _get_range_values(self, range_ref: str, current_tab: str) -> List[Any]:
        """Get all values from a range"""
        cells = self._parse_range(range_ref, current_tab)
        values = []
        
        for tab, row, col in cells:
            key = f"({row}, {col})"
            if tab in self.computed_values and key in self.computed_values[tab]:
                val = self.computed_values[tab][key]
                if not isinstance(val, dict) or 'error' not in val:
                    values.append(val)
        
        return values
    
    # =============================================================================
    # Excel Function Implementations
    # =============================================================================
    
    def _func_sum(self, args: List[str], tab: str, row: int, col: int) -> float:
        """SUM function"""
        total = 0
        for arg in args:
            if ':' in arg:
                # Range
                values = self._get_range_values(arg, tab)
                total += sum(float(v) for v in values if isinstance(v, (int, float)))
            else:
                # Single value
                val = self._eval_expression(arg, tab, row, col)
                if isinstance(val, (int, float)):
                    total += val
        return total
    
    def _func_average(self, args: List[str], tab: str, row: int, col: int) -> float:
        """AVERAGE function"""
        values = []
        for arg in args:
            if ':' in arg:
                values.extend(self._get_range_values(arg, tab))
            else:
                val = self._eval_expression(arg, tab, row, col)
                values.append(val)
        
        numeric_values = [v for v in values if isinstance(v, (int, float))]
        return sum(numeric_values) / len(numeric_values) if numeric_values else 0
    
    def _func_min(self, args: List[str], tab: str, row: int, col: int) -> float:
        """MIN function"""
        values = []
        for arg in args:
            if ':' in arg:
                values.extend(self._get_range_values(arg, tab))
            else:
                val = self._eval_expression(arg, tab, row, col)
                values.append(val)
        
        numeric_values = [v for v in values if isinstance(v, (int, float))]
        return min(numeric_values) if numeric_values else 0
    
    def _func_max(self, args: List[str], tab: str, row: int, col: int) -> float:
        """MAX function"""
        values = []
        for arg in args:
            if ':' in arg:
                values.extend(self._get_range_values(arg, tab))
            else:
                val = self._eval_expression(arg, tab, row, col)
                values.append(val)
        
        numeric_values = [v for v in values if isinstance(v, (int, float))]
        return max(numeric_values) if numeric_values else 0
    
    def _func_count(self, args: List[str], tab: str, row: int, col: int) -> int:
        """COUNT function"""
        count = 0
        for arg in args:
            if ':' in arg:
                values = self._get_range_values(arg, tab)
                count += sum(1 for v in values if isinstance(v, (int, float)))
            else:
                val = self._eval_expression(arg, tab, row, col)
                if isinstance(val, (int, float)):
                    count += 1
        return count
    
    def _func_if(self, args: List[str], tab: str, row: int, col: int) -> Any:
        """IF function"""
        if len(args) < 2:
            return 0
        
        # Evaluate condition
        condition = self._eval_expression(args[0], tab, row, col)
        
        # Handle comparison operators
        if isinstance(condition, str):
            condition = self._evaluate_comparison(condition, tab, row, col)
        
        # Return appropriate branch
        if condition:
            return self._eval_expression(args[1], tab, row, col) if len(args) > 1 else True
        else:
            return self._eval_expression(args[2], tab, row, col) if len(args) > 2 else False
    
    def _evaluate_comparison(self, expr: str, tab: str, row: int, col: int) -> bool:
        """Evaluate comparison expressions like 'A1 > 5' or 'IFERROR(A1,"")<>""'"""
        # First, check if this is a complex expression with functions
        # We need to evaluate sub-expressions first
        
        # Check for comparison operators
        for op in ['<>', '>=', '<=', '>', '<', '=']:
            if op in expr:
                # Find the operator position (avoid matching inside function calls)
                # Split carefully to handle nested functions
                parts = self._split_by_operator(expr, op)
                
                if len(parts) == 2:
                    left = self._eval_expression(parts[0].strip(), tab, row, col)
                    right = self._eval_expression(parts[1].strip(), tab, row, col)
                    
                    # Perform comparison
                    if op == '>=':
                        return left >= right
                    elif op == '<=':
                        return left <= right
                    elif op == '<>':
                        return left != right
                    elif op == '>':
                        return left > right
                    elif op == '<':
                        return left < right
                    elif op == '=':
                        return left == right
        
        # If no comparison operator, treat as boolean
        return bool(expr)
    
    def _split_by_operator(self, expr: str, operator: str) -> List[str]:
        """
        Split expression by operator, respecting parentheses and quoted strings.
        
        Example: 'IFERROR(A1,"")<>""' with '<>' -> ['IFERROR(A1,"")', '""']
        """
        parts = []
        current_part = []
        paren_depth = 0
        in_quotes = False
        i = 0
        
        while i < len(expr):
            char = expr[i]
            
            if char == '"':
                in_quotes = not in_quotes
                current_part.append(char)
                i += 1
            elif char == '(' and not in_quotes:
                paren_depth += 1
                current_part.append(char)
                i += 1
            elif char == ')' and not in_quotes:
                paren_depth -= 1
                current_part.append(char)
                i += 1
            elif paren_depth == 0 and not in_quotes:
                # Check if we're at the operator
                if expr[i:i+len(operator)] == operator:
                    # Found the operator at top level
                    parts.append(''.join(current_part))
                    current_part = []
                    i += len(operator)
                    
                    # Add remaining as second part
                    parts.append(expr[i:])
                    return parts
                else:
                    current_part.append(char)
                    i += 1
            else:
                current_part.append(char)
                i += 1
        
        # If we didn't find the operator at top level, return the whole expression
        return [''.join(current_part)]
    
    def _func_iferror(self, args: List[str], tab: str, row: int, col: int) -> Any:
        """IFERROR function"""
        if len(args) < 1:
            return 0
        
        try:
            result = self._eval_expression(args[0], tab, row, col)
            # Check if result is an error
            if isinstance(result, dict) and 'error' in result:
                return self._eval_expression(args[1], tab, row, col) if len(args) > 1 else ""
            return result
        except:
            return self._eval_expression(args[1], tab, row, col) if len(args) > 1 else ""
    
    def _func_sumifs(self, args: List[str], tab: str, row: int, col: int) -> float:
        """
        SUMIFS function
        
        Syntax: SUMIFS(sum_range, criteria_range1, criteria1, ...)
        """
        if len(args) < 3:
            return 0
        
        # Get sum range
        sum_range = self._get_range_values(args[0], tab)
        sum_cells = self._parse_range(args[0], tab)
        
        # Process criteria pairs
        matching_indices = set(range(len(sum_range)))
        
        for i in range(1, len(args), 2):
            if i + 1 >= len(args):
                break
            
            criteria_range = self._parse_range(args[i], tab)
            
            # Evaluate the criteria (may contain expressions like C$1&"*")
            criteria_arg = args[i + 1]
            if criteria_arg.startswith('"') and criteria_arg.endswith('"'):
                # Already a string literal
                criteria_str = criteria_arg.strip('"')
            else:
                # May be an expression - evaluate it
                criteria_val = self._eval_expression(criteria_arg, tab, row, col)
                criteria_str = str(criteria_val)
            
            # Support wildcards (e.g., "2024*")
            if '*' in criteria_str:
                pattern = criteria_str.replace('*', '.*')
                for idx, (ctab, crow, ccol) in enumerate(criteria_range):
                    if idx not in matching_indices:
                        continue
                    
                    key = f"({crow}, {ccol})"
                    if ctab in self.computed_values and key in self.computed_values[ctab]:
                        val = str(self.computed_values[ctab][key])
                        if not re.match(pattern, val):
                            matching_indices.discard(idx)
            else:
                # Exact match
                for idx, (ctab, crow, ccol) in enumerate(criteria_range):
                    if idx not in matching_indices:
                        continue
                    
                    key = f"({crow}, {ccol})"
                    if ctab in self.computed_values and key in self.computed_values[ctab]:
                        val = str(self.computed_values[ctab][key])
                        if val != criteria_str:
                            matching_indices.discard(idx)
        
        # Sum matching values
        total = 0
        for idx in matching_indices:
            if idx < len(sum_range) and isinstance(sum_range[idx], (int, float)):
                total += sum_range[idx]
        
        return total
    
    def _func_index(self, args: List[str], tab: str, row: int, col: int) -> Any:
        """INDEX function"""
        if len(args) < 2:
            return 0
        
        try:
            # Get range
            range_values = self._get_range_values(args[0], tab)
            
            # Get row index (1-based)
            row_idx_expr = args[1]
            row_idx_val = self._eval_expression(row_idx_expr, tab, row, col)
            
            # Handle nested function results
            if isinstance(row_idx_val, (int, float)):
                row_idx = int(row_idx_val) - 1
            else:
                return 0
            
            if 0 <= row_idx < len(range_values):
                return range_values[row_idx]
            
            return 0
        except Exception as e:
            return 0
    
    def _func_match(self, args: List[str], tab: str, row: int, col: int) -> int:
        """MATCH function"""
        if len(args) < 2:
            return 0
        
        try:
            # Get lookup value
            lookup_val = self._eval_expression(args[0], tab, row, col)
            
            # Get lookup range
            range_values = self._get_range_values(args[1], tab)
            
            # Match type (0 = exact, 1 = less than or equal, -1 = greater than or equal)
            match_type = 0
            if len(args) > 2:
                match_type = int(self._eval_expression(args[2], tab, row, col))
            
            # Find match (1-based index)
            if match_type == 0:
                # Exact match
                try:
                    return range_values.index(lookup_val) + 1
                except ValueError:
                    # Try string comparison
                    lookup_str = str(lookup_val)
                    for i, val in enumerate(range_values):
                        if str(val) == lookup_str:
                            return i + 1
                    return 0
            else:
                # For other match types, return 0 (not implemented)
                return 0
        except Exception as e:
            return 0
    
    def _func_and(self, args: List[str], tab: str, row: int, col: int) -> bool:
        """AND function"""
        for arg in args:
            val = self._eval_expression(arg, tab, row, col)
            if not val:
                return False
        return True
    
    def _func_or(self, args: List[str], tab: str, row: int, col: int) -> bool:
        """OR function"""
        for arg in args:
            val = self._eval_expression(arg, tab, row, col)
            if val:
                return True
        return False
    
    def _func_not(self, args: List[str], tab: str, row: int, col: int) -> bool:
        """NOT function"""
        if len(args) < 1:
            return True
        val = self._eval_expression(args[0], tab, row, col)
        return not val
    
    def _func_abs(self, args: List[str], tab: str, row: int, col: int) -> float:
        """ABS function"""
        if len(args) < 1:
            return 0
        val = self._eval_expression(args[0], tab, row, col)
        return abs(val) if isinstance(val, (int, float)) else 0
    
    def _func_max(self, args: List[str], tab: str, row: int, col: int) -> float:
        """MAX function"""
        if len(args) < 1:
            return 0
        
        values = []
        for arg in args:
            if ':' in arg:
                # Range
                range_vals = self._get_range_values(arg, tab)
                values.extend([v for v in range_vals if isinstance(v, (int, float))])
            else:
                # Single value
                val = self._eval_expression(arg, tab, row, col)
                if isinstance(val, (int, float)):
                    values.append(val)
        
        return max(values) if values else 0
    
    def _func_min(self, args: List[str], tab: str, row: int, col: int) -> float:
        """MIN function"""
        if len(args) < 1:
            return 0
        
        values = []
        for arg in args:
            if ':' in arg:
                # Range
                range_vals = self._get_range_values(arg, tab)
                values.extend([v for v in range_vals if isinstance(v, (int, float))])
            else:
                # Single value
                val = self._eval_expression(arg, tab, row, col)
                if isinstance(val, (int, float)):
                    values.append(val)
        
        return min(values) if values else 0
    
    def _func_round(self, args: List[str], tab: str, row: int, col: int) -> float:
        """ROUND function"""
        if len(args) < 1:
            return 0
        val = self._eval_expression(args[0], tab, row, col)
        decimals = int(self._eval_expression(args[1], tab, row, col)) if len(args) > 1 else 0
        return round(val, decimals) if isinstance(val, (int, float)) else 0
    
    def _func_left(self, args: List[str], tab: str, row: int, col: int) -> str:
        """LEFT function"""
        if len(args) < 1:
            return ""
        
        try:
            text = str(self._eval_expression(args[0], tab, row, col))
            num_chars = int(self._eval_expression(args[1], tab, row, col)) if len(args) > 1 else 1
            return text[:num_chars]
        except Exception as e:
            return ""
    
    def _func_value(self, args: List[str], tab: str, row: int, col: int) -> float:
        """VALUE function"""
        if len(args) < 1:
            return 0
        
        try:
            text = str(self._eval_expression(args[0], tab, row, col))
            # Remove common formatting
            text = text.replace(',', '').replace('$', '').strip()
            return float(text)
        except Exception as e:
            return 0
    
    def _func_columns(self, args: List[str], tab: str, row: int, col: int) -> int:
        """COLUMNS function"""
        if len(args) < 1:
            return 0
        
        # Parse range
        cells = self._parse_range(args[0], tab)
        
        # Count unique columns
        unique_cols = set(col for _, _, col in cells)
        return len(unique_cols)
    
    def _func_rows(self, args: List[str], tab: str, row: int, col: int) -> int:
        """ROWS function"""
        if len(args) < 1:
            return 0
        
        # Parse range
        cells = self._parse_range(args[0], tab)
        
        # Count unique rows
        unique_rows = set(row for _, row, _ in cells)
        return len(unique_rows)
    
    def _serialize_value(self, value: Any) -> Any:
        """Convert value to JSON-serializable format"""
        if isinstance(value, (int, float, str, bool, type(None))):
            return value
        elif isinstance(value, (list, tuple)):
            return [self._serialize_value(v) for v in value]
        elif isinstance(value, dict):
            return {k: self._serialize_value(v) for k, v in value.items()}
        else:
            return str(value)
    
    def save_to_json(self, results: Dict[str, Any], output_path: Union[str, Path]) -> None:
        """
        Save evaluation results to JSON file.
        
        Args:
            results: The results from evaluate_all_tabs()
            output_path: Path to save the JSON file
        """
        output_path = Path(output_path)
        
        # Ensure parent directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Convert all values to JSON-serializable format
        serializable_results = {}
        for tab_name, tab_data in results.items():
            serializable_results[tab_name] = {
                "cells": {k: self._serialize_value(v) for k, v in tab_data["cells"].items()},
                "metadata": tab_data["metadata"]
            }
        
        # Write to file
        with open(output_path, 'w') as f:
            json.dump(serializable_results, f, indent=2)
        
        file_size = output_path.stat().st_size
        print(f"\n✅ Evaluation results saved to: {output_path}")
        print(f"   • File size: {file_size:,} bytes ({file_size/1024:.1f} KB)")
        print(f"   • Tabs evaluated: {len(serializable_results)}")
        
        total_cells = sum(tab["metadata"]["total_cells"] for tab in serializable_results.values())
        print(f"   • Total cells: {total_cells}")
