#!/usr/bin/env python3
"""Price Adjustor - News-Based Financial Model Adjustment

This module builds a BRAND NEW financial model from scratch using adjusted assumptions.
It does NOT modify the original Excel file - instead it creates a new one.

Architecture:
1. Load original model to extract base assumptions from LLM_Inferred tab
2. Load screening data (news analysis)
3. Call LLM to get adjusted assumptions based on news
4. Build completely new Excel model using existing tab builders
5. All tabs automatically reference the adjusted assumptions

Key insight: We reuse all existing tab builders WITHOUT modification by simply
passing adjusted assumptions to AssumptionsTabBuilder, which creates the LLM_Inferred
tab that all other tabs reference.
"""

from typing import Dict, Any, List, Optional
from pathlib import Path
from datetime import datetime
import json
import openpyxl
from llms.config import get_llm
from logger import StockAnalystLogger


class PriceAdjustor:
    """
    Builds a new adjusted financial model based on news analysis.
    
    Workflow:
    1. Load original model + screening data
    2. Extract base assumptions from original model
    3. Call LLM to infer adjustments based on news
    4. Build brand new model from scratch with adjusted assumptions
    5. Save as new Excel file (original untouched)
    """
    
    def __init__(self, ticker: str, model_path, screening_path, logger: Optional[StockAnalystLogger] = None):
        self.ticker = ticker.upper()
        self.model_path = Path(model_path)
        self.screening_path = Path(screening_path)
        self.logger = logger
        
        if not self.model_path.exists():
            raise FileNotFoundError(f"Model not found: {self.model_path}")
        if not self.screening_path.exists():
            raise FileNotFoundError(f"Screening data not found: {self.screening_path}")
        
        # Original model (read-only, for extracting base assumptions)
        self.original_workbook = None
        
        # New model (built from scratch)
        self.new_workbook = None
        
        # Data
        self.screening_data = None
        self.base_assumptions = None
        self.adjusted_assumptions = None
        self.adjustment_metadata = None
        self.llm_cost = 0.0
        
        # JSON data from original model (needed for Raw/Keys_Map tabs)
        self.json_data = None
    
    def load_data(self):
        """Load original model and screening data."""
        if self.logger:
            self.logger.stage_start("LOADING DATA FOR PRICE ADJUSTMENT", f"Loading model and screening data - {self.ticker}")
        
        if self.logger:
            self.logger.info(f"[1/3] Loading original model from {self.model_path.name}...")
        self.original_workbook = openpyxl.load_workbook(self.model_path)
        if self.logger:
            self.logger.info(f"      ✅ Loaded {len(self.original_workbook.sheetnames)} tabs")
        
        if self.logger:
            self.logger.info(f"[2/3] Extracting Raw data from original model...")
        # Extract JSON-like data from Raw tab (needed to rebuild model)
        self.json_data = self._extract_json_from_raw_tab()
        if self.logger:
            self.logger.info(f"      ✅ Extracted {len(self.json_data.get('data', []))} data points")
        
        if self.logger:
            self.logger.info(f"[3/3] Loading screening data from {self.screening_path.name}...")
        with open(self.screening_path, 'r') as f:
            self.screening_data = json.load(f)
        
        summary = self.screening_data.get('analysis_summary', {})
        if self.logger:
            self.logger.info(f"      ✅ Loaded news analysis:")
            self.logger.info(f"         • Sentiment: {summary.get('overall_sentiment', 'N/A')}")
            self.logger.info(f"         • Catalysts: {summary.get('total_catalysts', 0)}")
            self.logger.info(f"         • Risks: {summary.get('total_risks', 0)}")
            self.logger.stage_end("LOADING DATA", success=True)
    
    def _extract_json_from_raw_tab(self) -> Dict:
        """Extract data from Raw tab to reconstruct JSON format."""
        if "Raw" not in self.original_workbook.sheetnames:
            raise ValueError("Raw tab not found in original model")
        
        ws = self.original_workbook["Raw"]
        
        # Raw tab structure: Column A=Statement, B=Field, C=Year, D=Value
        # Skip header row (row 1)
        data = []
        for row in range(2, ws.max_row + 1):
            statement = ws.cell(row, 1).value
            field = ws.cell(row, 2).value
            year = ws.cell(row, 3).value
            value = ws.cell(row, 4).value
            
            if statement and field and year:  # Skip empty rows
                data.append({
                    "statement": statement,
                    "field": field,
                    "year": str(year),
                    "value": value if value is not None else 0
                })
        
        return {
            "ticker": self.ticker,
            "data": data
        }
    
    def extract_base_assumptions(self):
        """Extract base assumptions from original model's LLM_Inferred tab."""
        if self.logger:
            self.logger.stage_start("EXTRACTING BASE ASSUMPTIONS", "Reading assumptions from original model")
        
        if "LLM_Inferred" not in self.original_workbook.sheetnames:
            raise ValueError("LLM_Inferred tab not found")
        
        ws = self.original_workbook["LLM_Inferred"]
        
        self.base_assumptions = {
            'wacc': ws['B2'].value or 0.09,
            'terminal_growth_rate': ws['B3'].value or 0.025,
            'revenue_growth_rates': [ws[f'{c}4'].value or 0.05 for c in 'BCDEF'],
            'gross_margins': [ws[f'{c}5'].value or 0.46 for c in 'BCDEF'],
            'ebitda_margins': [ws[f'{c}6'].value or 0.33 for c in 'BCDEF'],
            'operating_margins': [ws[f'{c}7'].value or 0.31 for c in 'BCDEF'],
            'dso_days': [ws[f'{c}8'].value or 45 for c in 'BCDEF'],
            'dio_days': [ws[f'{c}9'].value or 10 for c in 'BCDEF'],
            'dpo_days': [ws[f'{c}10'].value or 90 for c in 'BCDEF']
        }
        
        if self.logger:
            self.logger.info(f"✅ Base assumptions extracted:")
            self.logger.info(f"   • WACC: {self.base_assumptions['wacc']*100:.2f}%")
            self.logger.info(f"   • Revenue Growth FY1: {self.base_assumptions['revenue_growth_rates'][0]*100:.1f}%")
            self.logger.stage_end("EXTRACTING ASSUMPTIONS", success=True)
    
    def call_llm_for_adjustments(self):
        """Call LLM to infer adjustments based on news analysis."""
        if self.logger:
            self.logger.stage_start("LLM ADJUSTMENT INFERENCE", "Calling LLM to infer parameter adjustments")
        
        # Check if there's any news to analyze
        num_catalysts = len(self.screening_data.get('catalysts', []))
        num_risks = len(self.screening_data.get('risks', []))
        num_mitigations = len(self.screening_data.get('mitigations', []))
        
        if num_catalysts == 0 and num_risks == 0:
            if self.logger:
                self.logger.warning("⚠️  No catalysts or risks identified in news analysis")
                self.logger.info("📊 Using base assumptions UNCHANGED (no news adjustments needed)")
            
            # Return base assumptions as-is - no changes
            self.adjusted_assumptions = self.base_assumptions.copy()
            self.adjustment_metadata = {
                'adjustments_summary': {
                    'revenue_growth_fy1_bps': 0,
                    'revenue_growth_fy2_bps': 0,
                    'revenue_growth_fy3_bps': 0,
                    'revenue_growth_fy4_bps': 0,
                    'revenue_growth_fy5_bps': 0,
                    'gross_margin_fy1_bps': 0,
                    'gross_margin_fy2_bps': 0,
                    'gross_margin_fy3_bps': 0,
                    'gross_margin_fy4_bps': 0,
                    'gross_margin_fy5_bps': 0,
                    'operating_margin_fy1_bps': 0,
                    'operating_margin_fy2_bps': 0,
                    'operating_margin_fy3_bps': 0,
                    'operating_margin_fy4_bps': 0,
                    'operating_margin_fy5_bps': 0,
                },
                'factor_mappings': [],
                'note': 'No catalysts or risks identified - base case assumptions used unchanged'
            }
            
            if self.logger:
                self.logger.info("✅ Adjusted model will be IDENTICAL to base case (0 bps change on all parameters)")
                self.logger.stage_end("LLM ADJUSTMENT INFERENCE", success=True, stats={
                    "LLM Cost": "$0.0000 (skipped - no news)",
                    "Adjustments": "All 0 bps",
                    "News Factors": f"{num_catalysts} catalysts, {num_risks} risks"
                })
            return
        
        # Proceed with LLM call if there are catalysts or risks
        if self.logger:
            self.logger.info(f"📰 News factors found: {num_catalysts} catalysts, {num_risks} risks, {num_mitigations} mitigations")
            self.logger.info(f"🤖 Calling LLM to infer parameter adjustments based on news...")
        
        prompt_path = Path(__file__).parent.parent / "prompts" / "news_to_adjustments.md"
        with open(prompt_path, 'r') as f:
            prompt_template = f.read()
        
        # Use base assumptions for FY0
        fy0 = {
            'revenue_growth_fy0': f"{self.base_assumptions['revenue_growth_rates'][0]*100:.1f}",
            'gross_margin_fy0': f"{self.base_assumptions['gross_margins'][0]*100:.1f}",
            'operating_margin_fy0': f"{self.base_assumptions['operating_margins'][0]*100:.1f}"
        }
        
        # Format base assumptions
        base_fmt = {
            'base_revenue_growth': ', '.join([f"{r*100:.1f}%" for r in self.base_assumptions['revenue_growth_rates']]),
            'base_gross_margins': ', '.join([f"{m*100:.1f}%" for m in self.base_assumptions['gross_margins']]),
            'base_operating_margins': ', '.join([f"{m*100:.1f}%" for m in self.base_assumptions['operating_margins']]),
            'base_wacc': f"{self.base_assumptions['wacc']*100:.2f}",
            'base_terminal_growth': f"{self.base_assumptions['terminal_growth_rate']*100:.2f}"
        }
        
        # Format news
        summary = self.screening_data.get('analysis_summary', {})
        news_summary = f"Sentiment: {summary.get('overall_sentiment', 'neutral')} | Confidence: {summary.get('confidence_score', 0.5):.2f}"
        
        catalysts = self._format_factors(self.screening_data.get('catalysts', []), "catalyst")
        risks = self._format_factors(self.screening_data.get('risks', []), "risk")
        mitigations = self._format_factors(self.screening_data.get('mitigations', []), "mitigation")
        
        # Fill prompt
        prompt = prompt_template.format(
            company_name=self.ticker,
            ticker=self.ticker,
            sector="Technology",
            model_date=datetime.now().strftime("%Y-%m-%d"),
            **fy0,
            **base_fmt,
            news_summary=news_summary,
            catalysts_detail=catalysts,
            risks_detail=risks,
            mitigations_detail=mitigations
        )
        
        # Call LLM
        llm = get_llm()
        messages = [{"role": "user", "content": prompt}]
        
        if self.logger:
            self.logger.info("🤖 Sending request to LLM...")
        response, cost = llm(messages, temperature=0.3)
        self.llm_cost += cost
        if self.logger:
            self.logger.llm_call("Price Adjustment Parameter Inference", cost)
        
        # Parse response
        try:
            response_clean = response.strip()
            if response_clean.startswith('```'):
                lines = response_clean.split('\n')
                response_clean = '\n'.join(lines[1:-1]) if len(lines) > 2 else response_clean
            
            result = json.loads(response_clean)
            self.adjusted_assumptions = result.get('adjusted_assumptions', {})
            self.adjustment_metadata = {
                'adjustments_summary': result.get('adjustments_summary', {}),
                'factor_mappings': result.get('factor_mappings', [])
            }
            
            if self.logger:
                self.logger.info(f"✅ LLM adjustment successful!")
                
                adj_summary = self.adjustment_metadata['adjustments_summary']
                self.logger.info(f"📊 Key Adjustments:")
                for key in ['revenue_growth_fy1_bps', 'gross_margin_fy1_bps', 'operating_margin_fy1_bps']:
                    if key in adj_summary:
                        self.logger.info(f"   • {key}: {adj_summary[key]:+d} bps")
                
                # Log ALL parameter changes for transparency
                self.logger.info(f"\n📋 Complete Adjustment Details:")
                self.logger.info(f"   Base Assumptions:")
                self.logger.info(f"   • Revenue Growth Rates (FY1-FY5): {', '.join([f'{r*100:.1f}%' for r in self.base_assumptions['revenue_growth_rates']])}")
                self.logger.info(f"   • Gross Margins (FY1-FY5): {', '.join([f'{m*100:.1f}%' for m in self.base_assumptions['gross_margins']])}")
                self.logger.info(f"   • Operating Margins (FY1-FY5): {', '.join([f'{m*100:.1f}%' for m in self.base_assumptions['operating_margins']])}")
                self.logger.info(f"   • EBITDA Margins (FY1-FY5): {', '.join([f'{m*100:.1f}%' for m in self.base_assumptions['ebitda_margins']])}")
                self.logger.info(f"   • WACC: {self.base_assumptions['wacc']*100:.2f}%")
                self.logger.info(f"   • Terminal Growth: {self.base_assumptions['terminal_growth_rate']*100:.2f}%")
                
                # Show adjusted values
                self.logger.info(f"\n   Adjusted Parameters (from LLM):")
                for param, value in self.adjusted_assumptions.items():
                    if isinstance(value, (int, float)):
                        # Handle both percentages and basis points
                        if 'bps' in param or 'basis' in param.lower():
                            self.logger.info(f"   • {param}: {value:+.0f} bps")
                        elif param in ['wacc', 'terminal_growth_rate'] or 'rate' in param or 'margin' in param or 'growth' in param:
                            self.logger.info(f"   • {param}: {value*100:.2f}%")
                        else:
                            self.logger.info(f"   • {param}: {value}")
                    elif isinstance(value, list):
                        if all(isinstance(v, (int, float)) for v in value):
                            self.logger.info(f"   • {param}: {', '.join([f'{v*100:.1f}%' if v < 10 else f'{v:.0f}' for v in value])}")
                        else:
                            self.logger.info(f"   • {param}: {value}")
                    else:
                        self.logger.info(f"   • {param}: {value}")
                
                # Show news context
                num_catalysts = len(self.screening_data.get('catalysts', []))
                num_risks = len(self.screening_data.get('risks', []))
                num_mitigations = len(self.screening_data.get('mitigations', []))
                sentiment = self.screening_data.get('analysis_summary', {}).get('overall_sentiment', 'neutral')
                
                self.logger.info(f"\n   News Analysis Context:")
                self.logger.info(f"   • Sentiment: {sentiment.upper()}")
                self.logger.info(f"   • Catalysts: {num_catalysts}")
                self.logger.info(f"   • Risks: {num_risks}")
                self.logger.info(f"   • Mitigations: {num_mitigations}")
                
                if num_catalysts == 0 and num_risks == 0:
                    self.logger.warning(f"   ⚠️  WARNING: No catalysts or risks identified - adjustments may be minimal or default!")
                
                self.logger.stage_end("LLM ADJUSTMENT INFERENCE", success=True, stats={
                    "LLM Cost": f"${cost:.4f}",
                    "Adjustments": len(adj_summary),
                    "News Factors": f"{num_catalysts} catalysts, {num_risks} risks"
                })
        
        except json.JSONDecodeError as e:
            if self.logger:
                self.logger.error(f"❌ Failed to parse LLM response: {e}")
                self.logger.error(f"Response: {response[:500]}")
                self.logger.stage_end("LLM ADJUSTMENT INFERENCE", success=False)
            raise
    
    def _format_factors(self, factors, factor_type):
        """Format news factors for LLM prompt."""
        if not factors:
            return f"No {factor_type}s."
        
        lines = [f"\n## {factor_type.upper()}S ({len(factors)} total)"]
        for i, f in enumerate(factors, 1):
            desc = f.get('description', 'N/A')[:80]
            conf = f.get('confidence', 0.5)
            
            extra = []
            if factor_type == "catalyst":
                extra.append(f"timeline={f.get('timeline', 'N/A')}")
            elif factor_type == "risk":
                extra.append(f"sev={f.get('severity', 'med')}, like={f.get('likelihood', 'med')}")
            elif factor_type == "mitigation":
                extra.append(f"eff={f.get('effectiveness', 'med')}")
            
            lines.append(f"{i}. [{f.get('type', 'unknown')}] {desc} (conf={conf:.2f}, {', '.join(extra)})")
        
        return '\n'.join(lines)
    
    def build_new_model(self):
        """Build brand new Excel model from scratch using adjusted assumptions."""
        if self.logger:
            self.logger.stage_start("BUILDING ADJUSTED MODEL", "Building new Excel model from scratch with adjusted assumptions")
        
        # Import all tab builders
        from agents.fm.tabs.tab_raw import RawTabBuilder
        from agents.fm.tabs.tab_keys_map import KeysMapTabBuilder
        from agents.fm.tabs.tab_assumptions import AssumptionsTabBuilder
        from agents.fm.tabs.tab_historical import HistoricalTabBuilder
        from agents.fm.tabs.tab_projections import ProjectionsTabBuilder
        from agents.fm.tabs.tab_valuation_perpetual_growth_dcf import ValuationPerpetualGrowthDCFBuilder
        from agents.fm.tabs.tab_valuation_exit_multiple_dcf import ValuationExitMultipleDCFBuilder
        from agents.fm.tabs.tab_sensitivity import SensitivityTabBuilder
        from agents.fm.tabs.tab_summary import SummaryTabBuilder
        
        # Create new workbook
        self.new_workbook = openpyxl.Workbook()
        if 'Sheet' in self.new_workbook.sheetnames:
            self.new_workbook.remove(self.new_workbook['Sheet'])
        
        # Initialize builders
        raw_builder = RawTabBuilder()
        keys_map_builder = KeysMapTabBuilder()
        assumptions_builder = AssumptionsTabBuilder(llm_assumptions=self.adjusted_assumptions)
        historical_builder = HistoricalTabBuilder()
        projections_builder = ProjectionsTabBuilder()
        perpetual_growth_dcf_builder = ValuationPerpetualGrowthDCFBuilder()
        exit_multiple_dcf_builder = ValuationExitMultipleDCFBuilder()
        sensitivity_builder = SensitivityTabBuilder()
        summary_builder = SummaryTabBuilder()
        
        # Build all tabs in sequence (same order as original FinancialModelBuilder)
        if self.logger:
            self.logger.info("[1/9] Building Raw tab...")
        raw_builder.data_rows = self._convert_json_to_raw_rows()
        raw_builder.create_tab(self.new_workbook)
        if self.logger:
            self.logger.info(f"      ✅ Raw tab created ({len(raw_builder.data_rows)} rows)")
        
        if self.logger:
            self.logger.info("[2/9] Building Keys_Map tab...")
        keys_map_builder.build_from_raw_data(raw_builder.data_rows)
        keys_map_builder.create_tab(self.new_workbook)
        if self.logger:
            self.logger.info(f"      ✅ Keys_Map tab created ({len(keys_map_builder.field_mappings)} fields)")
        
        if self.logger:
            self.logger.info("[3/9] Building Assumptions tab (with ADJUSTED values)...")
        assumptions_builder.create_tab(self.new_workbook)
        if self.logger:
            self.logger.info("      ✅ Assumptions tab created (with LLM_Inferred containing adjusted values)")
        
        if self.logger:
            self.logger.info("[4/9] Building Historical tab...")
        historical_builder.create_tab(self.new_workbook)
        if self.logger:
            self.logger.info("      ✅ Historical tab created (5 years actuals)")
        
        if self.logger:
            self.logger.info("[5/9] Building Projections tab...")
        projections_builder.create_tab(self.new_workbook)
        if self.logger:
            self.logger.info("      ✅ Projections tab created (FY1-FY5 forecasts with adjusted assumptions)")
        
        if self.logger:
            self.logger.info("[6/9] Building Valuation (Perpetual Growth DCF) tab...")
        perpetual_growth_dcf_builder.create_tab(self.new_workbook)
        if self.logger:
            self.logger.info("      ✅ Valuation (DCF) tab created")
        
        if self.logger:
            self.logger.info("[7/9] Building Valuation (Exit Multiple DCF) tab...")
        exit_multiple_dcf_builder.create_tab(self.new_workbook)
        if self.logger:
            self.logger.info("      ✅ Valuation (Exit Multiple) tab created")
        
        if self.logger:
            self.logger.info("[8/9] Building Sensitivity tab...")
        sensitivity_builder.create_tab(self.new_workbook)
        if self.logger:
            self.logger.info("      ✅ Sensitivity tab created (2-way analysis)")
        
        if self.logger:
            self.logger.info("[9/9] Building Summary tab...")
        summary_builder.create_tab(self.new_workbook)
        if self.logger:
            self.logger.info("      ✅ Summary tab created (34 metrics)")
        
        # Set Summary as active sheet
        self.new_workbook.active = self.new_workbook["Summary"]
        
        if self.logger:
            self.logger.stage_end("BUILDING ADJUSTED MODEL", success=True, stats={
                "Total Tabs": len(self.new_workbook.sheetnames),
                "Active Sheet": "Summary"
            })
    
    def _convert_json_to_raw_rows(self):
        """Convert JSON data to Raw tab row format."""
        rows = []
        for item in self.json_data.get('data', []):
            rows.append({
                'statement': item['statement'],
                'field': item['field'],
                'year': item['year'],
                'value': item['value']
            })
        return rows
    
    def save(self, output_path=None):
        """Save the new adjusted model to Excel file."""
        if output_path is None:
            output_path = self.model_path.parent / f"{self.model_path.stem}_adjusted{self.model_path.suffix}"
        else:
            output_path = Path(output_path)
        
        if self.logger:
            self.logger.stage_start("SAVING ADJUSTED MODEL", "Saving new Excel file to disk")
        
        self.new_workbook.save(output_path)
        
        file_size = output_path.stat().st_size
        
        if self.logger:
            self.logger.file_operation("Adjusted model saved", output_path)
            self.logger.stage_end("SAVING ADJUSTED MODEL", success=True, stats={
                "File Size": f"{file_size:,} bytes",
                "Total Tabs": len(self.new_workbook.sheetnames),
                "Total LLM Cost": f"${self.llm_cost:.4f}",
                "Original Model": "UNTOUCHED"
            })
        
        return output_path
    
    def run(self):
        """Execute the full adjustment workflow."""
        self.load_data()
        self.extract_base_assumptions()
        self.call_llm_for_adjustments()
        self.build_new_model()


def adjust_price(ticker, model_path, screening_path, output_path=None, logger: Optional[StockAnalystLogger] = None):
    """
    Convenience function to adjust a financial model based on news analysis.
    
    Creates a brand new Excel model with adjusted assumptions.
    Original model is NOT modified.
    
    Args:
        ticker: Stock ticker symbol (e.g., "AAPL")
        model_path: Path to original financial model Excel file
        screening_path: Path to screening_data.json (news analysis)
        output_path: Optional path for new adjusted model (default: {original}_adjusted.xlsx)
        logger: Optional logger instance
    
    Returns:
        Path to the new adjusted model
    
    Example:
        >>> from src.price_adjustor import adjust_price
        >>> adjust_price("AAPL", "AAPL_financial_model.xlsx", "screening_data.json", logger=my_logger)
        # Creates: AAPL_financial_model_adjusted.xlsx (brand new file)
    """
    adjustor = PriceAdjustor(ticker, model_path, screening_path, logger)
    adjustor.run()
    return adjustor.save(output_path)
